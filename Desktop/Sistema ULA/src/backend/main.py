import os
import io
import re
import csv
import unicodedata
import pymysql
import bcrypt
import certifi
import pdfplumber
import json
import smtplib
import secrets
from contextlib import asynccontextmanager
from collections import defaultdict
from datetime import datetime, timedelta, date, time
from typing import Optional, List
import uuid
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Response, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from dotenv import load_dotenv
from io import BytesIO

# Cargar variables de entorno desde src/backend/.env
base_dir = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(base_dir, ".env"))

def migrar_columnas_verificacion():
    """Agrega las columnas y tablas necesarias si no existen."""
    try:
        connection = get_db_connection()
        with connection.cursor() as cursor:
            for sentencia in [
                "ALTER TABLE usuarios ADD COLUMN is_verified BOOLEAN NOT NULL DEFAULT FALSE",
                "ALTER TABLE usuarios ADD COLUMN verification_code VARCHAR(6) NULL",
                "ALTER TABLE usuarios ADD COLUMN reset_code VARCHAR(6) NULL",
                "ALTER TABLE aulas ADD COLUMN en_mantenimiento BOOLEAN NOT NULL DEFAULT FALSE",
                "ALTER TABLE aulas ADD COLUMN inicio_mantenimiento DATETIME NULL",
                "ALTER TABLE aulas ADD COLUMN fin_mantenimiento DATETIME NULL",
                "ALTER TABLE aulas ADD COLUMN aula_temporal VARCHAR(100) NULL",
                "ALTER TABLE usuarios ADD COLUMN verification_code_expira DATETIME NULL",
                "ALTER TABLE usuarios ADD COLUMN reset_code_expira DATETIME NULL",
                "ALTER TABLE usuarios ADD COLUMN created_at DATETIME DEFAULT NOW()",
                "ALTER TABLE horarios ADD COLUMN fecha_clase DATE NULL",
                "ALTER TABLE horarios ADD COLUMN semestre VARCHAR(50) NULL",
                "ALTER TABLE horarios ADD COLUMN cuatrimestre VARCHAR(50) NULL",
                "ALTER TABLE horarios ADD COLUMN grupo VARCHAR(50) NULL",
                "ALTER TABLE horarios ADD COLUMN estado_slug VARCHAR(50) DEFAULT 'programada'",
                "ALTER TABLE horarios ADD COLUMN nota_reprogramacion VARCHAR(255) DEFAULT ''",
                "ALTER TABLE horarios ADD COLUMN fecha_reposicion DATE NULL",
                "ALTER TABLE horarios ADD COLUMN es_reposicion BOOLEAN DEFAULT FALSE",
                "ALTER TABLE examenes_calendario ADD COLUMN archivo_origen VARCHAR(255) NULL",
                """CREATE TABLE IF NOT EXISTS docentes (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    nombre VARCHAR(200) NOT NULL,
                    especialidad VARCHAR(200) DEFAULT '',
                    materias TEXT DEFAULT '[]',
                    correo VARCHAR(200) DEFAULT '',
                    created_at DATETIME DEFAULT NOW()
                )""",
                """CREATE TABLE IF NOT EXISTS suplencias (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    docente_id INT NOT NULL,
                    suplente_id INT NOT NULL,
                    materia VARCHAR(200) DEFAULT '',
                    dia VARCHAR(50) DEFAULT '',
                    fecha DATE,
                    hora_inicio TIME,
                    hora_fin TIME,
                    activa BOOLEAN DEFAULT TRUE,
                    created_at DATETIME DEFAULT NOW()
                )""",
                """CREATE TABLE IF NOT EXISTS suplencias_horarios (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    docente_nombre VARCHAR(200) NOT NULL,
                    suplente_nombre VARCHAR(200) NOT NULL,
                    materia VARCHAR(200) DEFAULT '',
                    dia VARCHAR(50) DEFAULT '',
                    fecha DATE,
                    hora_inicio TIME,
                    hora_fin TIME,
                    activa BOOLEAN DEFAULT TRUE,
                    created_at DATETIME DEFAULT NOW()
                )""",
                """CREATE TABLE IF NOT EXISTS calendarios (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    tipo VARCHAR(50) NOT NULL,
                    carrera VARCHAR(50) DEFAULT '',
                    archivo_nombre VARCHAR(255) NOT NULL,
                    archivo_url VARCHAR(500) NOT NULL,
                    created_at DATETIME DEFAULT NOW()
                )""",
                "ALTER TABLE calendarios ADD COLUMN archivo_datos LONGBLOB NULL",
                "ALTER TABLE calendarios ADD COLUMN ciclo_escolar VARCHAR(20) DEFAULT '2025-2026'",
                """CREATE TABLE IF NOT EXISTS examenes_calendario (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    carrera VARCHAR(50) NOT NULL,
                    periodo VARCHAR(100) NOT NULL,
                    semestre VARCHAR(100) NOT NULL,
                    dia VARCHAR(30) DEFAULT '',
                    fecha VARCHAR(100) DEFAULT '',
                    materia VARCHAR(250) NOT NULL,
                    created_at DATETIME DEFAULT NOW()
                )""",
                """CREATE TABLE IF NOT EXISTS calendario_institucional (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    plan VARCHAR(50) NOT NULL,
                    ciclo VARCHAR(50) NOT NULL,
                    periodo INT NOT NULL,
                    tipo_evento VARCHAR(100) NOT NULL,
                    descripcion VARCHAR(300) NOT NULL,
                    fecha_inicio DATE NOT NULL,
                    fecha_fin DATE NOT NULL,
                    suspende_clases BOOLEAN DEFAULT FALSE,
                    created_at DATETIME DEFAULT NOW()
                )""",
                """CREATE TABLE IF NOT EXISTS aulas_liberadas (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    aula_nombre VARCHAR(100) NOT NULL,
                    fecha DATE NOT NULL,
                    docente VARCHAR(200) DEFAULT '',
                    asignatura VARCHAR(250) DEFAULT '',
                    horario VARCHAR(100) DEFAULT '',
                    motivo VARCHAR(250) DEFAULT 'Liberación manual',
                    activa BOOLEAN DEFAULT TRUE,
                    created_at DATETIME DEFAULT NOW()
                )""",
                """CREATE TABLE IF NOT EXISTS auditoria_exportaciones (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    usuario_nombre VARCHAR(255) DEFAULT 'Administrador',
                    usuario_correo VARCHAR(255) DEFAULT '',
                    fecha_hora DATETIME DEFAULT NOW(),
                    tipo_exportacion VARCHAR(50) NOT NULL,
                    es_historial_completo BOOLEAN DEFAULT FALSE,
                    filtros_aplicados TEXT NULL,
                    cantidad_registros INT DEFAULT 0,
                    enviado_email BOOLEAN DEFAULT FALSE,
                    destinatarios_email TEXT NULL
                )""",
                """CREATE TABLE IF NOT EXISTS reportes_compartidos (
                    id VARCHAR(64) PRIMARY KEY,
                    fecha_hora DATETIME DEFAULT NOW(),
                    nombre_archivo VARCHAR(255) NOT NULL,
                    tipo VARCHAR(50) NOT NULL,
                    contenido LONGBLOB NOT NULL
                )""",
                """CREATE TABLE IF NOT EXISTS clases_historico (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    clase_id INT NULL,
                    fecha DATE NOT NULL,
                    dia_semana VARCHAR(20) NOT NULL,
                    semana_numero INT NOT NULL,
                    docente_nombre VARCHAR(255) NOT NULL,
                    licenciatura_codigo VARCHAR(100) DEFAULT '',
                    semestre VARCHAR(100) DEFAULT '',
                    grupo VARCHAR(50) DEFAULT '',
                    asignatura_nombre VARCHAR(255) NOT NULL,
                    horario_inicio TIME NOT NULL,
                    horario_fin TIME NOT NULL,
                    aula_original VARCHAR(100) DEFAULT '',
                    aula_actual VARCHAR(100) DEFAULT '',
                    estado_slug VARCHAR(100) NOT NULL,
                    estado_label VARCHAR(255) NOT NULL,
                    incidencia_detalle TEXT NULL,
                    created_at DATETIME DEFAULT NOW(),
                    INDEX idx_fecha (fecha),
                    INDEX idx_semana (semana_numero),
                    INDEX idx_licenciatura (licenciatura_codigo),
                    INDEX idx_docente (docente_nombre),
                    INDEX idx_semestre (semestre)
                )""",
                """CREATE TABLE IF NOT EXISTS historial_reprogramaciones_semanal (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    docente VARCHAR(255) NOT NULL,
                    licenciatura VARCHAR(255),
                    asignatura VARCHAR(255),
                    horario_reposicion VARCHAR(255),
                    fecha_reposicion DATE,
                    aula_asignada VARCHAR(100),
                    nota_reprogramacion TEXT,
                    motivo TEXT,
                    fecha_archivado DATETIME DEFAULT CURRENT_TIMESTAMP,
                    semana_anio VARCHAR(20),
                    INDEX idx_fecha_rep (fecha_reposicion),
                    INDEX idx_docente_rep (docente)
                )""",
            ]:
                try:
                    cursor.execute(sentencia)
                except Exception:
                    pass
        connection.commit()
        connection.close()
    except Exception as e:
        print(f"Advertencia en migración: {e}")


def _limpiar_para_comparar(texto: str) -> str:
    if not texto:
        return ""
    texto_nfd = unicodedata.normalize('NFD', texto)
    texto_sin_acento = "".join(c for c in texto_nfd if unicodedata.category(c) != 'Mn')
    return texto_sin_acento.lower().strip()


def _normalizar_nombre_formato(nombre: str) -> str:
    if not nombre:
        return ""
    nombre = nombre.strip()
    if ',' in nombre:
        parts = nombre.split(',', 1)
        nombre = f"{parts[1].strip()} {parts[0].strip()}"
    words = [w.capitalize() for w in nombre.split()]
    return " ".join(words)


def _construir_mapa_docentes_canonicos(lista_nombres_raw: list) -> dict:
    mapa_raw_a_candidato = {}
    candidatos_unicos = set()
    for raw in lista_nombres_raw:
        if not raw or not raw.strip():
            continue
        cand = _normalizar_nombre_formato(raw)
        mapa_raw_a_candidato[raw] = cand
        candidatos_unicos.add(cand)

    candidatos_ordenados = sorted(list(candidatos_unicos), key=lambda x: len(x), reverse=True)
    candidato_a_canonico = {}
    
    for cand in candidatos_ordenados:
        clean_cand = _limpiar_para_comparar(cand)
        cand_tokens = set(clean_cand.split())
        
        encontrado = None
        for canonico_existente in candidato_a_canonico.values():
            clean_existente = _limpiar_para_comparar(canonico_existente)
            existente_tokens = set(clean_existente.split())
            if cand_tokens.issubset(existente_tokens) and len(cand_tokens) >= 2:
                encontrado = canonico_existente
                break
        
        if encontrado:
            candidato_a_canonico[cand] = encontrado
        else:
            candidato_a_canonico[cand] = cand

    mapa_final = {}
    for raw, cand in mapa_raw_a_candidato.items():
        mapa_final[raw] = candidato_a_canonico.get(cand, cand)
    return mapa_final


def seed_calendario_institucional():
    """Poblar el calendario institucional con datos si no existen para el ciclo."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            # Verificar si ya existen eventos para 2025-2026 y 2026-2027
            cursor.execute("SELECT COUNT(*) as cnt FROM calendario_institucional WHERE ciclo IN ('2025-2026', '2026-2027')")
            if cursor.fetchone()['cnt'] > 0:
                print("Calendario institucional ya tiene datos base, omitiendo seed para evitar duplicados.")
                return

            
            eventos = [
                # ── SEMESTRE ACTUAL (09 feb 2026 – 31 jul 2026) según PDF del usuario ──
                ('semestral','2025-2026',2,'inicio_periodo','Inicio de subciclo escolar','2026-02-09','2026-02-09',False),
                ('semestral','2025-2026',2,'inhabil','Inhábil: Constitución Mexicana','2026-02-02','2026-02-02',True),
                ('semestral','2025-2026',2,'inhabil','Inhábil: Natalicio de Benito Juárez','2026-03-16','2026-03-16',True),
                ('semestral','2025-2026',2,'examen_parcial','Primer examen Parcial','2026-03-17','2026-03-25',False),
                ('semestral','2025-2026',2,'vacaciones','Vacaciones de semana santa','2026-03-30','2026-04-10',True),
                ('semestral','2025-2026',2,'inhabil','Inhábil: Día del trabajo','2026-05-01','2026-05-01',True),
                ('semestral','2025-2026',2,'evaluacion','Evaluación docente','2026-05-04','2026-05-07',False),
                ('semestral','2025-2026',2,'inhabil','Inhábil: Día del Maestro','2026-05-15','2026-05-15',True),
                ('semestral','2025-2026',2,'examen_parcial','Segundo examen parcial','2026-05-25','2026-06-05',False),
                ('semestral','2025-2026',2,'examen_ordinario','Exámenes ordinarios','2026-06-22','2026-07-03',False),
                ('semestral','2025-2026',2,'inscripcion','Inscripciones a exámenes extraordinarios','2026-07-06','2026-07-07',False),
                ('semestral','2025-2026',2,'examen_extraordinario','Exámenes extraordinarios','2026-07-08','2026-07-21',False),
                ('semestral','2025-2026',2,'entrega','Entrega reporte de actividades','2026-07-22','2026-07-22',False),
                ('semestral','2025-2026',2,'entrega','Entrega plan de actividades','2026-07-29','2026-07-29',False),
                ('semestral','2025-2026',2,'fin_periodo','Fin de subciclo escolar','2026-07-31','2026-07-31',False),

                # ── SEMESTRE 1: 31 ago 2026 – 31 ene 2027 ──
                ('semestral','2026-2027',1,'inicio_periodo','Inicio de periodo escolar','2026-08-31','2026-08-31',False),
                ('semestral','2026-2027',1,'inhabil','Inhábil: Independencia de México','2026-09-16','2026-09-16',True),
                ('semestral','2026-2027',1,'examen_parcial','Primer examen parcial','2026-10-05','2026-10-14',False),
                ('semestral','2026-2027',1,'inhabil','Inhábil: Día de muertos','2026-11-02','2026-11-02',True),
                ('semestral','2026-2027',1,'evaluacion','Evaluación docente','2026-11-03','2026-11-06',False),
                ('semestral','2026-2027',1,'inhabil','Inhábil: Revolución Mexicana','2026-11-16','2026-11-16',True),
                ('semestral','2026-2027',1,'examen_parcial','Segundo examen parcial','2026-11-18','2026-11-27',False),
                ('semestral','2026-2027',1,'vacaciones','Vacaciones navideñas','2026-12-21','2027-01-01',True),
                ('semestral','2026-2027',1,'examen_ordinario','Exámenes ordinarios','2027-01-04','2027-01-13',False),
                ('semestral','2026-2027',1,'inscripcion','Inscripciones a exámenes extraordinarios','2027-01-08','2027-01-19',False),
                ('semestral','2026-2027',1,'examen_extraordinario','Exámenes extraordinarios','2027-01-18','2027-01-27',False),
                ('semestral','2026-2027',1,'fin_periodo','Fin de periodo escolar','2027-01-31','2027-01-31',False),
                # ── SEMESTRE 2: 10 feb 2027 – 31 jul 2027 ──
                ('semestral','2026-2027',2,'inhabil','Inhábil: Día de la Constitución','2027-02-01','2027-02-01',True),
                ('semestral','2026-2027',2,'inhabil','Inhábil: Carnaval','2027-02-08','2027-02-09',True),
                ('semestral','2026-2027',2,'inicio_periodo','Inicio de periodo escolar','2027-02-10','2027-02-10',False),
                ('semestral','2026-2027',2,'inhabil','Inhábil: Natalicio de Benito Juárez','2027-03-15','2027-03-15',True),
                ('semestral','2026-2027',2,'vacaciones','Vacaciones de semana santa','2027-03-22','2027-04-02',True),
                ('semestral','2026-2027',2,'examen_parcial','Primer examen parcial','2027-04-05','2027-04-14',False),
                ('semestral','2026-2027',2,'inhabil','Inhábil: Día del trabajo','2027-05-01','2027-05-01',True),
                ('semestral','2026-2027',2,'evaluacion','Evaluación docente','2027-05-03','2027-05-06',False),
                ('semestral','2026-2027',2,'inhabil','Inhábil: Día del Maestro','2027-05-14','2027-05-14',True),
                ('semestral','2026-2027',2,'examen_parcial','Segundo examen parcial','2027-05-17','2027-05-26',False),
                ('semestral','2026-2027',2,'examen_ordinario','Exámenes ordinarios','2027-06-23','2027-07-02',False),
                ('semestral','2026-2027',2,'inscripcion','Inscripciones a exámenes extraordinarios','2027-06-29','2027-07-08',False),
                ('semestral','2026-2027',2,'examen_extraordinario','Exámenes extraordinarios','2027-07-07','2027-07-16',False),
                ('semestral','2026-2027',2,'fin_periodo','Fin de periodo escolar','2027-07-31','2027-07-31',False),
                
                # ── CUATRIMESTRE 1 (Enero-Abril 2026) ──
                ('cuatrimestral','2025-2026',1,'inicio_periodo','Inicio de Cuatrimestre','2026-01-05','2026-01-05',False),
                ('cuatrimestral','2025-2026',1,'inhabil','Inhábil: Día de Reyes','2026-01-06','2026-01-06',True),
                ('cuatrimestral','2025-2026',1,'inhabil','Inhábil: Constitución Mexicana','2026-02-02','2026-02-02',True),
                ('cuatrimestral','2025-2026',1,'inhabil','Inhábil: Natalicio de Benito Juárez','2026-03-16','2026-03-16',True),
                ('cuatrimestral','2025-2026',1,'vacaciones','Vacaciones de semana santa','2026-03-30','2026-04-10',True),
                ('cuatrimestral','2025-2026',1,'examen_parcial','Primer Parcial','2026-02-16','2026-02-21',False),
                ('cuatrimestral','2025-2026',1,'examen_parcial','Segundo Parcial','2026-03-23','2026-03-28',False),
                ('cuatrimestral','2025-2026',1,'examen_ordinario','Exámenes Finales','2026-04-20','2026-04-25',False),
                ('cuatrimestral','2025-2026',1,'fin_periodo','Fin de Cuatrimestre','2026-04-30','2026-04-30',False),

                # ── CUATRIMESTRE 2 (Mayo-Agosto 2026) ──
                ('cuatrimestral','2025-2026',2,'inicio_periodo','Inicio de Cuatrimestre','2026-05-04','2026-05-04',False),
                ('cuatrimestral','2025-2026',2,'inhabil','Inhábil: Día del trabajo (Recorrido)','2026-05-05','2026-05-05',True),
                ('cuatrimestral','2025-2026',2,'inhabil','Inhábil: Día del Maestro','2026-05-15','2026-05-15',True),
                ('cuatrimestral','2025-2026',2,'examen_parcial','Primer Parcial','2026-06-08','2026-06-15',False),
                ('cuatrimestral','2025-2026',2,'evaluacion','Evaluación docente','2026-07-06','2026-07-09',False),
                ('cuatrimestral','2025-2026',2,'examen_parcial','Segundo Parcial','2026-07-20','2026-07-27',False),
                ('cuatrimestral','2025-2026',2,'examen_ordinario','Exámenes ordinarios','2026-08-03','2026-08-10',False),
                ('cuatrimestral','2025-2026',2,'inscripcion','Inscripción a exámenes extraordinarios','2026-08-11','2026-08-12',False),
                ('cuatrimestral','2025-2026',2,'examen_extraordinario','Exámenes extraordinarios','2026-08-14','2026-08-21',False),
                ('cuatrimestral','2025-2026',2,'fin_periodo','Fin de Cuatrimestre','2026-08-31','2026-08-31',False),
                # ── CUATRIMESTRE 3 (Septiembre-Diciembre 2026) ── según PDF del usuario
                ('cuatrimestral','2026-2027',3,'inicio_periodo','Inicio de Cuatrimestre','2026-08-31','2026-08-31',False),
                ('cuatrimestral','2026-2027',3,'inhabil','Inhábil: Independencia de México','2026-09-16','2026-09-16',True),
                ('cuatrimestral','2026-2027',3,'examen_parcial','Primer examen parcial','2026-09-28','2026-10-05',False),
                ('cuatrimestral','2026-2027',3,'evaluacion','Evaluación docente','2026-10-26','2026-10-29',False),
                ('cuatrimestral','2026-2027',3,'inhabil','Inhábil: Día de muertos','2026-11-02','2026-11-02',True),
                ('cuatrimestral','2026-2027',3,'examen_parcial','Segundo examen parcial','2026-11-03','2026-11-10',False),
                ('cuatrimestral','2026-2027',3,'inhabil','Inhábil: Revolución Mexicana','2026-11-16','2026-11-16',True),
                ('cuatrimestral','2026-2027',3,'examen_ordinario','Exámenes ordinarios','2026-11-30','2026-12-07',False),
                ('cuatrimestral','2026-2027',3,'inscripcion','Inscripción a exámenes extraordinarios','2026-12-04','2026-12-11',False),
                ('cuatrimestral','2026-2027',3,'examen_extraordinario','Exámenes extraordinarios','2026-12-10','2026-12-17',False),
                ('cuatrimestral','2026-2027',3,'vacaciones','Vacaciones navideñas','2026-12-21','2026-12-31',True),
                ('cuatrimestral','2026-2027',3,'fin_periodo','Fin de Cuatrimestre','2026-12-31','2026-12-31',False),

                # ── CUATRIMESTRE 1 (Enero-Abril 2027) ──
                ('cuatrimestral','2026-2027',1,'inicio_periodo','Inicio de Cuatrimestre','2027-01-04','2027-01-04',False),
                ('cuatrimestral','2026-2027',1,'inhabil','Inhábil: Día de Reyes','2027-01-06','2027-01-06',True),
                ('cuatrimestral','2026-2027',1,'inhabil','Inhábil: Día de la Constitución','2027-02-01','2027-02-01',True),
                ('cuatrimestral','2026-2027',1,'inhabil','Inhábil: Natalicio de Benito Juárez','2027-03-15','2027-03-15',True),
                ('cuatrimestral','2026-2027',1,'vacaciones','Vacaciones de semana santa','2027-03-22','2027-04-02',True),
                ('cuatrimestral','2026-2027',1,'examen_parcial','Primer Parcial','2027-02-15','2027-02-20',False),
                ('cuatrimestral','2026-2027',1,'examen_parcial','Segundo Parcial','2027-03-22','2027-03-27',False),
                ('cuatrimestral','2026-2027',1,'examen_ordinario','Exámenes Finales','2027-04-19','2027-04-24',False),
                ('cuatrimestral','2026-2027',1,'fin_periodo','Fin de Cuatrimestre','2027-04-30','2027-04-30',False),

                # ── CUATRIMESTRE 2 (Mayo-Agosto 2027) ──
                ('cuatrimestral','2026-2027',2,'inicio_periodo','Inicio de Cuatrimestre','2027-05-03','2027-05-03',False),
                ('cuatrimestral','2026-2027',2,'inhabil','Inhábil: Día del trabajo','2027-05-01','2027-05-01',True),
                ('cuatrimestral','2026-2027',2,'inhabil','Inhábil: Día del Maestro','2027-05-14','2027-05-14',True),
                ('cuatrimestral','2026-2027',2,'examen_parcial','Primer Parcial','2027-06-14','2027-06-19',False),
                ('cuatrimestral','2026-2027',2,'examen_parcial','Segundo Parcial','2027-07-19','2027-07-24',False),
                ('cuatrimestral','2026-2027',2,'examen_ordinario','Exámenes Finales','2027-08-09','2027-08-14',False),
                ('cuatrimestral','2026-2027',2,'fin_periodo','Fin de Cuatrimestre','2027-08-20','2027-08-20',False),
            ]
            for ev in eventos:
                cursor.execute(
                    "INSERT INTO calendario_institucional (plan,ciclo,periodo,tipo_evento,descripcion,fecha_inicio,fecha_fin,suspende_clases) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                    ev
                )
        connection.commit()
        print("Calendario institucional sembrado exitosamente.")
    except Exception as e:
        print(f"Advertencia en seed calendario: {e}")
    finally:
        connection.close()


def force_reseed_calendario():
    """Limpia TODOS los datos del calendario y reinserta el seed limpio."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM calendario_institucional")
        connection.commit()
        print("Calendario institucional limpiado.")
    finally:
        connection.close()
    # Ahora el seed se insertará porque la tabla está vacía
    seed_calendario_institucional()


@asynccontextmanager
async def lifespan(_app: FastAPI):
    try:
        migrar_columnas_verificacion()
        seed_calendario_institucional()
    except Exception as e:
        print(f"Advertencia en inicio de app (DB connection issue): {e}")
    yield


app = FastAPI(title="API SIPREF ULA", lifespan=lifespan)

# Configurar directorio de uploads
UPLOAD_DIR = os.path.join(base_dir, "uploads", "calendarios")
os.makedirs(UPLOAD_DIR, exist_ok=True)
app.mount("/uploads", StaticFiles(directory=os.path.join(base_dir, "uploads")), name="uploads")

@app.get("/")
def read_root():
    return {"status": "online", "message": "Sistema ULA Backend API is running", "version": "1.0"}


import urllib.request
import urllib.error

def _enviar_smtp(correo_destino: str, asunto: str, html: str):
    """Envía correos usando la API de EmailJS (vía HTTP) para evadir el bloqueo de puertos de Render."""
    service_id = os.getenv("EMAILJS_SERVICE_ID")
    template_id = os.getenv("EMAILJS_TEMPLATE_ID")
    public_key = os.getenv("EMAILJS_PUBLIC_KEY")
    private_key = os.getenv("EMAILJS_PRIVATE_KEY")

    if not service_id or not template_id or not public_key:
        raise ValueError("Faltan las credenciales de EmailJS en las variables de entorno.")

    payload = {
        "service_id": service_id,
        "template_id": template_id,
        "user_id": public_key,
        "accessToken": private_key,
        "template_params": {
            "to_email": correo_destino,
            "asunto": asunto,
            "html_content": html
        }
    }

    req = urllib.request.Request(
        'https://api.emailjs.com/api/v1.0/email/send',
        data=json.dumps(payload).encode('utf-8'),
        headers={
            'Content-Type': 'application/json',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        },
        method='POST'
    )

    try:
        with urllib.request.urlopen(req, timeout=10) as response:
            pass # Si es 200 OK, el correo se envió correctamente
    except urllib.error.HTTPError as e:
        error_msg = e.read().decode('utf-8')
        raise ValueError(f"Error de EmailJS ({e.code}): {error_msg}")
    except Exception as e:
        raise ValueError(f"Error de conexión con EmailJS: {str(e)}")


def enviar_correo_otp(correo_destino: str, codigo: str, nombre: str):
    html = f"""
    <html><body style="font-family:Arial,sans-serif;background:#f4f6fb;padding:32px;">
      <div style="max-width:420px;margin:auto;background:#fff;border-radius:16px;
                  padding:36px;text-align:center;box-shadow:0 4px 20px rgba(0,0,0,.08);">
        <h2 style="color:#1c355e;margin-bottom:4px;">SIPREF</h2>
        <p style="color:#75777f;font-size:13px;">Universidad Latino</p>
        <p style="color:#44464e;">Hola <b>{nombre}</b>, tu código de verificación es:</p>
        <div style="font-size:38px;font-weight:bold;letter-spacing:10px;color:#1c355e;
                    background:#f0f4ff;padding:18px;border-radius:10px;margin:24px 0;">
          {codigo}
        </div>
        <p style="color:#75777f;font-size:12px;">Caduca en 15 minutos. No lo compartas con nadie.</p>
      </div>
    </body></html>
    """
    sender_email = os.getenv("SMTP_USER", "")
    smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", 587))
    smtp_pass = os.getenv("SMTP_PASSWORD", "") or os.getenv("SMTP_PASS", "")

    if sender_email and smtp_pass:
        try:
            msg = MIMEMultipart()
            msg['From'] = sender_email
            msg['To'] = correo_destino
            msg['Subject'] = "Código de verificación — SIPREF ULA"
            msg.attach(MIMEText(html, 'html'))
            with smtplib.SMTP(smtp_host, smtp_port, timeout=4.0) as server:
                server.starttls()
                server.login(sender_email, smtp_pass)
                server.sendmail(sender_email, [correo_destino], msg.as_string())
            return
        except Exception as e:
            print(f"SMTP fallo para OTP, probando EmailJS fallback: {e}")

    # Fallback si falla SMTP o si no hay credenciales SMTP
    _enviar_smtp(correo_destino, "Código de verificación — SIPREF ULA", html)

# Configuración de CORS para permitir la conexión desde Vite (React)
app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5173",
        "http://127.0.0.1:5174",
        "*"
    ],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------------------------------------------------
# CONEXIÓN A LA BASE DE DATOS TiDB CLOUD (OFICIAL WINDOWS)
# ---------------------------------------------------------
def get_db_connection():
    db_host = os.getenv("DB_HOST")
    db_user = os.getenv("DB_USER")
    db_password = os.getenv("DB_PASSWORD")
    db_name = os.getenv("DB_NAME")
    db_port = int(os.getenv("DB_PORT", 4000))

    missing_vars = [name for name, value in (
        ("DB_HOST", db_host),
        ("DB_USER", db_user),
        ("DB_PASSWORD", db_password),
        ("DB_NAME", db_name),
    ) if not value]
    
    if missing_vars:
        raise RuntimeError(f"Faltan variables de entorno: {', '.join(missing_vars)}")

    # SOLUCIÓN: Usar certifi.where() asegura una conexión SSL exitosa a TiDB 
    # en cualquier entorno, ignorando rutas estáticas locales.
    return pymysql.connect(
        host=db_host,
        user=db_user,
        password=db_password,
        db=db_name,
        port=db_port,
        cursorclass=pymysql.cursors.DictCursor,
        ssl_verify_cert=True,
        ssl_verify_identity=True,
        ssl={"ca": certifi.where()} # <--- Generación automática del certificado
    )

# ---------------------------------------------------------
# MODELOS DE DATOS (PYDANTIC)
# ---------------------------------------------------------
class RegistroUsuario(BaseModel):
    nombre: str
    correo: str
    turno: str
    password: str

class LoginUsuario(BaseModel):
    correo: str
    password: str

class Horario(BaseModel):
    docente: str
    licenciatura: str
    asignatura: str
    horario: str
    aulaAsignada: str
    archivo: str
    semestre: str = ""
    cuatrimestre: str = ""
    grupo: str = ""

class Aula(BaseModel):
    nombre: str
    edificio: str = ""
    capacidad: int = 0
    equipos: list = []
    estado: str = "Activo"

class MantenimientoUpdate(BaseModel):
    en_mantenimiento: bool
    inicio_mantenimiento: Optional[str] = None  # ISO datetime, ej: "2026-05-25T08:00"
    fin_mantenimiento: Optional[str] = None     # ISO datetime, ej: "2026-05-25T18:00"
    aula_temporal: Optional[str] = None

class Docente(BaseModel):
    nombre: str
    especialidad: str = ""
    materias: list = []
    correo: str = ""

class Suplencia(BaseModel):
    docente_id: int
    suplente_id: int
    materia: str = ""
    dia: str = ""
    fecha: str
    hora_inicio: str
    hora_fin: str

class SuplenciaHorarios(BaseModel):
    docente_nombre: str
    suplente_nombre: str
    materia: str = ""
    dia: str = ""
    fecha: str
    hora_inicio: str
    hora_fin: str

class ActualizarUsuario(BaseModel):
    nombre: str
    turno: str

class CambiarPassword(BaseModel):
    password_actual: str
    password_nueva: str

class VerificacionCorreo(BaseModel):
    correo: str
    codigo: str

class RecuperarContrasena(BaseModel):
    correo: str

class RestablecerContrasena(BaseModel):
    correo: str
    codigo: str
    nueva_password: str

class ReprogramacionRequest(BaseModel):
    docente_nombre: str
    clase_original_id: Optional[int] = None
    clases_originales_ids: Optional[List[int]] = None
    clase_original_horario: str
    clase_original_asignatura: str
    nueva_fecha: str
    nuevo_dia: str
    nueva_hora_inicio: str
    nueva_hora_fin: str
    nueva_aula: str
    motivo: Optional[str] = ""

# ---------------------------------------------------------
# HELPERS
# ---------------------------------------------------------
from urllib.parse import unquote

def obtener_usuario(request: Request) -> str:
    user = request.headers.get("x-usuario", "Sistema")
    return unquote(user)

def registrar_bitacora(
    cursor,
    usuario: str,
    tipo_operacion: str,
    modulo_afectado: str,
    registro_id: int = None,
    docente: str = None,
    suplente: str = None,
    licenciatura: str = None,
    grado_cuatrimestre: str = None,
    asignatura: str = None,
    grupo: str = None,
    dia_anterior: str = None,
    dia_nuevo: str = None,
    hora_anterior: str = None,
    hora_nueva: str = None,
    aula_anterior: str = None,
    aula_nueva: str = None,
    turno_anterior: str = None,
    turno_nuevo: str = None,
    estado_anterior: str = None,
    estado_nuevo: str = None,
    motivo: str = None,
    datos_anteriores: dict = None,
    datos_nuevos: dict = None
):
    query = """
        INSERT INTO bitacora_cambios (
            usuario, tipo_operacion, modulo_afectado, registro_id,
            docente, suplente, licenciatura, grado_cuatrimestre, asignatura, grupo,
            dia_anterior, dia_nuevo, hora_anterior, hora_nueva,
            aula_anterior, aula_nueva, turno_anterior, turno_nuevo,
            estado_anterior, estado_nuevo, motivo, datos_anteriores, datos_nuevos
        ) VALUES (
            %s, %s, %s, %s,
            %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s, %s
        )
    """
    d_ant = json.dumps(datos_anteriores, ensure_ascii=False) if datos_anteriores else None
    d_nue = json.dumps(datos_nuevos, ensure_ascii=False) if datos_nuevos else None
    
    cursor.execute(query, (
        usuario, tipo_operacion, modulo_afectado, registro_id,
        docente, suplente, licenciatura, grado_cuatrimestre, asignatura, grupo,
        dia_anterior, dia_nuevo, hora_anterior, hora_nueva,
        aula_anterior, aula_nueva, turno_anterior, turno_nuevo,
        estado_anterior, estado_nuevo, motivo, d_ant, d_nue
    ))

_DIAS_MAP = {'domingo': 0, 'lunes': 1, 'martes': 2, 'miercoles': 3, 'jueves': 4, 'viernes': 5, 'sabado': 6}

def _parse_horario_minutos(horario_str: str):
    """Convierte 'Lunes 07:00-09:00' → (dia_index, inicio_min, fin_min). Retorna (None,None,None) si no parsea."""
    if not horario_str:
        return None, None, None
    partes = horario_str.strip().split(' ', 1)
    if len(partes) < 2:
        return None, None, None
    dia_raw = unicodedata.normalize('NFD', partes[0].lower())
    dia_raw = ''.join(c for c in dia_raw if unicodedata.category(c) != 'Mn')
    dia_index = _DIAS_MAP.get(dia_raw)
    tiempo = re.sub(r'-+', '-', partes[1].replace(' ', ''))
    m = re.match(r'(\d{1,2}):(\d{2})-(\d{1,2}):(\d{2})', tiempo)
    if not m:
        return dia_index, None, None
    h_i, m_i, h_f, m_f = map(int, m.groups())
    return dia_index, h_i * 60 + m_i, h_f * 60 + m_f

def _aplicar_mantenimiento(horarios: list) -> list:
    """Reemplaza aula_asignada por aula_temporal si el aula está en mantenimiento vigente."""
    ahora = datetime.now()
    for h in horarios:
        en_mant = h.pop('en_mantenimiento', None)
        inicio_mant = h.pop('inicio_mantenimiento', None)
        fin_mant = h.pop('fin_mantenimiento', None)
        aula_temp = h.pop('aula_temporal', None)
        if not en_mant or not aula_temp:
            continue

        inicio_dt = None
        if inicio_mant:
            if isinstance(inicio_mant, datetime):
                inicio_dt = inicio_mant
            else:
                try:
                    inicio_dt = datetime.fromisoformat(str(inicio_mant))
                except Exception:
                    pass

        fin_dt = None
        if fin_mant:
            if isinstance(fin_mant, datetime):
                fin_dt = fin_mant
            else:
                try:
                    fin_dt = datetime.fromisoformat(str(fin_mant))
                except Exception:
                    pass

        vigente = True
        if inicio_dt and ahora < inicio_dt:
            vigente = False
        if fin_dt and ahora > fin_dt:
            vigente = False

        if vigente:
            h['aula_original'] = h['aula_asignada']
            h['aula_asignada'] = aula_temp
            h['aula_reasignada'] = True
    return horarios

def _timedelta_to_str(td):
    if isinstance(td, timedelta):
        total = int(td.total_seconds())
        return f"{total // 3600:02d}:{(total % 3600) // 60:02d}"
    return str(td)[:5] if td else ""

def _normalizar_licenciatura(texto: str) -> str:
    """Elimina duplicación de palabras en el nombre de licenciatura extraído del PDF."""
    if not texto:
        return texto
    # Palabras pegadas repetidas: "LICENCIATURALICENCIATURA" → "LICENCIATURA"
    texto = re.sub(r'([A-Za-záéíóúÁÉÍÓÚñÑ]{5,})\1', r'\1', texto)
    # Palabras separadas repetidas: "Licenciatura Licenciatura" → "Licenciatura"
    texto = re.sub(r'\b(\w+)\s+\1\b', r'\1', texto, flags=re.IGNORECASE)
    texto = ' '.join(texto.split()).strip()
    
    # Extraer siglas si están entre paréntesis, por ejemplo: "(ISC Plan 2020)" o "(Der Plan 2024)" -> "ISC", "DER"
    match = re.search(r'\(\s*([A-Za-z]{2,6})\b', texto)
    if match:
        return match.group(1).upper()
        
    return texto


def _time_to_mins(t):
    if isinstance(t, timedelta):
        return int(t.total_seconds() // 60)
    try:
        parts = str(t).split(':')
        return int(parts[0]) * 60 + int(parts[1])
    except Exception:
        return 0

# ---------------------------------------------------------
# ENDPOINTS (RUTAS DE LA API)
# ---------------------------------------------------------

@app.get("/api/ciclos-disponibles")
def ciclos_disponibles(fecha: str = None):
    """Retorna la lista de ciclos escolares que tienen calendarios cargados, más el ciclo actual y siguiente."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT ciclo_escolar FROM calendarios ORDER BY ciclo_escolar DESC")
            rows = cursor.fetchall()
            ciclos_db = [r['ciclo_escolar'] for r in rows if r['ciclo_escolar']]
            
            # Generar ciclo actual y siguiente basado en la fecha (o fecha de prueba)
            if fecha:
                now_str = fecha
                now = datetime.strptime(fecha, "%Y-%m-%d")
            else:
                now_str = datetime.now().strftime("%Y-%m-%d")
                now = datetime.now()
                
            # 1. Intentar deducir el ciclo actual mirando los límites de fechas en la BD
            cursor.execute("""
                SELECT ciclo FROM calendario_institucional 
                GROUP BY ciclo
                HAVING MIN(fecha_inicio) <= %s AND MAX(fecha_fin) >= %s
                LIMIT 1
            """, (now_str, now_str))
            row_ciclo = cursor.fetchone()
            
            if row_ciclo:
                ciclo_actual = row_ciclo['ciclo']
                try:
                    parts = ciclo_actual.split('-')
                    ciclo_siguiente = f"{int(parts[0])+1}-{int(parts[1])+1}"
                except:
                    year = now.year
                    ciclo_siguiente = f"{year+1}-{year+2}"
            else:
                # 2. Fallback heurístico si no hay datos que cubran esa fecha
                year = now.year
                if now.month >= 8:
                    ciclo_actual = f"{year}-{year+1}"
                    ciclo_siguiente = f"{year+1}-{year+2}"
                else:
                    ciclo_actual = f"{year-1}-{year}"
                    ciclo_siguiente = f"{year}-{year+1}"
            
            todos = list(set(ciclos_db + [ciclo_actual, ciclo_siguiente]))
            todos.sort(reverse=True)
            return {"ciclos": todos, "ciclo_actual": ciclo_actual}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.get("/api/calendarios")
def obtener_calendarios(ciclo_escolar: str = ""):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            if ciclo_escolar:
                cursor.execute(
                    "SELECT id, tipo, carrera, archivo_nombre, archivo_url, ciclo_escolar FROM calendarios WHERE ciclo_escolar = %s",
                    (ciclo_escolar,)
                )
            else:
                cursor.execute("SELECT id, tipo, carrera, archivo_nombre, archivo_url, ciclo_escolar FROM calendarios")
            calendarios = cursor.fetchall()
        return calendarios
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()

@app.delete("/api/calendarios/ciclo/{ciclo_escolar}")
def eliminar_ciclo_escolar(ciclo_escolar: str):
    """Elimina completamente un ciclo escolar de la base de datos (calendarios, eventos y exámenes)."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            # Obtener nombres de archivos de exámenes de este ciclo para borrar sus registros
            cursor.execute("SELECT archivo_nombre FROM calendarios WHERE ciclo_escolar = %s AND tipo = 'examenes'", (ciclo_escolar,))
            archivos_examenes = cursor.fetchall()
            for arc in archivos_examenes:
                cursor.execute("DELETE FROM examenes_calendario WHERE archivo_origen = %s", (arc['archivo_nombre'],))
                
            # Eliminar eventos institucionales sincronizados
            cursor.execute("DELETE FROM calendario_institucional WHERE ciclo = %s", (ciclo_escolar,))
            
            # Eliminar de la tabla calendarios (PDFs)
            cursor.execute("DELETE FROM calendarios WHERE ciclo_escolar = %s", (ciclo_escolar,))
            
            connection.commit()
        return {"message": f"El ciclo escolar {ciclo_escolar} ha sido eliminado correctamente."}
    except Exception as e:
        connection.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


import shutil


# -------------------------------------------------------
# PARSER DE PDF DE EXÁMENES
# -------------------------------------------------------
def parsear_pdf_examenes(file_bytes: bytes) -> list:
    """Parsea un PDF de calendario de exámenes de Universidad Latino y retorna datos estructurados."""
    resultados = []
    meses = ['enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
             'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
    dias_semana = ['lunes', 'martes', 'miércoles', 'jueves', 'viernes', 'sábado', 'domingo']

    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            tables = page.extract_tables()

            # Identificar periodos en orden de aparición en el texto
            periodos = []
            text_upper = text.upper()
            # Buscar cada periodo y su posición para ordenarlos
            posiciones = []
            for nombre, buscar in [
                ('Primer Parcial', 'PRIMER PARCIAL'),
                ('Segundo Parcial', 'SEGUNDO PARCIAL'),
                ('Ordinarios', 'ORDINARIO'),
                ('Extraordinarios', 'EXTRAORDINARIO'),
            ]:
                pos = text_upper.find(buscar)
                if pos != -1:
                    posiciones.append((pos, nombre))
            posiciones.sort(key=lambda x: x[0])
            periodos = [p[1] for p in posiciones]

            # Filtrar tablas que realmente contienen datos de exámenes
            exam_tables = []
            for table in tables:
                if not table or len(table) < 2:
                    continue
                flat = ' '.join(str(c or '') for row in table for c in row).lower()
                has_days = any(d in flat for d in dias_semana)
                has_dates = any(m in flat for m in meses)
                if has_days or has_dates:
                    exam_tables.append(table)

            # Procesar cada tabla y emparejarla con su periodo
            for idx, table in enumerate(exam_tables):
                periodo = periodos[idx] if idx < len(periodos) else f"Periodo {idx + 1}"

                # Encontrar fila de días y fila de fechas
                dia_row_idx = None
                fecha_row_idx = None
                for row_idx, row in enumerate(table):
                    row_text = ' '.join(str(c or '') for c in row).lower()
                    if any(d in row_text for d in dias_semana) and dia_row_idx is None:
                        dia_row_idx = row_idx
                    if any(m in row_text for m in meses) and fecha_row_idx is None:
                        fecha_row_idx = row_idx

                if dia_row_idx is None and fecha_row_idx is None:
                    continue

                # Si solo encontramos una, asumir la otra
                if dia_row_idx is None:
                    dia_row_idx = max(0, fecha_row_idx - 1)
                if fecha_row_idx is None:
                    fecha_row_idx = dia_row_idx + 1

                dias = table[dia_row_idx] if dia_row_idx < len(table) else []
                fechas = table[fecha_row_idx] if fecha_row_idx < len(table) else []

                # Filas de datos empiezan después de la fila de fechas
                data_start = max(dia_row_idx, fecha_row_idx) + 1

                for row in table[data_start:]:
                    if not row or not row[0]:
                        continue
                    semestre = str(row[0]).strip()
                    # Saltar filas que no sean semestres reales
                    if not semestre or len(semestre) < 2:
                        continue
                    # Saltar filas que parecen notas
                    if 'nota' in semestre.lower() or 'todos' in semestre.lower():
                        continue

                    for col in range(1, len(row)):
                        materia = str(row[col] or '').strip()
                        # Limpiar saltos de línea dentro de la materia
                        materia = ' '.join(materia.split())
                        if materia and len(materia) > 1:
                            dia = str(dias[col] or '').strip() if col < len(dias) else ''
                            fecha = str(fechas[col] or '').strip() if col < len(fechas) else ''
                            resultados.append({
                                'periodo': periodo,
                                'semestre': semestre,
                                'dia': dia,
                                'fecha': fecha,
                                'materia': materia
                            })

    return resultados

def _parsear_calendario_institucional_pdf(contenido_pdf: bytes):
    """
    Parsea el contenido binario de un PDF de Calendario Institucional.
    Busca páginas con tablas [Mes, Fecha/periodo, Actividad/Actividades].
    Soporta formatos de fecha: "15", "08 al 21", "11 y 12", "30-31", "30 al 31".
    Fusiona eventos que cruzan meses (ej: Vacaciones dic→ene) en un solo registro.
    Retorna lista de tuplas (plan, ciclo, periodo, tipo_evento, descripcion, fecha_inicio, fecha_fin, suspende_clases).
    """
    import io
    import re
    
    MESES = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
        'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
        'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
    }
    
    def clasificar_evento(descripcion: str):
        desc_lower = descripcion.lower()
        if 'inhábil' in desc_lower or 'inhabil' in desc_lower:
            return 'inhabil', True
        elif 'vacaciones' in desc_lower:
            return 'vacaciones', True
        elif 'extraordinario' in desc_lower and 'inscripci' in desc_lower:
            return 'inscripcion', False
        elif 'extraordinario' in desc_lower:
            return 'examen_extraordinario', False
        elif 'examen parcial' in desc_lower or ('parcial' in desc_lower and 'examen' not in desc_lower.replace('parcial','')):
            return 'examen_parcial', False
        elif 'parcial' in desc_lower:
            return 'examen_parcial', False
        elif 'ordinario' in desc_lower or 'exámenes finales' in desc_lower:
            return 'examen_ordinario', False
        elif 'inscripci' in desc_lower:
            return 'inscripcion', False
        elif 'evaluación docente' in desc_lower or 'evaluacion docente' in desc_lower:
            return 'evaluacion', False
        elif 'inicio de subciclo' in desc_lower or 'inicio de periodo' in desc_lower or 'inicio de cuatrimestre' in desc_lower:
            return 'inicio_periodo', False
        elif 'fin de periodo' in desc_lower or 'fin de subciclo' in desc_lower or 'fin de cuatrimestre' in desc_lower:
            return 'fin_periodo', False
        elif 'entrega' in desc_lower:
            return 'entrega', False
        elif 'inducción' in desc_lower or 'induccion' in desc_lower:
            return 'induccion', False
        elif 'junta' in desc_lower:
            return 'otro', False
        elif 'inicio' in desc_lower:
            return 'inicio_periodo', False
        else:
            return 'otro', False
    
    def parsear_fecha_periodo(fecha_str, mes_actual, anio_mes):
        """Parsea múltiples formatos de fecha: '15', '08 al 21', '11 y 12', '30-31'."""
        fecha_str = fecha_str.strip()
        # Formato: "08 al 21" (rango con 'al')
        match_rango = re.match(r'(\d{1,2})\s*al\s*(\d{1,2})', fecha_str)
        if match_rango:
            dia_ini = int(match_rango.group(1))
            dia_fin = int(match_rango.group(2))
            return f"{anio_mes}-{mes_actual:02d}-{dia_ini:02d}", f"{anio_mes}-{mes_actual:02d}-{dia_fin:02d}"
        # Formato: "11 y 12" (rango con 'y')
        match_y = re.match(r'(\d{1,2})\s*y\s*(\d{1,2})', fecha_str)
        if match_y:
            dia_ini = int(match_y.group(1))
            dia_fin = int(match_y.group(2))
            return f"{anio_mes}-{mes_actual:02d}-{dia_ini:02d}", f"{anio_mes}-{mes_actual:02d}-{dia_fin:02d}"
        # Formato: "30-31" (rango con guión pegado)
        match_guion = re.match(r'^(\d{1,2})\s*[-–]\s*(\d{1,2})$', fecha_str)
        if match_guion:
            dia_ini = int(match_guion.group(1))
            dia_fin = int(match_guion.group(2))
            return f"{anio_mes}-{mes_actual:02d}-{dia_ini:02d}", f"{anio_mes}-{mes_actual:02d}-{dia_fin:02d}"
        # Formato: "15" (día simple)
        match_dia = re.match(r'^(\d{1,2})$', fecha_str)
        if match_dia:
            dia = int(match_dia.group(1))
            fecha = f"{anio_mes}-{mes_actual:02d}-{dia:02d}"
            return fecha, fecha
        return None, None
    
    def limpiar_descripcion(texto):
        """Limpia la descripción removiendo sufijos de fecha, detalles de licenciaturas, etc."""
        limpio = texto
        # Remover detalles de distribución por licenciatura (ej: "03 LE. / 04 LG. / 05 LN. / 06 LP.")
        limpio = re.sub(r':\s*\d{2}\s*\w+\.?\s*/.*$', '', limpio)
        # Remover referencias a fechas futuras (ej: ": 29 de julio")
        limpio = re.sub(r':\s*\d{1,2}\s+de\s+\w+.*$', '', limpio)
        # Remover referencias a periodos (ej: "correspondientes a Sep 2025 - Ene")
        limpio = re.sub(r'\s+correspondientes?\s+a\s+.*$', '', limpio, flags=re.IGNORECASE)
        # Remover fechas sueltas al final
        limpio = re.sub(r':\s*\d{1,2}\s+de\s+\w+\s+de(?:l)?\s+\d{4}$', '', limpio)
        limpio = limpio.strip().rstrip(':').strip()
        return limpio
    
    def detectar_periodo(text, plan, MESES):
        """Detecta el periodo del subciclo a partir del texto de la página."""
        # Patrón completo: "01 de septiembre de 2025 – 31 de enero de 2026"
        periodo_match = re.search(
            r'(\d{1,2})\s+de\s+(\w+)(?:\s+de(?:l)?\s+(\d{4}))?\s*[–\-]\s*(\d{1,2})\s+de\s+(\w+)(?:\s+de(?:l)?\s+(\d{4}))?',
            text
        )
        if periodo_match:
            mes_inicio_nombre = periodo_match.group(2).lower()
            mes_inicio_num = MESES.get(mes_inicio_nombre, 0)
            if mes_inicio_num == 0:
                return 1
            if plan == 'semestral':
                return 2 if 2 <= mes_inicio_num <= 7 else 1
            elif plan == 'cuatrimestral':
                if 1 <= mes_inicio_num <= 4:
                    return 1
                elif 5 <= mes_inicio_num <= 8:
                    return 2
                else:
                    return 3
                    
        # Fallback para Cuatrimestral basado en texto
        if plan == 'cuatrimestral':
            text_l = text.lower()
            if 'sep-dic' in text_l or 'septiembre' in text_l and 'diciembre' in text_l:
                return 3
            if 'may-ago' in text_l or 'mayo' in text_l and 'agosto' in text_l:
                return 2
            if 'ene-abr' in text_l or 'enero' in text_l and 'abril' in text_l:
                return 1
                
        return 1
    
    def calcular_anio_mes(plan, periodo, mes_actual_num, anio_base, anio_siguiente):
        """Calcula el año correcto para un mes dado según el plan y periodo."""
        if plan == 'semestral':
            # Semestre 1: ago-ene → ago-dic = anio_base, ene = anio_siguiente
            # Semestre 2: feb-jul → siempre anio_siguiente
            if periodo == 1:
                return anio_base if mes_actual_num >= 8 else anio_siguiente
            else:
                return anio_siguiente
        else:
            # Cuatrimestre 3: sep-dic = anio_base, ene = anio_siguiente
            # Cuatrimestre 1: ene-abr = anio_siguiente  
            # Cuatrimestre 2: may-ago = anio_siguiente
            if periodo == 3:
                return anio_base if mes_actual_num >= 8 else anio_siguiente
            elif periodo == 2:
                # Todo Cuatrimestre 2 (Mayo-Agosto) sucede en el anio_siguiente
                return anio_siguiente
            else:
                # Todo Cuatrimestre 1 (Enero-Abril) sucede en el anio_siguiente
                return anio_siguiente
    
    pdf = pdfplumber.open(io.BytesIO(contenido_pdf))
    eventos_crudos = []
    
    for page in pdf.pages:
        text = page.extract_text() or ''
        tables = page.extract_tables()
        tabla_actividades = None
        for table in tables:
            if table and len(table) > 1 and table[0]:
                header = ' '.join([str(c or '') for c in table[0]]).lower()
                # Soportar "Actividad" y "Actividades"
                if 'mes' in header and ('actividad' in header or 'fecha' in header):
                    tabla_actividades = table
                    break
        

        text_lower = text.lower()
        if 'cuatrimestral' in text_lower or 'cuatrimestre' in text_lower:
            plan = 'cuatrimestral'
        elif 'semestral' in text_lower or 'semestre' in text_lower:
            plan = 'semestral'
        else:
            continue
        
        ciclo_match = re.search(r'[Cc]iclo\s+(\d{4})\s*[–\-]\s*(\d{4})', text)
        if ciclo_match:
            ciclo = f"{ciclo_match.group(1)}-{ciclo_match.group(2)}"
            anio_base = int(ciclo_match.group(1))
            anio_siguiente = int(ciclo_match.group(2))
        else:
            continue
        
        if not tabla_actividades:
            continue
        
        periodo = detectar_periodo(text, plan, MESES)
        
        mes_actual_num = None
        for fila in tabla_actividades[1:]:
            if not fila or len(fila) < 3:
                continue
            mes_col = str(fila[0] or '').strip()
            fecha_col = str(fila[1] or '').strip()
            actividad_col = str(fila[2] or '').strip()
            actividad_col = ' '.join(actividad_col.split())
            if not fecha_col or not actividad_col:
                continue
            
            if mes_col:
                for nombre_mes, num_mes in MESES.items():
                    if nombre_mes in mes_col.lower():
                        mes_actual_num = num_mes
                        break
            
            if mes_actual_num is None:
                continue
            
            anio_mes = calcular_anio_mes(plan, periodo, mes_actual_num, anio_base, anio_siguiente)
            
            fecha_inicio, fecha_fin = parsear_fecha_periodo(fecha_col, mes_actual_num, anio_mes)
            if not fecha_inicio:
                continue
            
            tipo_evento, suspende = clasificar_evento(actividad_col)
            desc_limpia = limpiar_descripcion(actividad_col)
            
            eventos_crudos.append((plan, ciclo, periodo, tipo_evento, desc_limpia, fecha_inicio, fecha_fin, 1 if suspende else 0))
    
    pdf.close()
    
    # ── POST-PROCESAMIENTO: Fusionar eventos consecutivos del mismo tipo ──
    # Eventos como "Vacaciones navideñas (Inician)" dic 22-31 + "Vacaciones navideñas (Culminan)" ene 01-02
    # se fusionan en un solo "Vacaciones navideñas" dic 22 → ene 02
    eventos_finales = []
    fusionados = set()
    
    for i, ev in enumerate(eventos_crudos):
        if i in fusionados:
            continue
        plan_i, ciclo_i, per_i, tipo_i, desc_i, fi_i, ff_i, susp_i = ev
        desc_base = re.sub(r'\s*\((?:inicia[n]?|culmina[n]?|inicio|fin)\)\s*$', '', desc_i, flags=re.IGNORECASE).strip()
        
        # Buscar el siguiente evento del mismo tipo para fusionar
        for j in range(i + 1, min(i + 5, len(eventos_crudos))):
            if j in fusionados:
                continue
            plan_j, ciclo_j, per_j, tipo_j, desc_j, fi_j, ff_j, susp_j = eventos_crudos[j]
            desc_base_j = re.sub(r'\s*\((?:inicia[n]?|culmina[n]?|inicio|fin)\)\s*$', '', desc_j, flags=re.IGNORECASE).strip()
            
            if (plan_i == plan_j and ciclo_i == ciclo_j and per_i == per_j and 
                tipo_i == tipo_j and desc_base.lower() == desc_base_j.lower()):
                # Fusionar: tomar fecha_inicio más temprana y fecha_fin más tardía
                fi_i = min(fi_i, fi_j)
                ff_i = max(ff_i, ff_j)
                desc_i = desc_base  # Usar nombre limpio sin "(Inician)"/"(Culminan)"
                fusionados.add(j)
        
        eventos_finales.append((plan_i, ciclo_i, per_i, tipo_i, desc_i if desc_i == desc_base else desc_i, fi_i, ff_i, susp_i))
    
    return eventos_finales


def _guardar_eventos_calendario(eventos):
    """Guarda la lista de eventos en la BD con deduplicación inteligente (no borra el seed)."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            insertados = 0
            for ev in eventos:
                plan, ciclo, periodo, tipo_evento, desc, fi, ff, susp = ev
                # Verificar si ya existe un evento similar (mismo plan, ciclo, periodo, tipo, rango de fechas)
                cursor.execute("""
                    SELECT id FROM calendario_institucional
                    WHERE plan=%s AND ciclo=%s AND periodo=%s AND tipo_evento=%s AND fecha_inicio=%s AND fecha_fin=%s
                    LIMIT 1
                """, (plan, ciclo, periodo, tipo_evento, fi, ff))
                if not cursor.fetchone():
                    cursor.execute(
                        "INSERT INTO calendario_institucional (plan,ciclo,periodo,tipo_evento,descripcion,fecha_inicio,fecha_fin,suspende_clases) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
                        ev
                    )
                    insertados += 1
        connection.commit()
        return insertados
    finally:
        connection.close()


@app.post("/api/calendario-institucional/sincronizar/{ciclo_escolar}")
def sincronizar_calendario_desde_bd(ciclo_escolar: str):
    """Lee el PDF del calendario general ya subido en la BD para un ciclo específico y extrae los eventos."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT archivo_datos FROM calendarios WHERE tipo = 'general' AND ciclo_escolar = %s LIMIT 1", (ciclo_escolar,))
            row = cursor.fetchone()
            if not row or not row.get('archivo_datos'):
                raise HTTPException(status_code=404, detail=f"No hay un calendario institucional subido para el ciclo {ciclo_escolar}. Sube primero el PDF.")
            contenido_pdf = row['archivo_datos']
    finally:
        connection.close()
    
    eventos = _parsear_calendario_institucional_pdf(contenido_pdf)
    
    if not eventos:
        raise HTTPException(status_code=400, detail="No se pudieron extraer eventos del PDF. Verifica que el archivo sea un Calendario Institucional válido.")
    
    insertados = _guardar_eventos_calendario(eventos)
    
    return {"status": "success", "message": f"Se extrajeron {len(eventos)} eventos del PDF. {insertados} nuevos insertados (sin duplicar el seed)."}

@app.get("/api/calendario-institucional/eventos")
def obtener_eventos_institucionales():
    """Obtiene todos los eventos parseados de la base de datos."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT id, plan, ciclo, periodo, tipo_evento, descripcion, fecha_inicio, fecha_fin, suspende_clases FROM calendario_institucional ORDER BY fecha_inicio ASC")
            eventos = cursor.fetchall()
            # Convertir fechas a string si es necesario
            for ev in eventos:
                if hasattr(ev['fecha_inicio'], 'strftime'):
                    ev['fecha_inicio'] = ev['fecha_inicio'].strftime('%Y-%m-%d')
                if hasattr(ev['fecha_fin'], 'strftime'):
                    ev['fecha_fin'] = ev['fecha_fin'].strftime('%Y-%m-%d')
            return {"status": "success", "eventos": eventos}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()

@app.post("/api/calendario-institucional/csv")
async def subir_calendario_institucional_pdf(archivo: UploadFile = File(...)):
    """Sube un PDF nuevo y extrae los eventos."""
    if not archivo.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=400, detail="Solo se permiten archivos PDF.")
    
    contenido = await archivo.read()
    eventos = _parsear_calendario_institucional_pdf(contenido)
    
    if not eventos:
        raise HTTPException(status_code=400, detail="No se pudieron extraer eventos del PDF.")
    
    _guardar_eventos_calendario(eventos)
    
    return {"status": "success", "message": f"Se extrajeron y guardaron {len(eventos)} eventos del calendario institucional."}


@app.post("/api/calendarios/upload")
async def subir_calendario(
    tipo: str = Form(...),
    carrera: str = Form(""),
    ciclo_escolar: str = Form("2025-2026"),
    archivo: UploadFile = File(...)
):
    if archivo.content_type != "application/pdf":
        raise HTTPException(status_code=400, detail="Solo se permiten archivos PDF.")

    file_bytes = await archivo.read()

    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            if tipo == 'examenes' and carrera:
                # ── EXÁMENES: Parsear PDF y guardar datos estructurados ──
                try:
                    datos = parsear_pdf_examenes(file_bytes)
                except Exception as parse_err:
                    raise HTTPException(status_code=400, detail=f"No se pudo leer el PDF: {str(parse_err)}")

                if not datos:
                    raise HTTPException(status_code=400, detail="No se encontraron datos de exámenes en el PDF. Verifica que el formato sea correcto.")

                # Borrar SOLO los datos si se está resubiendo el MISMO archivo (para permitir correcciones)
                cursor.execute("DELETE FROM examenes_calendario WHERE carrera = %s AND archivo_origen = %s", (carrera, archivo.filename))

                # Insertar cada examen extraído
                for d in datos:
                    cursor.execute(
                        "INSERT INTO examenes_calendario (carrera, periodo, semestre, dia, fecha, materia, archivo_origen) VALUES (%s, %s, %s, %s, %s, %s, %s)",
                        (carrera, d['periodo'], d['semestre'], d['dia'], d['fecha'], d['materia'], archivo.filename)
                    )

                # Registrar en la tabla de calendarios (como historial)
                cursor.execute("DELETE FROM calendarios WHERE tipo = %s AND carrera = %s AND archivo_nombre = %s AND ciclo_escolar = %s", (tipo, carrera, archivo.filename, ciclo_escolar))
                cursor.execute(
                    "INSERT INTO calendarios (tipo, carrera, archivo_nombre, archivo_url, ciclo_escolar) VALUES (%s, %s, %s, %s, %s)",
                    (tipo, carrera, archivo.filename, f"/api/examenes-calendario/{carrera}", ciclo_escolar)
                )

                connection.commit()
                return {
                    "message": f"Se extrajeron {len(datos)} exámenes del PDF y se guardaron en la base de datos.",
                    "total": len(datos),
                    "datos": datos
                }
            else:
                # ── GENERAL: Guardar PDF como BLOB ──
                archivo_url = f"/api/calendarios/view/{tipo}/{ciclo_escolar}" + (f"/{carrera}" if carrera else "")
                cursor.execute("DELETE FROM calendarios WHERE tipo = %s AND carrera = %s AND ciclo_escolar = %s", (tipo, carrera, ciclo_escolar))
                cursor.execute(
                    "INSERT INTO calendarios (tipo, carrera, archivo_nombre, archivo_url, archivo_datos, ciclo_escolar) VALUES (%s, %s, %s, %s, %s, %s)",
                    (tipo, carrera, archivo.filename, archivo_url, file_bytes, ciclo_escolar)
                )
                connection.commit()
                return {"message": "Calendario subido y guardado exitosamente", "url": archivo_url}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


def _es_fecha_pasada(fecha_str: str) -> bool:
    """Devuelve True si la fecha (ej. '06 de junio') tiene más de 5 días de haber pasado, asumiendo el ciclo actual."""
    if not fecha_str:
        return True
    fecha_str = str(fecha_str).lower().replace('de ', '')
    partes = fecha_str.split()
    if len(partes) < 2 or not partes[0].isdigit():
        return False
    dia = int(partes[0])
    mes_texto = partes[1]
    meses_map = {
        'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4, 'mayo': 5, 'junio': 6,
        'julio': 7, 'agosto': 8, 'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
    }
    mes_num = meses_map.get(mes_texto)
    if not mes_num: return False
    
    today = datetime.now()
    year = today.year
    # Si el mes del examen es enero/febrero y hoy es octubre-diciembre, el examen es del próximo año
    if mes_num < 3 and today.month > 9:
        year += 1
    # Si el mes del examen es noviembre/diciembre y hoy es enero-febrero, el examen es del año pasado
    elif mes_num > 9 and today.month < 3:
        year -= 1
        
    try:
        fecha_exam = datetime(year, mes_num, dia)
        # Se considera "pasado" si ya pasaron más de 5 días desde la fecha del examen
        return (today - fecha_exam).days > 5
    except ValueError:
        return False


@app.get("/api/examenes-calendario/{carrera}")
def obtener_examenes_calendario(carrera: str):
    """Retorna los exámenes extraídos para la carrera y elimina automáticamente los calendarios finalizados."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT id, periodo, semestre, dia, fecha, materia, archivo_origen FROM examenes_calendario WHERE carrera = %s ORDER BY id",
                (carrera,)
            )
            datos = cursor.fetchall()
            
            if not datos:
                return []
                
            # Agrupar por archivo_origen (o periodo si no hay archivo_origen)
            grupos = {}
            for d in datos:
                clave = d.get('archivo_origen') or 'desconocido'
                if clave not in grupos:
                    grupos[clave] = []
                grupos[clave].append(d)
                
            # Identificar el grupo activo
            grupo_activo = None
            archivos_a_borrar = []
            
            for clave, examenes in grupos.items():
                # Verificar si TODOS los exámenes de este archivo ya pasaron
                todos_pasaron = all(_es_fecha_pasada(e['fecha']) for e in examenes)
                
                if todos_pasaron:
                    archivos_a_borrar.append(clave)
                elif grupo_activo is None:
                    # El primer archivo que no haya pasado es el activo
                    grupo_activo = examenes
                    
            # Borrar de la DB los que ya pasaron
            for clave in archivos_a_borrar:
                cursor.execute("DELETE FROM examenes_calendario WHERE carrera = %s AND archivo_origen = %s", (carrera, clave))
                cursor.execute("DELETE FROM calendarios WHERE carrera = %s AND archivo_nombre = %s", (carrera, clave))
            if archivos_a_borrar:
                connection.commit()
                
            return grupo_activo if grupo_activo else []
            
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.get("/api/examenes-hoy")
def examenes_hoy():
    """Retorna todos los exámenes programados"""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT DISTINCT carrera, materia, periodo, semestre, dia, fecha FROM examenes_calendario ORDER BY carrera, semestre"
            )
            datos = cursor.fetchall()
        return datos
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.get("/api/calendarios/view/{tipo}")
@app.get("/api/calendarios/view/{tipo}/{ciclo_escolar}")
@app.get("/api/calendarios/view/{tipo}/{ciclo_escolar}/{carrera}")
def ver_calendario(tipo: str, ciclo_escolar: str = "", carrera: str = ""):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            if ciclo_escolar:
                cursor.execute(
                    "SELECT archivo_datos, archivo_nombre FROM calendarios WHERE tipo = %s AND carrera = %s AND ciclo_escolar = %s LIMIT 1",
                    (tipo, carrera, ciclo_escolar)
                )
            else:
                cursor.execute(
                    "SELECT archivo_datos, archivo_nombre FROM calendarios WHERE tipo = %s AND carrera = %s LIMIT 1",
                    (tipo, carrera)
                )
            row = cursor.fetchone()
            if not row or not row.get('archivo_datos'):
                raise HTTPException(status_code=404, detail="Calendario no encontrado")

            return Response(
                content=row['archivo_datos'],
                media_type="application/pdf",
                headers={"Content-Disposition": f'inline; filename="{row["archivo_nombre"]}"'}
            )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.post("/api/calendario-institucional/reseed")
def reseed_calendario():
    """Limpia el calendario y reinserta los datos del seed limpio."""
    try:
        force_reseed_calendario()
        return {"status": "success", "message": "Calendario reiniciado con datos limpios del seed."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ── CRUD de Eventos del Calendario ──

@app.post("/api/calendario-institucional/evento")
def crear_evento_calendario(evento: dict):
    """Crea un nuevo evento en el calendario institucional."""
    required = ['plan', 'ciclo', 'periodo', 'tipo_evento', 'descripcion', 'fecha_inicio', 'fecha_fin']
    for field in required:
        if field not in evento:
            raise HTTPException(status_code=400, detail=f"Campo requerido: {field}")
    
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO calendario_institucional (plan, ciclo, periodo, tipo_evento, descripcion, fecha_inicio, fecha_fin, suspende_clases)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                evento['plan'], evento['ciclo'], int(evento['periodo']),
                evento['tipo_evento'], evento['descripcion'],
                evento['fecha_inicio'], evento['fecha_fin'],
                1 if evento.get('suspende_clases', False) else 0
            ))
        connection.commit()
        return {"status": "success", "message": "Evento creado exitosamente."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.put("/api/calendario-institucional/evento/{evento_id}")
def actualizar_evento_calendario(evento_id: int, evento: dict):
    """Actualiza un evento existente."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE calendario_institucional
                SET plan=%s, ciclo=%s, periodo=%s, tipo_evento=%s, descripcion=%s,
                    fecha_inicio=%s, fecha_fin=%s, suspende_clases=%s
                WHERE id=%s
            """, (
                evento.get('plan'), evento.get('ciclo'), int(evento.get('periodo', 1)),
                evento.get('tipo_evento'), evento.get('descripcion'),
                evento.get('fecha_inicio'), evento.get('fecha_fin'),
                1 if evento.get('suspende_clases', False) else 0,
                evento_id
            ))
        connection.commit()
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Evento no encontrado.")
        return {"status": "success", "message": "Evento actualizado exitosamente."}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.delete("/api/calendario-institucional/evento/{evento_id}")
def eliminar_evento_calendario(evento_id: int):
    """Elimina un evento del calendario."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM calendario_institucional WHERE id=%s", (evento_id,))
        connection.commit()
        if cursor.rowcount == 0:
            raise HTTPException(status_code=404, detail="Evento no encontrado.")
        return {"status": "success", "message": "Evento eliminado exitosamente."}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.get("/ping")
def ping():
    return {"status": "ok", "message": "Servidor activo"}


@app.get("/api/estado-academico")
def estado_academico(plan: str = "semestral", fecha: str = None):
    """Retorna el estado académico actual para un tipo de plan, opcionalmente en una fecha específica (para pruebas)."""
    if fecha:
        today = datetime.strptime(fecha, "%Y-%m-%d").date()
    else:
        today = datetime.now().date()
        
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            # 1. Buscar eventos activos hoy independientemente de si estamos en periodo
            cursor.execute("""
                SELECT tipo_evento, descripcion, fecha_inicio, fecha_fin, suspende_clases
                FROM calendario_institucional
                WHERE plan = %s AND fecha_inicio <= %s AND fecha_fin >= %s
                ORDER BY suspende_clases DESC
            """, (plan, today, today))
            eventos_hoy = cursor.fetchall()
            for ev in eventos_hoy:
                ev['fecha_inicio'] = str(ev['fecha_inicio'])
                ev['fecha_fin'] = str(ev['fecha_fin'])

            # 2. Buscar si estamos dentro de un periodo activo (entre inicio y fin)
            cursor.execute("""
                SELECT ciclo, periodo, fecha_inicio FROM calendario_institucional
                WHERE plan = %s AND tipo_evento = 'inicio_periodo' AND fecha_inicio <= %s
                ORDER BY fecha_inicio DESC LIMIT 1
            """, (plan, today))
            ultimo_inicio = cursor.fetchone()

            cursor.execute("""
                SELECT ciclo, periodo, fecha_fin FROM calendario_institucional
                WHERE plan = %s AND tipo_evento = 'fin_periodo' AND fecha_fin >= %s
                ORDER BY fecha_fin ASC LIMIT 1
            """, (plan, today))
            proximo_fin = cursor.fetchone()

            # Determinar si estamos en periodo activo
            en_periodo = False
            periodo_info = None
            if ultimo_inicio and proximo_fin:
                if ultimo_inicio['ciclo'] == proximo_fin['ciclo'] and ultimo_inicio['periodo'] == proximo_fin['periodo']:
                    en_periodo = True
                    periodo_info = {'ciclo': ultimo_inicio['ciclo'], 'periodo': ultimo_inicio['periodo']}

            # 3. Si no hay clases regulares (no en periodo) y TAMPOCO hay eventos hoy, entonces es receso
            if not en_periodo and not eventos_hoy:
                # Buscar próximo inicio de periodo
                cursor.execute("""
                    SELECT descripcion, fecha_inicio FROM calendario_institucional
                    WHERE plan = %s AND tipo_evento = 'inicio_periodo' AND fecha_inicio > %s
                    ORDER BY fecha_inicio ASC LIMIT 1
                """, (plan, today))
                proximo = cursor.fetchone()
                msg = f"Próximo inicio: {proximo['fecha_inicio']}" if proximo else "Sin periodo programado"
                return {
                    'hay_clases': False,
                    'estado': 'receso',
                    'descripcion': 'Periodo de receso entre semestres',
                    'detalle': msg,
                    'eventos_hoy': [],
                    'periodo_actual': None
                }

            hay_clases = True
            estado = 'clases'
            descripcion = 'Día de clases regular'

            # Las clases regulares finalizan cuando inician los exámenes ordinarios
            if en_periodo and periodo_info:
                cursor.execute("""
                    SELECT fecha_inicio FROM calendario_institucional
                    WHERE plan = %s AND ciclo = %s AND periodo = %s AND tipo_evento = 'examen_ordinario'
                    ORDER BY fecha_inicio ASC LIMIT 1
                """, (plan, periodo_info['ciclo'], periodo_info['periodo']))
                ord_inicio = cursor.fetchone()
                if ord_inicio and today >= ord_inicio['fecha_inicio']:
                    hay_clases = False
                    estado = 'examen_ordinario'
                    descripcion = 'Periodo de Evaluación Final'

            for ev in eventos_hoy:
                if ev['suspende_clases']:
                    hay_clases = False
                    estado = ev['tipo_evento']
                    descripcion = ev['descripcion']
                    break
                # Si hay eventos como exámenes, inscripciones, entregas, etc. actualizamos estado pero respetamos la suspensión de clases final
                if 'examen' in ev['tipo_evento'] or ev['tipo_evento'] in ['inscripcion', 'entrega', 'evaluacion']:
                    estado = ev['tipo_evento']
                    descripcion = ev['descripcion']

            return {
                'hay_clases': hay_clases,
                'estado': estado,
                'descripcion': descripcion,
                'eventos_hoy': eventos_hoy,
                'periodo_actual': periodo_info
            }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.get("/api/calendario-institucional")
def listar_calendario_institucional(plan: str = "semestral"):
    """Retorna todos los eventos del calendario institucional."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT id, plan, ciclo, periodo, tipo_evento, descripcion, fecha_inicio, fecha_fin, suspende_clases FROM calendario_institucional WHERE plan = %s ORDER BY fecha_inicio",
                (plan,)
            )
            datos = cursor.fetchall()
            for d in datos:
                d['fecha_inicio'] = str(d['fecha_inicio'])
                d['fecha_fin'] = str(d['fecha_fin'])
        return datos
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()

# ---------------------------------------------------------
# PROCESAMIENTO DE EVENTOS DEL CALENDARIO INSTITUCIONAL
# ---------------------------------------------------------

_MESES_MAP_PARSE = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
    'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
    'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12
}

def _clasificar_tipo_evento(texto: str) -> tuple:
    """Clasifica un texto descriptivo de evento en (tipo_evento, suspende_clases)."""
    t = texto.lower().strip()
    if any(k in t for k in ['inicio de subciclo', 'inicio de periodo', 'inicio de cuatrimestre', 'inicio de clases', 'inicio del ciclo']):
        return 'inicio_periodo', False
    if any(k in t for k in ['fin de subciclo', 'fin de periodo', 'fin de cuatrimestre', 'fin de clases', 'fin del ciclo']):
        return 'fin_periodo', False
    if any(k in t for k in ['vacaciones', 'semana santa', 'asueto', 'receso']):
        return 'vacaciones', True
    if any(k in t for k in ['inhábil', 'inhabil', 'festivo', 'día de muertos', 'dia de muertos',
                             'independencia', 'revolución', 'revolucion', 'constitución', 'constitucion',
                             'natalicio', 'carnaval', 'día del trabajo', 'dia del trabajo',
                             'día del maestro', 'dia del maestro']):
        return 'inhabil', True
    if any(k in t for k in ['parcial', '1er parcial', '2do parcial', '1ro.', '2do.', 'primer parcial', 'segundo parcial']):
        return 'examen_parcial', False
    if 'extraordinario' in t:
        return 'examen_extraordinario', False
    if any(k in t for k in ['ordinario', 'finales', 'examen final']):
        return 'examen_ordinario', False
    if any(k in t for k in ['inscripción', 'inscripcion', 'trámite', 'tramite', 'reinscripción', 'reinscripcion']):
        return 'inscripcion', False
    if any(k in t for k in ['entrega', 'reporte de actividades', 'plan de actividades']):
        return 'entrega', False
    if any(k in t for k in ['evaluación docente', 'evaluacion docente', 'eval. docente']):
        return 'evaluacion', False
    return 'otro', False


def _parse_fecha_texto(texto: str, anio_default: int) -> list:
    """Parsea texto de fecha en español -> lista de (fecha_inicio, fecha_fin)."""
    texto = texto.strip().lower().replace('de ', '').replace('del ', '')
    resultados = []
    anio_match = re.search(r'(\d{4})', texto)
    if anio_match:
        anio_default = int(anio_match.group(1))
        texto = texto.replace(anio_match.group(1), '').strip()

    m = re.match(r'(\d{1,2})\s+(\w+)\s+al\s+(\d{1,2})\s+(\w+)', texto)
    if m:
        d1, mes1, d2, mes2 = int(m.group(1)), m.group(2), int(m.group(3)), m.group(4)
        m1, m2 = _MESES_MAP_PARSE.get(mes1), _MESES_MAP_PARSE.get(mes2)
        if m1 and m2:
            a1 = anio_default + 1 if m1 <= 2 and m2 >= 8 else anio_default
            a2 = anio_default + 1 if m2 <= 2 and m1 >= 8 else anio_default
            try: return [(datetime(a1, m1, d1).date(), datetime(a2, m2, d2).date())]
            except ValueError: pass

    m = re.match(r'(\d{1,2})\s*(?:al|-)\s*(\d{1,2})\s+(\w+)', texto)
    if m:
        d1, d2, mes = int(m.group(1)), int(m.group(2)), m.group(3)
        mn = _MESES_MAP_PARSE.get(mes)
        if mn:
            try: return [(datetime(anio_default, mn, d1).date(), datetime(anio_default, mn, d2).date())]
            except ValueError: pass

    m = re.match(r'(\d{1,2})(?:\s*,\s*|\s+y\s+)(\d{1,2})\s+(\w+)', texto)
    if m:
        d1, d2, mes = int(m.group(1)), int(m.group(2)), m.group(3)
        mn = _MESES_MAP_PARSE.get(mes)
        if mn:
            try:
                resultados.append((datetime(anio_default, mn, d1).date(), datetime(anio_default, mn, d1).date()))
                resultados.append((datetime(anio_default, mn, d2).date(), datetime(anio_default, mn, d2).date()))
                return resultados
            except ValueError: pass

    partes = re.split(r'\s*,\s*', texto)
    if len(partes) > 1:
        for parte in partes:
            sub = _parse_fecha_texto(parte.strip(), anio_default)
            resultados.extend(sub)
        if resultados: return resultados

    m = re.match(r'(\d{1,2})\s+(\w+)', texto)
    if m:
        dia, mes = int(m.group(1)), m.group(2)
        mn = _MESES_MAP_PARSE.get(mes)
        if mn:
            try:
                d = datetime(anio_default, mn, dia).date()
                return [(d, d)]
            except ValueError: pass
    return resultados


def _deduplicar_eventos(eventos: list) -> list:
    """Combina eventos con la misma fecha_inicio+fecha_fin+plan en un solo registro."""
    agrupados = defaultdict(list)
    for ev in eventos:
        clave = (str(ev.get('fecha_inicio', '')), str(ev.get('fecha_fin', '')), ev.get('plan', ''))
        agrupados[clave].append(ev)
    resultado = []
    for clave, grupo in agrupados.items():
        if len(grupo) == 1:
            resultado.append(grupo[0])
        else:
            combinado = dict(grupo[0])
            descripciones, tipos, suspende = [], set(), False
            for ev in grupo:
                desc = ev.get('descripcion', '')
                if desc and desc not in descripciones: descripciones.append(desc)
                tipos.add(ev.get('tipo_evento', ''))
                if ev.get('suspende_clases'): suspende = True
            combinado['descripcion'] = ' | '.join(descripciones)
            if 'inhabil' in tipos: combinado['tipo_evento'] = 'inhabil'
            combinado['suspende_clases'] = suspende
            combinado['eventos_combinados'] = len(grupo)
            resultado.append(combinado)
    resultado.sort(key=lambda x: str(x.get('fecha_inicio', '')))
    return resultado


def _parsear_texto_eventos(texto_raw: str, plan: str, ciclo: str, periodo: int) -> list:
    """Parsea un bloque de texto libre con eventos del calendario institucional."""
    try: anio_base = int(ciclo.split('-')[0])
    except (ValueError, IndexError): anio_base = datetime.now().year
    eventos = []
    for linea in texto_raw.strip().split('\n'):
        linea = linea.strip()
        if not linea or len(linea) < 5: continue
        descripcion, fecha_texto = linea, ''
        for sep in [':', ' - ', ' – ']:
            if sep in linea:
                partes = linea.split(sep, 1)
                if len(partes) == 2 and re.search(r'\d', partes[1]):
                    descripcion, fecha_texto = partes[0].strip(), partes[1].strip()
                    break
        if not fecha_texto:
            fecha_match = re.search(r'(\d{1,2}\s+(?:de\s+)?\w+(?:\s+(?:al|-)?\s*\d{1,2}\s+(?:de\s+)?\w+)?)', linea, re.IGNORECASE)
            if fecha_match:
                fecha_texto = fecha_match.group(1)
                descripcion = linea[:fecha_match.start()].strip().rstrip(':').rstrip('-').strip()
        if not fecha_texto: continue
        tipo_evento, suspende = _clasificar_tipo_evento(descripcion)
        sub_eventos = fecha_texto.split('/')
        for sub in sub_eventos:
            sub = sub.strip()
            sub_clean = re.sub(r'^\d+(?:ro|do|er|to|vo)\.\s*', '', sub)
            fechas = _parse_fecha_texto(sub_clean, anio_base)
            sub_desc = descripcion
            prefix_match = re.match(r'(\d+(?:ro|do|er|to|vo)\.\s*)', sub)
            if prefix_match and len(sub_eventos) > 1:
                sub_desc = f"{descripcion} ({prefix_match.group(1).strip()})"
            for fecha_inicio, fecha_fin in fechas:
                eventos.append({'plan': plan, 'ciclo': ciclo, 'periodo': periodo,
                    'tipo_evento': tipo_evento, 'descripcion': sub_desc,
                    'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin, 'suspende_clases': suspende})
    return eventos


class EventoCalendarioInput(BaseModel):
    plan: str
    ciclo: str
    periodo: int
    tipo_evento: str
    descripcion: str
    fecha_inicio: str
    fecha_fin: str
    suspende_clases: bool = False

class TextoEventosInput(BaseModel):
    plan: str
    ciclo: str
    periodo: int
    texto: str


@app.post("/api/calendario-institucional/parsear-texto")
def parsear_texto_eventos_endpoint(datos: TextoEventosInput):
    """Parsea texto libre con eventos y retorna los eventos estructurados (sin guardar)."""
    try:
        eventos = _parsear_texto_eventos(datos.texto, datos.plan, datos.ciclo, datos.periodo)
        eventos_dedup = _deduplicar_eventos(eventos)
        for ev in eventos_dedup:
            ev['fecha_inicio'] = str(ev['fecha_inicio'])
            ev['fecha_fin'] = str(ev['fecha_fin'])
        return {"total_parseados": len(eventos), "total_deduplicados": len(eventos_dedup), "eventos": eventos_dedup}
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error al parsear texto: {str(e)}")


@app.post("/api/calendario-institucional/importar-texto")
def importar_texto_eventos(datos: TextoEventosInput):
    """Parsea texto libre y guarda los eventos en la DB (con deduplicacion)."""
    try:
        eventos = _parsear_texto_eventos(datos.texto, datos.plan, datos.ciclo, datos.periodo)
        if not eventos:
            raise HTTPException(status_code=400, detail="No se pudieron extraer eventos del texto proporcionado.")
        eventos_dedup = _deduplicar_eventos(eventos)
        connection = get_db_connection()
        insertados = 0
        try:
            with connection.cursor() as cursor:
                for ev in eventos_dedup:
                    cursor.execute("""SELECT id FROM calendario_institucional
                        WHERE plan=%s AND ciclo=%s AND periodo=%s AND fecha_inicio=%s AND fecha_fin=%s AND descripcion=%s LIMIT 1""",
                        (ev['plan'], ev['ciclo'], ev['periodo'], ev['fecha_inicio'], ev['fecha_fin'], ev['descripcion']))
                    if not cursor.fetchone():
                        cursor.execute("""INSERT INTO calendario_institucional (plan,ciclo,periodo,tipo_evento,descripcion,fecha_inicio,fecha_fin,suspende_clases)
                            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                            (ev['plan'], ev['ciclo'], ev['periodo'], ev['tipo_evento'], ev['descripcion'], ev['fecha_inicio'], ev['fecha_fin'], ev['suspende_clases']))
                        insertados += 1
            connection.commit()
            return {"message": f"Se importaron {insertados} eventos nuevos.", "insertados": insertados, "total_parseados": len(eventos), "total_deduplicados": len(eventos_dedup)}
        finally:
            connection.close()
    except HTTPException: raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/calendario-institucional/evento")
def crear_evento_institucional(evento: EventoCalendarioInput):
    """Crea un evento individual con deduplicacion automatica."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""SELECT id, descripcion FROM calendario_institucional
                WHERE plan=%s AND ciclo=%s AND periodo=%s AND fecha_inicio=%s AND fecha_fin=%s LIMIT 1""",
                (evento.plan, evento.ciclo, evento.periodo, evento.fecha_inicio, evento.fecha_fin))
            existente = cursor.fetchone()
            if existente:
                if evento.descripcion not in existente['descripcion']:
                    nueva_desc = f"{existente['descripcion']} | {evento.descripcion}"
                    cursor.execute("UPDATE calendario_institucional SET descripcion=%s WHERE id=%s", (nueva_desc, existente['id']))
                    connection.commit()
                    return {"message": "Evento existente actualizado (descripciones combinadas)", "id": existente['id'], "combinado": True}
                return {"message": "El evento ya existe", "id": existente['id'], "duplicado": True}
            cursor.execute("""INSERT INTO calendario_institucional (plan,ciclo,periodo,tipo_evento,descripcion,fecha_inicio,fecha_fin,suspende_clases)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s)""",
                (evento.plan, evento.ciclo, evento.periodo, evento.tipo_evento, evento.descripcion, evento.fecha_inicio, evento.fecha_fin, evento.suspende_clases))
        connection.commit()
        return {"message": "Evento creado exitosamente", "id": cursor.lastrowid}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.delete("/api/calendario-institucional/evento/{evento_id}")
def eliminar_evento_institucional(evento_id: int):
    """Elimina un evento del calendario institucional."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("DELETE FROM calendario_institucional WHERE id=%s", (evento_id,))
        connection.commit()
        return {"message": "Evento eliminado"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.get("/api/calendario-institucional/deduplicado")
def obtener_eventos_deduplicados(plan: str = "semestral", ciclo: str = ""):
    """Retorna los eventos del calendario institucional deduplicados por fecha."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            if ciclo:
                cursor.execute("SELECT id,plan,ciclo,periodo,tipo_evento,descripcion,fecha_inicio,fecha_fin,suspende_clases FROM calendario_institucional WHERE plan=%s AND ciclo=%s ORDER BY fecha_inicio", (plan, ciclo))
            else:
                cursor.execute("SELECT id,plan,ciclo,periodo,tipo_evento,descripcion,fecha_inicio,fecha_fin,suspende_clases FROM calendario_institucional WHERE plan=%s ORDER BY fecha_inicio", (plan,))
            datos = cursor.fetchall()
            for d in datos:
                d['fecha_inicio'] = str(d['fecha_inicio'])
                d['fecha_fin'] = str(d['fecha_fin'])
        return _deduplicar_eventos(datos)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()



@app.post("/api/registro")
def registrar_usuario(datos: RegistroUsuario):
    correo_limpio = datos.correo.strip().lower()
    if not correo_limpio.endswith("universidadlatino.edu.mx"):
        raise HTTPException(status_code=400, detail="El correo debe pertenecer a la Universidad Latino.")

    password_hasheada = bcrypt.hashpw(datos.password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT id FROM usuarios WHERE correo = %s", (correo_limpio,))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail="Este correo ya se encuentra registrado.")

            cursor.execute(
                "INSERT INTO usuarios (nombre, correo, turno, password, is_verified) VALUES (%s, %s, %s, %s, TRUE)",
                (datos.nombre.strip(), correo_limpio, datos.turno, password_hasheada)
            )
        connection.commit()
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error en la base de datos: {str(e)}")
    finally:
        connection.close()

    return {"message": "Usuario registrado correctamente. Ya puedes iniciar sesión."}


@app.delete("/api/usuarios/{usuario_id}")
def eliminar_usuario(usuario_id: int, admin_correo: str):
    correo_admin_permitido = "yenri.moo@alumno.universidadlatino.edu.mx"
    
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT correo FROM usuarios WHERE id = %s", (usuario_id,))
            user_a_eliminar = cursor.fetchone()
            
            if not user_a_eliminar:
                raise HTTPException(status_code=404, detail="Usuario no encontrado.")
                
            correo_eliminar = user_a_eliminar['correo']
            
            # Un usuario puede eliminarse a sí mismo. Para eliminar a otros, debe ser el admin.
            if admin_correo.lower() != correo_eliminar.lower() and admin_correo.lower() != correo_admin_permitido:
                raise HTTPException(status_code=403, detail="Solo el administrador puede eliminar a otros usuarios.")
                
            cursor.execute("DELETE FROM usuarios WHERE id = %s", (usuario_id,))
        connection.commit()
        return {"message": "Cuenta eliminada correctamente"}
    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al eliminar la cuenta: {str(e)}")
    finally:
        connection.close()


@app.put("/api/usuarios/{usuario_id}")
def actualizar_usuario(usuario_id: int, datos: ActualizarUsuario):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "UPDATE usuarios SET nombre = %s, turno = %s WHERE id = %s",
                (datos.nombre.strip(), datos.turno, usuario_id)
            )
        connection.commit()
        return {"message": "Perfil actualizado correctamente"}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al actualizar perfil: {str(e)}")
    finally:
        connection.close()


@app.get("/api/usuarios/{usuario_id}")
def obtener_usuario_por_id(usuario_id: int):
    """Devuelve los datos públicos de un usuario por ID (sin contraseña)."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT id, nombre, correo, turno, created_at FROM usuarios WHERE id = %s",
                (usuario_id,)
            )
            u = cursor.fetchone()
        if not u:
            raise HTTPException(status_code=404, detail="Usuario no encontrado.")
        return {
            'id':         u['id'],
            'nombre':     u['nombre'],
            'correo':     u['correo'],
            'turno':      u['turno'] or '',
            'created_at': u['created_at'].isoformat() if u.get('created_at') else None,
        }
    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error: {str(e)}")
    finally:
        connection.close()


@app.get("/api/usuarios")
def listar_usuarios():
    """Lista todos los usuarios registrados (sin datos sensibles)."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, nombre, correo, turno, is_verified, created_at
                FROM usuarios ORDER BY nombre
            """)
            usuarios = cursor.fetchall()
        result = []
        for u in (usuarios or []):
            result.append({
                'id':          u['id'],
                'nombre':      u['nombre'],
                'correo':      u['correo'],
                'turno':       u['turno'] or '',
                'is_verified': bool(u.get('is_verified', False)),
                'created_at':  u['created_at'].isoformat() if u.get('created_at') else None,
            })
        return result
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener usuarios: {str(e)}")
    finally:
        connection.close()


@app.put("/api/usuarios/{usuario_id}/cambiar-password")
def cambiar_password(usuario_id: int, datos: CambiarPassword):
    if len(datos.password_nueva) < 6:
        raise HTTPException(status_code=400, detail="La nueva contraseña debe tener al menos 6 caracteres.")
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT password FROM usuarios WHERE id = %s", (usuario_id,))
            usuario = cursor.fetchone()
            if not usuario:
                raise HTTPException(status_code=404, detail="Usuario no encontrado.")
            if not bcrypt.checkpw(datos.password_actual.encode('utf-8'), usuario['password'].encode('utf-8')):
                raise HTTPException(status_code=400, detail="La contraseña actual es incorrecta.")
            nueva_hash = bcrypt.hashpw(datos.password_nueva.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            cursor.execute("UPDATE usuarios SET password = %s WHERE id = %s", (nueva_hash, usuario_id))
        connection.commit()
        return {"message": "Contraseña actualizada correctamente."}
    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al cambiar contraseña: {str(e)}")
    finally:
        connection.close()


@app.post("/api/verificar-correo")
def verificar_correo(datos: VerificacionCorreo):
    correo_limpio = datos.correo.strip().lower()
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT id, verification_code, verification_code_expira, is_verified FROM usuarios WHERE correo = %s",
                (correo_limpio,)
            )
            usuario = cursor.fetchone()

            if not usuario:
                raise HTTPException(status_code=404, detail="Usuario no encontrado.")
            if usuario['is_verified']:
                return {"message": "El correo ya estaba verificado."}
            if usuario.get('verification_code_expira') and datetime.now() > usuario['verification_code_expira']:
                raise HTTPException(status_code=400, detail="El código ha expirado. Regístrate de nuevo para recibir un código actualizado.")
            if usuario['verification_code'] != datos.codigo.strip():
                raise HTTPException(status_code=400, detail="Código incorrecto. Inténtalo de nuevo.")

            cursor.execute(
                "UPDATE usuarios SET is_verified = TRUE, verification_code = NULL, verification_code_expira = NULL WHERE id = %s",
                (usuario['id'],)
            )
        connection.commit()
        return {"message": "Correo verificado correctamente. Ya puedes iniciar sesión."}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error en la base de datos: {str(e)}")
    finally:
        connection.close()


@app.post("/api/login")
def login_usuario(datos: LoginUsuario):
    correo_limpio = datos.correo.strip().lower()

    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT id, nombre, correo, turno, password, is_verified, created_at FROM usuarios WHERE correo = %s",
                (correo_limpio,)
            )
            usuario = cursor.fetchone()

            if not usuario:
                raise HTTPException(status_code=401, detail="Correo o contraseña incorrectos.")

            password_valida = bcrypt.checkpw(
                datos.password.encode('utf-8'),
                usuario['password'].encode('utf-8')
            )

            if not password_valida:
                raise HTTPException(status_code=401, detail="Correo o contraseña incorrectos.")

            return {
                "message": "Login exitoso",
                "usuario": {
                    "id":         usuario['id'],
                    "nombre":     usuario['nombre'],
                    "correo":     usuario['correo'],
                    "turno":      usuario['turno'],
                    "created_at": usuario['created_at'].isoformat() if usuario.get('created_at') else None,
                    "logged_at":  datetime.now().isoformat(),
                }
            }

    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error en la base de datos: {str(e)}")
    finally:
        connection.close()


# ---------------------------------------------------------
# ENDPOINT PARA PROCESAR PDF Y ARCHIVOS DE IMAGEN DE HORARIOS
# ---------------------------------------------------------
@app.post("/upload-pdf")
async def procesar_pdf(archivo: UploadFile = File(...)):
    """
    Parser inteligente que acepta PDF e imágenes (PNG, JPG).
    Para PDF: usa pdfplumber.
    Para imágenes: usa easyocr con detección de tabla por posición.
    """
    try:
        contenido = await archivo.read()
        
        # Validar tipo de archivo
        tipo_archivo = archivo.content_type
        if tipo_archivo not in ['application/pdf', 'image/png', 'image/jpeg', 'image/jpg']:
            raise ValueError("Formato no soportado. Usa PDF o imágenes (PNG, JPG).")
        
        licenciatura_extraida = "Licenciatura no identificada"
        semestre_extraido = ""
        cuatrimestre_extraido = ""
        grupo_extraido = ""
        mapa_docentes = {}
        horarios_compilados = []
        
        # ========== PROCESAR PDF ==========
        if tipo_archivo == 'application/pdf':
            pdf_file = BytesIO(contenido)
            
            with pdfplumber.open(pdf_file) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    texto_pagina = page.extract_text() or ""
                    for linea in texto_pagina.split('\n'):
                        linea_upper = linea.upper()
                        if "LICENCIATURA EN" in linea_upper or "LICENCIATURA" in linea_upper:
                            licenciatura_extraida = _normalizar_licenciatura(linea.strip())
                        match_sem = re.search(r'SEMESTRE:\s*([^\s]+)', linea_upper)
                        if match_sem: semestre_extraido = match_sem.group(1)
                        match_cuat = re.search(r'CUATRIMESTRE:\s*([^\s]+)', linea_upper)
                        if match_cuat: cuatrimestre_extraido = match_cuat.group(1)
                        match_grupo = re.search(r'GRUPO:\s*([^\s]+)', linea_upper)
                        if match_grupo: grupo_extraido = match_grupo.group(1)
                    
                    tables = page.extract_tables()
                    if not tables:
                        continue
                    
                    tabla_matriz = None
                    tabla_directorio = None
                    
                    for table in tables:
                        if not table or not table[0]: continue
                        fila_texto = " ".join([str(c).lower() for c in table[0] if c])
                        
                        if "lunes" in fila_texto or "martes" in fila_texto:
                            tabla_matriz = table
                        elif "asignatura" in fila_texto or "docente" in fila_texto:
                            tabla_directorio = table

                    if not tabla_directorio and len(tables) >= 2:
                        tabla_directorio = tables[-1]
                    if not tabla_matriz and len(tables) >= 1:
                        tabla_matriz = tables[0]

                    if not tabla_matriz or not tabla_directorio:
                        continue
                    
                    # Construir mapa de docentes
                    for row in tabla_directorio:
                        if not row or len(row) < 3: continue
                        asignatura_raw = str(row[0] or "").replace('\n', ' ').strip()
                        docente_raw = str(row[2] or "").replace('\n', ' ').strip()
                        
                        asignatura_raw = " ".join(asignatura_raw.split()).lower()
                        docente_raw = " ".join(docente_raw.split())
                        
                        if asignatura_raw and docente_raw and "asignatura" not in asignatura_raw:
                            mapa_docentes[asignatura_raw] = docente_raw
                    
                    # Procesar matriz de horarios
                    dias_indices = {}
                    start_row = 0
                    
                    for r_idx, row in enumerate(tabla_matriz):
                        fila_str = " ".join([str(c).lower() for c in row if c])
                        if "lunes" in fila_str:
                            start_row = r_idx + 1
                            for c_idx, cell in enumerate(row):
                                cell_val = str(cell or "").strip().lower().replace('\n', '')
                                for d in ["lunes", "martes", "miércoles", "miercoles", "jueves", "viernes"]:
                                    if d in cell_val:
                                        dia_oficial = d.replace('miercoles', 'miércoles').capitalize()
                                        dias_indices[dia_oficial] = c_idx
                            break

                    for row_num in range(start_row, len(tabla_matriz)):
                        row = tabla_matriz[row_num]
                        if not row or not row[0]: continue
                        
                        horario_slot = str(row[0] or "").replace('\n', '-').replace(' ', '').strip()
                        if len(horario_slot) < 5: continue
                        
                        for dia_nombre, idx_col in dias_indices.items():
                            if idx_col < len(row):
                                asignatura_celda = str(row[idx_col] or "").replace('\n', ' ').strip()
                                asignatura_celda = " ".join(asignatura_celda.split())
                                
                                if asignatura_celda and asignatura_celda.lower() not in ["", "sin especificar", "horario"]:
                                    key_busqueda = asignatura_celda.lower()
                                    docente_encontrado = mapa_docentes.get(key_busqueda, "Sin especificar")
                                    
                                    if docente_encontrado == "Sin especificar":
                                        for key_mapa, doc in mapa_docentes.items():
                                            if key_mapa in key_busqueda or key_busqueda in key_mapa:
                                                docente_encontrado = doc
                                                break
                                    
                                    horarios_compilados.append({
                                        "id": f"{page_num}_{row_num}_{idx_col}",
                                        "docente": docente_encontrado,
                                        "licenciatura": licenciatura_extraida,
                                        "asignatura": asignatura_celda.title(),
                                        "horario_resumen": f"{dia_nombre} {horario_slot}",
                                        "aula_asignada": "",
                                        "semestre": semestre_extraido,
                                        "cuatrimestre": cuatrimestre_extraido,
                                        "grupo": grupo_extraido
                                    })
        
        # ========== PROCESAR IMÁGENES ==========
        else:
            try:
                import numpy as np
                import easyocr
                from PIL import Image

                imagen = Image.open(BytesIO(contenido))
                imagen_np = np.array(imagen)

                # Usar reader con mejor configuración para imágenes
                reader = easyocr.Reader(['es', 'en'], gpu=False, verbose=False)
                resultados = reader.readtext(imagen_np)

                if not resultados:
                    raise ValueError("No se pudo extraer texto de la imagen. Verifica que sea una imagen clara.")

                # Construir lista de elementos con posición central (umbral de confianza más bajo)
                elementos = []
                for bbox, texto, conf in resultados:
                    if conf > 0.25 and texto.strip():  # Reducido de 0.3 a 0.25 para más tolerancia
                        y_centro = (bbox[0][1] + bbox[2][1]) / 2
                        x_centro = (bbox[0][0] + bbox[2][0]) / 2
                        elementos.append({"texto": texto.strip(), "x": x_centro, "y": y_centro, "conf": conf})

                if not elementos:
                    raise ValueError("No se extrajo texto suficiente de la imagen.")

                elementos.sort(key=lambda e: e["y"])

                # Extraer licenciatura, semestre, cuatrimestre, grupo
                for elem in elementos:
                    texto_upper = elem["texto"].upper()
                    if "LICENCIATURA" in texto_upper:
                        licenciatura_extraida = _normalizar_licenciatura(elem["texto"].strip())
                    match_sem = re.search(r'SEMESTRE:\s*([^\s]+)', texto_upper)
                    if match_sem: semestre_extraido = match_sem.group(1)
                    match_cuat = re.search(r'CUATRIMESTRE:\s*([^\s]+)', texto_upper)
                    if match_cuat: cuatrimestre_extraido = match_cuat.group(1)
                    match_grupo = re.search(r'GRUPO:\s*([^\s]+)', texto_upper)
                    if match_grupo: grupo_extraido = match_grupo.group(1)

                # Agrupar elementos en filas (aumentar tolerancia a 25px para mejor agrupación)
                filas = []
                fila_actual = []
                y_actual = None
                tolerancia = 25  # Aumentado de 15 a 25px
                for elem in elementos:
                    if y_actual is None or abs(elem["y"] - y_actual) <= tolerancia:
                        fila_actual.append(elem)
                        y_actual = elem["y"] if y_actual is None else (y_actual + elem["y"]) / 2
                    else:
                        if fila_actual:
                            filas.append(sorted(fila_actual, key=lambda e: e["x"]))
                        fila_actual = [elem]
                        y_actual = elem["y"]
                if fila_actual:
                    filas.append(sorted(fila_actual, key=lambda e: e["x"]))

                # Detectar columnas de días - MÁS FLEXIBLE
                dias_keywords = {
                    "lunes": "Lunes", "l": "Lunes",
                    "martes": "Martes", "m": "Martes", "ma": "Martes",
                    "miércoles": "Miércoles", "miercoles": "Miércoles", "x": "Miércoles", "mi": "Miércoles",
                    "jueves": "Jueves", "j": "Jueves", "ju": "Jueves",
                    "viernes": "Viernes", "v": "Viernes", "vi": "Viernes"
                }
                cols_dias = {}
                y_fila_dias = None

                for fila in filas:
                    for elem in fila:
                        texto_lower = elem["texto"].lower().strip()
                        # Solo aceptar palabras cortas o que coincidan exactamente para días
                        if len(texto_lower) <= 10:  # Los días no son textos largos
                            for d_key, d_val in dias_keywords.items():
                                if d_key in texto_lower and d_val not in cols_dias:
                                    cols_dias[d_val] = elem["x"]
                                    y_fila_dias = elem["y"]
                                    break

                # Construir mapa asignatura→docente desde la tabla de directorio al pie
                mapa_docentes_img = {}
                for fila in filas:
                    textos = [e["texto"] for e in fila]
                    tiene_numero = any(re.search(r"\d{1,3}", t) for t in textos)
                    # Buscar nombres propios (capitalizados)
                    nombres = [t for t in textos if len(t.split()) >= 2 and t[0].isupper() and not any(c.isdigit() for c in t)]
                    if tiene_numero and len(nombres) >= 1:
                        # Mapear el primer elemento con nombres
                        key = textos[0].lower() if textos else "unknown"
                        valor = nombres[-1] if nombres else "Sin especificar"
                        if key and key != "sin especificar":
                            mapa_docentes_img[key] = valor

                # OPCIÓN 1: Extraer horarios de la tabla de días si se detectaron
                if y_fila_dias is not None and cols_dias:
                    for fila in filas:
                        if not fila or fila[0]["y"] <= y_fila_dias:
                            continue
                        primer_texto = fila[0]["texto"]
                        # Buscar patrón de horario (HH:MM o HH-MM)
                        if not re.search(r"\d{1,2}[:\-]\d{2}", primer_texto):
                            continue
                        horario_slot = primer_texto

                        for elem in fila[1:]:
                            celda = elem["texto"].strip()
                            if not celda or len(celda) < 2:
                                continue
                            # No procesar si es hora
                            if re.search(r"\d{1,2}[:\-]\d{2}", celda):
                                continue

                            # Asignar día basado en la columna X más cercana
                            dia_asignado = min(cols_dias, key=lambda d: abs(cols_dias[d] - elem["x"]))

                            key_busqueda = celda.lower()
                            docente_img = mapa_docentes_img.get(key_busqueda, "Sin especificar")
                            if docente_img == "Sin especificar":
                                for k, v in mapa_docentes_img.items():
                                    if k in key_busqueda or key_busqueda in k:
                                        docente_img = v
                                        break

                            entrada_id = f"img_{dia_asignado}_{horario_slot}_{celda}"
                            ya_existe = any(
                                h["id"] == entrada_id and h["asignatura"].lower() == celda.lower()
                                for h in horarios_compilados
                            )
                            if not ya_existe:
                                horarios_compilados.append({
                                    "id": entrada_id,
                                    "docente": docente_img,
                                    "licenciatura": licenciatura_extraida,
                                    "asignatura": celda.title(),
                                    "horario_resumen": f"{dia_asignado} {horario_slot}",
                                    "aula_asignada": "",
                                    "semestre": semestre_extraido,
                                    "cuatrimestre": cuatrimestre_extraido,
                                    "grupo": grupo_extraido
                                })

                # OPCIÓN 2: Si no se estructuró bien, generar horarios de texto directo
                if not horarios_compilados and elementos:
                    contador = 0
                    for elem in elementos:
                        texto = elem["texto"].strip()
                        # Filtrar elementos muy cortos o que sean números solos
                        if len(texto) > 2 and not re.fullmatch(r"\d+", texto) and "licenciatura" not in texto.lower():
                            docente_encontrado = "Sin especificar"
                            # Buscar en el mapa si existe
                            for k, v in mapa_docentes_img.items():
                                if k in texto.lower() or texto.lower() in k:
                                    docente_encontrado = v
                                    break

                            horarios_compilados.append({
                                "id": f"img_fallback_{contador}",
                                "docente": docente_encontrado,
                                "licenciatura": licenciatura_extraida,
                                "asignatura": texto.title(),
                                "horario_resumen": "Verificar en imagen",
                                "aula_asignada": "",
                                "semestre": semestre_extraido,
                                "cuatrimestre": cuatrimestre_extraido,
                                "grupo": grupo_extraido
                            })
                            contador += 1
                            if contador >= 20:  # Máximo 20 fallback para evitar spam
                                break

                # OPCIÓN 3: Si aún no hay nada, al menos mostrar que se leyó texto
                if not horarios_compilados:
                    horarios_compilados.append({
                        "id": "img_0",
                        "docente": "Verificar manualmente",
                        "licenciatura": licenciatura_extraida,
                        "asignatura": f"Se detectó imagen. Se extrajeron {len(elementos)} elementos de texto. Verifica que el formato sea correcto.",
                        "horario_resumen": "Revisar imagen",
                        "aula_asignada": "",
                        "semestre": semestre_extraido,
                        "cuatrimestre": cuatrimestre_extraido,
                        "grupo": grupo_extraido
                    })

            except ImportError:
                horarios_compilados.append({
                    "id": "img_error",
                    "docente": "Error de Extracción",
                    "licenciatura": "No disponible",
                    "asignatura": "easyocr no instalado. Ejecuta: pip install easyocr numpy pillow",
                    "horario_resumen": "Instala la dependencia",
                    "aula_asignada": ""
                })
            except Exception as e:
                import traceback
                error_detail = str(e)
                print(f"Error procesando imagen: {error_detail}")
                print(traceback.format_exc())
                raise ValueError(f"Error procesando imagen: {error_detail}")
        
        if not horarios_compilados:
            raise ValueError("No se extrajeron horarios. Verifica la estructura del archivo.")
        
        return {
            "message": "Archivo procesado exitosamente",
            "datos_extraidos": {
                "lista_horarios": horarios_compilados
            }
        }
    
    except Exception as e:
        print(f"Error procesando archivo: {str(e)}")
        raise HTTPException(status_code=400, detail=f"Error al procesar archivo: {str(e)}")


@app.post("/api/recuperar-contrasena")
def recuperar_contrasena(datos: RecuperarContrasena):
    correo_limpio = datos.correo.strip().lower()
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT id, nombre FROM usuarios WHERE correo = %s", (correo_limpio,))
            usuario = cursor.fetchone()
            if not usuario:
                raise HTTPException(status_code=404, detail="No existe una cuenta con ese correo.")

            codigo = str(secrets.randbelow(900000) + 100000)
            expira_reset = datetime.now() + timedelta(minutes=15)
            cursor.execute("UPDATE usuarios SET reset_code = %s, reset_code_expira = %s WHERE id = %s", (codigo, expira_reset, usuario['id']))
        connection.commit()
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error en la base de datos: {str(e)}")
    finally:
        connection.close()

    try:
        html_reset = f"""
        <html><body style="font-family:Arial,sans-serif;background:#f4f6fb;padding:32px;">
          <div style="max-width:420px;margin:auto;background:#fff;border-radius:16px;
                      padding:36px;text-align:center;box-shadow:0 4px 20px rgba(0,0,0,.08);">
            <h2 style="color:#1c355e;margin-bottom:4px;">SIPREF</h2>
            <p style="color:#75777f;font-size:13px;">Universidad Latino</p>
            <p style="color:#44464e;">Hola <b>{usuario['nombre']}</b>, tu código para restablecer la contraseña es:</p>
            <div style="font-size:38px;font-weight:bold;letter-spacing:10px;color:#1c355e;
                        background:#f0f4ff;padding:18px;border-radius:10px;margin:24px 0;">
              {codigo}
            </div>
            <p style="color:#75777f;font-size:12px;">Caduca en 15 minutos. Si no solicitaste esto, ignora este correo.</p>
          </div>
        </body></html>
        """
        sender_email = os.getenv("SMTP_USER", "")
        smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
        smtp_port = int(os.getenv("SMTP_PORT", 587))
        smtp_pass = os.getenv("SMTP_PASSWORD", "") or os.getenv("SMTP_PASS", "")

        enviado_smtp = False
        if sender_email and smtp_pass:
            try:
                msg = MIMEMultipart()
                msg['From'] = sender_email
                msg['To'] = correo_limpio
                msg['Subject'] = "Restablecer contraseña — SIPREF ULA"
                msg.attach(MIMEText(html_reset, 'html'))
                with smtplib.SMTP(smtp_host, smtp_port) as server:
                    server.starttls()
                    server.login(sender_email, smtp_pass)
                    server.sendmail(sender_email, [correo_limpio], msg.as_string())
                enviado_smtp = True
            except Exception as smtp_err:
                print(f"SMTP fallo para restablecer contraseña, intentando EmailJS: {smtp_err}")

        if not enviado_smtp:
            _enviar_smtp(correo_limpio, "Restablecer contraseña — SIPREF ULA", html_reset)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"No se pudo enviar el correo: {str(e)}")

    return {"message": "Código enviado al correo."}


@app.post("/api/restablecer-contrasena")
def restablecer_contrasena(datos: RestablecerContrasena):
    correo_limpio = datos.correo.strip().lower()
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT id, reset_code, reset_code_expira FROM usuarios WHERE correo = %s", (correo_limpio,))
            usuario = cursor.fetchone()
            if not usuario:
                raise HTTPException(status_code=404, detail="Usuario no encontrado.")
            if usuario.get('reset_code_expira') and datetime.now() > usuario['reset_code_expira']:
                raise HTTPException(status_code=400, detail="El código ha expirado. Solicita un nuevo correo de recuperación.")
            if not usuario['reset_code'] or usuario['reset_code'] != datos.codigo.strip():
                raise HTTPException(status_code=400, detail="Código incorrecto.")

            if len(datos.nueva_password) < 6:
                raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres.")
            nueva_hash = bcrypt.hashpw(datos.nueva_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            cursor.execute(
                "UPDATE usuarios SET password = %s, reset_code = NULL, reset_code_expira = NULL WHERE id = %s",
                (nueva_hash, usuario['id'])
            )
        connection.commit()
        return {"message": "Contraseña actualizada correctamente."}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error en la base de datos: {str(e)}")
    finally:
        connection.close()


# ---------------------------------------------------------
# ENDPOINTS PARA GESTIÓN DE HORARIOS
# ---------------------------------------------------------
def _verificar_conflicto_aula(cursor, aula: str, horario_str: str, excluir_id: int = None) -> str | None:
    """Retorna mensaje de conflicto si el aula ya está ocupada en ese rango horario, o None si no hay conflicto."""
    if not aula or aula in ("Por asignar", ""):
        return None
    dia_idx, inicio, fin = _parse_horario_minutos(horario_str)
    if dia_idx is None or inicio is None or fin is None:
        return None
    query = "SELECT horario, asignatura, docente FROM horarios WHERE aula_asignada = %s AND TRIM(aula_asignada) != '' AND TRIM(aula_asignada) != 'Por asignar'"
    params = [aula]
    if excluir_id is not None:
        query += " AND id != %s"
        params.append(excluir_id)
    cursor.execute(query, params)
    existentes = cursor.fetchall()
    for ex in (existentes or []):
        e_dia, e_inicio, e_fin = _parse_horario_minutos(ex.get('horario', ''))
        if e_dia != dia_idx or e_inicio is None or e_fin is None:
            continue
        # Hay solapamiento si NO se cumple: e_fin <= inicio OR e_inicio >= fin
        if not (e_fin <= inicio or e_inicio >= fin):
            return (f"El aula '{aula}' ya está ocupada en ese horario: "
                    f"{ex['asignatura']} con {ex['docente']} ({ex['horario']})")
    return None


@app.post("/api/guardar-horarios")
def guardar_horarios(horarios: list[dict], request: Request):
    connection = get_db_connection()
    usuario = obtener_usuario(request)
    try:
        with connection.cursor() as cursor:
            # Verificar conflictos en bloque para mayor velocidad
            aulas_involucradas = list(set(h.get("aulaAsignada", h.get("aula_asignada", "Por asignar")) for h in horarios))
            aulas_filtradas = [a for a in aulas_involucradas if a not in ("Por asignar", "")]
            
            existentes_por_aula = {}
            if aulas_filtradas:
                format_strings = ','.join(['%s'] * len(aulas_filtradas))
                query = f"SELECT aula_asignada, horario, asignatura, docente FROM horarios WHERE aula_asignada IN ({format_strings}) AND TRIM(aula_asignada) != '' AND TRIM(aula_asignada) != 'Por asignar'"
                cursor.execute(query, tuple(aulas_filtradas))
                for row in cursor.fetchall():
                    aula = row['aula_asignada']
                    if aula not in existentes_por_aula:
                        existentes_por_aula[aula] = []
                    existentes_por_aula[aula].append(row)

            conflictos = []
            for h in horarios:
                aula = h.get("aulaAsignada", h.get("aula_asignada", "Por asignar"))
                if aula in ("Por asignar", ""):
                    continue
                horario_str = h.get("horario", h.get("horario_resumen", ""))
                dia_idx, inicio, fin = _parse_horario_minutos(horario_str)
                if dia_idx is None or inicio is None or fin is None:
                    continue
                
                for ex in existentes_por_aula.get(aula, []):
                    e_dia, e_inicio, e_fin = _parse_horario_minutos(ex.get('horario', ''))
                    if e_dia != dia_idx or e_inicio is None or e_fin is None:
                        continue
                    if not (e_fin <= inicio or e_inicio >= fin):
                        msg = (f"El aula '{aula}' ya está ocupada en ese horario: "
                               f"{ex['asignatura']} con {ex['docente']} ({ex['horario']})")
                        if msg not in conflictos:
                            conflictos.append(msg)
                            
            if conflictos:
                raise HTTPException(status_code=409, detail=conflictos[0])

            for h in horarios:
                sql = """
                    INSERT INTO horarios
                    (docente, licenciatura, asignatura, horario, aula_asignada, archivo, fecha_creacion, fecha_clase, semestre, cuatrimestre, grupo)
                    VALUES (%s, %s, %s, %s, %s, %s, NOW(), CURDATE(), %s, %s, %s)
                """
                cursor.execute(sql, (
                    h.get("docente", ""),
                    h.get("licenciatura", ""),
                    h.get("asignatura", ""),
                    h.get("horario", h.get("horario_resumen", "")),
                    h.get("aulaAsignada", h.get("aula_asignada", "Por asignar")),
                    h.get("archivo", ""),
                    h.get("semestre", ""),
                    h.get("cuatrimestre", ""),
                    h.get("grupo", "")
                ))
                horario_id = cursor.lastrowid
                registrar_bitacora(
                    cursor, usuario, 'crear_clase', 'horarios', registro_id=horario_id,
                    docente=h.get("docente"), asignatura=h.get("asignatura"),
                    licenciatura=h.get("licenciatura"), 
                    grado_cuatrimestre=h.get("semestre") or h.get("cuatrimestre"),
                    grupo=h.get("grupo"), 
                    hora_nueva=h.get("horario", h.get("horario_resumen", "")),
                    aula_nueva=h.get("aulaAsignada", h.get("aula_asignada", "Por asignar"))
                )
        connection.commit()
        return {"message": "Horarios guardados"}
    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al guardar horarios: {str(e)}")
    finally:
        connection.close()


# ---------------------------------------------------------
# ENDPOINTS PARA ARCHIVOS (agrupados por nombre de archivo)
# ---------------------------------------------------------
@app.get("/api/archivos")
def listar_archivos():
    """Devuelve la lista de archivos cargados con su conteo de horarios y metadatos."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    archivo,
                    COUNT(*) AS total_horarios,
                    SUM(CASE WHEN aula_asignada IS NOT NULL AND TRIM(aula_asignada) != '' AND TRIM(aula_asignada) != 'Por asignar' THEN 1 ELSE 0 END) AS aulas_asignadas,
                    MAX(licenciatura) AS licenciatura,
                    MAX(semestre) AS semestre,
                    MAX(cuatrimestre) AS cuatrimestre,
                    MAX(grupo) AS grupo,
                    MAX(fecha_creacion) AS fecha_carga,
                    CASE
                        WHEN MAX(semestre) IS NOT NULL AND MAX(semestre) != '' THEN 'Semestral'
                        WHEN MAX(cuatrimestre) IS NOT NULL AND MAX(cuatrimestre) != '' THEN 'Cuatrimestral'
                        ELSE 'No definido'
                    END AS plan,
                    GROUP_CONCAT(DISTINCT CASE WHEN aula_asignada IS NOT NULL AND TRIM(aula_asignada) != '' AND TRIM(aula_asignada) != 'Por asignar' THEN aula_asignada END ORDER BY aula_asignada SEPARATOR ', ') AS aulas_ocupadas,
                    MIN(SUBSTRING_INDEX(SUBSTRING_INDEX(horario, ' ', -1), '-', 1)) AS rango_inicio,
                    MAX(SUBSTRING_INDEX(horario, '-', -1)) AS rango_fin
                FROM horarios
                WHERE archivo IS NOT NULL AND TRIM(archivo) != ''
                GROUP BY archivo
                ORDER BY MAX(fecha_creacion) DESC
            """)
            archivos = cursor.fetchall()
            # Construir rango horario y turno
            for a in (archivos or []):
                inicio = a.pop('rango_inicio', '')
                fin = a.pop('rango_fin', '')
                if inicio and fin:
                    a['rango_horario'] = f"{inicio} - {fin}"
                    try:
                        hora = int(inicio.split(':')[0])
                        if hora < 14:
                            a['turno'] = 'Matutino'
                        else:
                            a['turno'] = 'Vespertino'
                    except:
                        a['turno'] = None
                else:
                    a['rango_horario'] = None
                    a['turno'] = None
            return archivos or []
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.get("/api/archivos/{nombre_archivo}/horarios")
def obtener_horarios_por_archivo(nombre_archivo: str):
    """Devuelve los horarios pertenecientes a un archivo específico."""
    from urllib.parse import unquote
    nombre_archivo = unquote(nombre_archivo)
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, docente, licenciatura, asignatura, horario, 
                       aula_asignada, archivo, semestre, cuatrimestre, grupo,
                       fecha_creacion
                FROM horarios
                WHERE archivo = %s
                ORDER BY docente, horario
            """, (nombre_archivo,))
            horarios = cursor.fetchall()
            return horarios or []
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.delete("/api/archivos/{nombre_archivo}")
def eliminar_archivo(nombre_archivo: str, request: Request):
    """Elimina todos los horarios asociados a un archivo."""
    from urllib.parse import unquote
    nombre_archivo = unquote(nombre_archivo)
    connection = get_db_connection()
    usuario = obtener_usuario(request)
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) as total FROM horarios WHERE archivo = %s", (nombre_archivo,))
            count = cursor.fetchone()
            if not count or count['total'] == 0:
                raise HTTPException(status_code=404, detail="Archivo no encontrado")
            
            registrar_bitacora(
                cursor, usuario, 'eliminar_archivo', 'horarios',
                motivo=f"Eliminacion de archivo: {nombre_archivo} ({count['total']} horarios)"
            )
            
            cursor.execute("DELETE FROM horarios WHERE archivo = %s", (nombre_archivo,))
            connection.commit()
            return {"message": f"Archivo '{nombre_archivo}' eliminado con {count['total']} horarios."}
    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.get("/api/horarios")
def obtener_horarios():
    """
    Devuelve el catálogo completo de horarios programados (semana recurrente).
    Usa GROUP BY para deduplicar entradas idénticas sin filtrar por fecha de carga,
    igual que /api/clases-hoy, para que el dashboard muestre exactamente los mismos
    registros que Gestión de Docentes.
    """
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT
                    MIN(h.id)           AS id,
                    h.docente,
                    h.licenciatura,
                    h.asignatura,
                    h.horario,
                    h.aula_asignada,
                    MIN(h.archivo)      AS archivo,
                    MAX(h.fecha_creacion) AS fecha_creacion,
                    MAX(h.semestre)     AS semestre,
                    MAX(h.cuatrimestre) AS cuatrimestre,
                    MAX(h.grupo)        AS grupo,
                    COALESCE(MAX(h.estado_slug), 'programada') AS estado_slug,
                    MAX(h.nota_reprogramacion) AS nota_reprogramacion,
                    MAX(h.fecha_reposicion) AS fecha_reposicion,
                    MAX(h.es_reposicion) AS es_reposicion,
                    a.en_mantenimiento,
                    a.inicio_mantenimiento,
                    a.fin_mantenimiento,
                    a.aula_temporal
                FROM horarios h
                LEFT JOIN aulas a ON a.nombre = h.aula_asignada
                WHERE h.docente IS NOT NULL
                  AND TRIM(h.docente) != ''
                  AND TRIM(h.docente) != 'Sin especificar'
                GROUP BY
                    h.docente, h.licenciatura, h.asignatura,
                    h.horario, h.aula_asignada,
                    a.en_mantenimiento, a.inicio_mantenimiento, a.fin_mantenimiento, a.aula_temporal
                ORDER BY h.horario
                LIMIT 1000
            """)
            horarios = cursor.fetchall()
        return _aplicar_mantenimiento(horarios) if horarios else []
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener horarios: {str(e)}")
    finally:
        connection.close()


@app.get("/api/clases-hoy")
def obtener_clases_hoy(dia: Optional[int] = None, mins: Optional[int] = None, fecha: Optional[str] = None):
    """
    Clases EN CURSO ahora mismo, respetando el calendario académico por plan.
    dia   = getDay() JS (0=Dom…6=Sab). Si se omite, usa el día real del servidor.
    mins  = minutos desde medianoche. Si se omite, usa la hora real del servidor.
    fecha = fecha ISO (YYYY-MM-DD). Para consultar eventos del calendario institucional.
    """
    ahora = datetime.now()
    dia_hoy    = dia  if dia  is not None else (ahora.isoweekday() % 7)
    mins_ahora = mins if mins is not None else (ahora.hour * 60 + ahora.minute)

    # Domingo → no hay clases regulares
    if dia_hoy == 0:
        return []

    if fecha:
        try:
            date_obj = datetime.strptime(fecha, "%Y-%m-%d").date()
        except Exception:
            date_obj = ahora.date()
    else:
        date_obj = ahora.date()

    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            # ── 1) Obtener estado académico de cada plan ──────────────────
            hay_clases_plan = {}
            for plan in ('semestral', 'cuatrimestral'):
                # Verificar suspensión de clases por evento
                cursor.execute("""
                    SELECT tipo_evento, suspende_clases FROM calendario_institucional
                    WHERE plan = %s AND fecha_inicio <= %s AND fecha_fin >= %s
                    ORDER BY suspende_clases DESC
                """, (plan, date_obj, date_obj))
                eventos = cursor.fetchall()

                clases_ok = True

                # Si algún evento suspende clases → sin clases para ese plan
                for ev in eventos:
                    if ev.get('suspende_clases'):
                        clases_ok = False
                        break

                # Verificar si estamos en periodo activo
                if clases_ok:
                    cursor.execute("""
                        SELECT ciclo, periodo FROM calendario_institucional
                        WHERE plan = %s AND tipo_evento = 'inicio_periodo' AND fecha_inicio <= %s
                        ORDER BY fecha_inicio DESC LIMIT 1
                    """, (plan, date_obj))
                    ultimo_inicio = cursor.fetchone()

                    cursor.execute("""
                        SELECT ciclo, periodo FROM calendario_institucional
                        WHERE plan = %s AND tipo_evento = 'fin_periodo' AND fecha_fin >= %s
                        ORDER BY fecha_fin ASC LIMIT 1
                    """, (plan, date_obj))
                    proximo_fin = cursor.fetchone()

                    en_periodo = False
                    periodo_info = None
                    if ultimo_inicio and proximo_fin:
                        if ultimo_inicio['ciclo'] == proximo_fin['ciclo'] and ultimo_inicio['periodo'] == proximo_fin['periodo']:
                            en_periodo = True
                            periodo_info = {'ciclo': ultimo_inicio['ciclo'], 'periodo': ultimo_inicio['periodo']}

                    if not en_periodo and not eventos:
                        clases_ok = False  # Receso

                    # Verificar si estamos en periodo de exámenes ordinarios/extraordinarios
                    if en_periodo and periodo_info:
                        cursor.execute("""
                            SELECT fecha_inicio FROM calendario_institucional
                            WHERE plan = %s AND ciclo = %s AND periodo = %s
                              AND (tipo_evento = 'examen_ordinario' OR tipo_evento = 'examen_extraordinario')
                            ORDER BY fecha_inicio ASC LIMIT 1
                        """, (plan, periodo_info['ciclo'], periodo_info['periodo']))
                        ord_inicio = cursor.fetchone()
                        if ord_inicio and date_obj >= ord_inicio['fecha_inicio']:
                            clases_ok = "examenes"  # Periodo de evaluación final

                hay_clases_plan[plan] = clases_ok

            # Si NINGÚN plan tiene clases hoy → lista vacía
            if not hay_clases_plan['semestral'] and not hay_clases_plan['cuatrimestral']:
                return []

            # ── 1b) Obtener aulas liberadas manualmente para esta fecha ─────
            cursor.execute("""
                SELECT aula_nombre
                FROM aulas_liberadas
                WHERE fecha = %s AND activa = TRUE
            """, (date_obj,))
            liberadas_rows = cursor.fetchall()
            liberadas_set = {r['aula_nombre'] for r in (liberadas_rows or [])}

            # ── 2) Obtener clases con licenciatura, semestre, cuatrimestre y grupo ─
            cursor.execute("""
                SELECT h.docente, h.licenciatura, h.asignatura, h.horario, h.aula_asignada,
                       MAX(NULLIF(TRIM(h.semestre), '')) as semestre,
                       MAX(NULLIF(TRIM(h.cuatrimestre), '')) as cuatrimestre,
                       MAX(NULLIF(TRIM(h.grupo), '')) as grupo,
                       COALESCE(MAX(h.estado_slug), 'programada') AS estado_slug,
                       MAX(h.nota_reprogramacion) AS nota_reprogramacion,
                       MAX(h.fecha_reposicion) AS fecha_reposicion,
                       MAX(h.es_reposicion) AS es_reposicion,
                       a.en_mantenimiento, a.inicio_mantenimiento, a.fin_mantenimiento, a.aula_temporal
                FROM horarios h
                LEFT JOIN aulas a ON a.nombre = h.aula_asignada
                WHERE h.docente IS NOT NULL AND TRIM(h.docente) != ''
                GROUP BY h.docente, h.licenciatura, h.asignatura, h.horario, h.aula_asignada,
                         a.en_mantenimiento, a.inicio_mantenimiento, a.fin_mantenimiento, a.aula_temporal
            """)
            todas = cursor.fetchall()

            # Función de normalización robusta para comparar materias
            def normalize_subject(s: str) -> str:
                if not s: return ""
                import unicodedata
                s = s.strip().lower()
                s = ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')
                words = []
                for w in s.split():
                    if len(w) > 3 and w.endswith('s'):
                        words.append(w[:-1])
                    else:
                        words.append(w)
                return ' '.join(words)

            # Si hay algún plan en estado "examenes", necesitamos consultar qué exámenes hay hoy
            if "examenes" in hay_clases_plan.values():
                meses_nombres = ['', 'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                                 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
                dias_semana = ['lunes', 'martes', 'miércoles', 'jueves', 'viernes', 'sábado', 'domingo']
                dia_str = dias_semana[dia_hoy - 1] if 1 <= dia_hoy <= 7 else ''
                # Formato "07 de agosto" (como lo almacena el parser del PDF)
                date_str_texto = f"{date_obj.day:02d} de {meses_nombres[date_obj.month]}"
                date_str_iso = date_obj.strftime("%Y-%m-%d")
                date_str_ddmmyyyy = date_obj.strftime("%d/%m/%Y")
                
                cursor.execute("""
                    SELECT materia FROM examenes_calendario 
                    WHERE LOWER(fecha) = LOWER(%s)
                       OR LOWER(fecha) = LOWER(%s)
                       OR LOWER(fecha) = LOWER(%s)
                       OR (TRIM(fecha) = '' AND LOWER(dia) = LOWER(%s))
                """, (date_str_texto, date_str_iso, date_str_ddmmyyyy, dia_str))
                examenes_hoy_materias = {normalize_subject(row['materia']) for row in cursor.fetchall()}
            else:
                examenes_hoy_materias = set()

        # Agrupar bloques consecutivos de la misma clase (ej. 17:10-18:00 y 18:00-18:50 -> 17:10-18:50)
        mapa_grupos = defaultdict(list)
        for clase in (todas or []):
            # Filtrar por estado académico del plan
            plan_clase = 'cuatrimestral' if clase.get('cuatrimestre') else 'semestral'
            estado_plan = hay_clases_plan.get(plan_clase, True)
            if not estado_plan: continue
            if estado_plan == "examenes":
                clase_asig_norm = normalize_subject(clase.get('asignatura', ''))
                match_examen = False
                for ex_mat in examenes_hoy_materias:
                    if len(ex_mat) >= 3 and (ex_mat in clase_asig_norm or clase_asig_norm in ex_mat):
                        match_examen = True
                        break
                if not match_examen:
                    continue

            if clase.get('es_reposicion'):
                if str(clase.get('fecha_reposicion')) != str(date_obj):
                    continue
            dia_idx, inicio, fin = _parse_horario_minutos(clase.get('horario', ''))
            if dia_idx != dia_hoy or inicio is None or fin is None:
                continue
            key = (clase.get('docente'), clase.get('licenciatura'), clase.get('asignatura'), clase.get('aula_asignada', ''))
            item = dict(clase)
            item['_inicio'] = inicio
            item['_fin'] = fin
            mapa_grupos[key].append(item)

        clases_agrupadas = []
        for key, lista in mapa_grupos.items():
            lista.sort(key=lambda c: c['_inicio'])
            actual = dict(lista[0])
            for i in range(1, len(lista)):
                sig = lista[i]
                if sig['_inicio'] - actual['_fin'] <= 10:
                    actual['_fin'] = max(actual['_fin'], sig['_fin'])
                else:
                    clases_agrupadas.append(actual)
                    actual = dict(sig)
            clases_agrupadas.append(actual)

        en_curso = []
        nombres_dias = ["Domingo", "Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]
        for clase in clases_agrupadas:
            inicio = clase['_inicio']
            fin = clase['_fin']
            if not (inicio <= mins_ahora <= fin):
                continue

            # Formatear horario de bloque continuo (ej. "Lunes 17:10---18:50")
            h_ini = f"{inicio // 60:02d}:{inicio % 60:02d}"
            h_fin = f"{fin // 60:02d}:{fin % 60:02d}"
            dia_str = nombres_dias[dia_hoy]
            clase['horario'] = f"{dia_str} {h_ini}---{h_fin}"

            # Omitir si el aula fue liberada manualmente hoy
            if clase.get('aula_asignada') in liberadas_set:
                continue

            # Determinar plan de la clase
            es_cuatri = bool(clase.get('cuatrimestre'))
            plan_clase = 'cuatrimestral' if es_cuatri else 'semestral'

            # Solo incluir si el plan de esta clase tiene clases hoy
            if hay_clases_plan.get(plan_clase, False):
                en_curso.append(clase)

        return _aplicar_mantenimiento(en_curso)
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener clases: {str(e)}")
    finally:
        connection.close()


class LiberarAulaRequest(BaseModel):
    aula_nombre: str
    fecha: Optional[str] = None
    docente: Optional[str] = ""
    asignatura: Optional[str] = ""
    horario: Optional[str] = ""
    motivo: Optional[str] = "Liberación manual"


@app.post("/api/aulas/liberar")
def liberar_aula(datos: LiberarAulaRequest, request: Request):
    usuario = obtener_usuario(request)
    ahora = datetime.now()
    fecha_val = datos.fecha if datos.fecha else ahora.strftime("%Y-%m-%d")
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE aulas_liberadas
                SET activa = FALSE
                WHERE aula_nombre = %s AND fecha = %s
            """, (datos.aula_nombre, fecha_val))

            cursor.execute("""
                INSERT INTO aulas_liberadas (aula_nombre, fecha, docente, asignatura, horario, motivo, activa)
                VALUES (%s, %s, %s, %s, %s, %s, TRUE)
            """, (datos.aula_nombre, fecha_val, datos.docente or "", datos.asignatura or "", datos.horario or "", datos.motivo or "Liberación manual"))
            
            registrar_bitacora(
                cursor, usuario, 'liberar_aula', 'aulas', 
                aula_anterior=datos.aula_nombre, aula_nueva=None,
                docente=datos.docente, asignatura=datos.asignatura,
                hora_anterior=datos.horario, motivo=datos.motivo
            )

            # Registrar snapshot inmediato en clases_historico
            date_obj = datetime.strptime(fecha_val, "%Y-%m-%d").date()
            _guardar_registro_historico(
                cursor,
                clase_id=None,
                fecha_obj=date_obj,
                docente_nombre=datos.docente or 'Docente',
                licenciatura_codigo='',
                semestre='',
                grupo='',
                asignatura_nombre=datos.asignatura or 'Clase',
                horario_inicio_str='00:00:00',
                horario_fin_str='23:59:59',
                aula_original=datos.aula_nombre,
                aula_actual='Fuera de Aula',
                estado_slug='aula_liberada',
                estado_label='Aula Liberada',
                incidencia_detalle=datos.motivo or 'Liberación manual'
            )
            connection.commit()
        return {"message": f"Aula {datos.aula_nombre} liberada con éxito"}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al liberar aula: {str(e)}")
    finally:
        connection.close()


@app.get("/api/aulas/liberadas")
def obtener_aulas_liberadas(fecha: Optional[str] = None):
    ahora = datetime.now()
    fecha_val = fecha if fecha else ahora.strftime("%Y-%m-%d")
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, aula_nombre, fecha, docente, asignatura, horario, motivo, created_at
                FROM aulas_liberadas
                WHERE fecha = %s AND activa = TRUE
                ORDER BY id DESC
            """, (fecha_val,))
            return cursor.fetchall() or []
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener liberaciones: {str(e)}")
    finally:
        connection.close()


@app.delete("/api/aulas/liberar/{liberacion_id}")
def reactivar_aula(liberacion_id: int, request: Request):
    usuario = obtener_usuario(request)
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT aula_nombre FROM aulas_liberadas WHERE id = %s", (liberacion_id,))
            row = cursor.fetchone()
            if row:
                registrar_bitacora(
                    cursor, usuario, 'reactivar_aula', 'aulas',
                    registro_id=liberacion_id, aula_nueva=row['aula_nombre']
                )
            cursor.execute("UPDATE aulas_liberadas SET activa = FALSE WHERE id = %s", (liberacion_id,))
            connection.commit()
        return {"message": "Liberación cancelada y aula reactivada"}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al reactivar aula: {str(e)}")
    finally:
        connection.close()


# ---------------------------------------------------------
# ENDPOINTS PARA GESTIÓN DE AULAS
# ---------------------------------------------------------
@app.get("/api/aulas/ocupacion")
def obtener_ocupacion_aulas():
    """
    Calcula ocupación de cada aula basándose en sus horarios asignados en la BD.
    Matutino  : inicio < 14:00 (840 min).
    Vespertino: inicio >= 14:00.
    No depende de la hora actual — refleja la semana completa.
    Incluye lista de carreras (licenciaturas) con sus horarios por aula.
    """
    LIMITE_VESPERTINO = 14 * 60  # 840 min = 14:00
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT aula_asignada, horario, licenciatura, semestre, cuatrimestre
                FROM horarios
                WHERE aula_asignada IS NOT NULL
                  AND TRIM(aula_asignada) != ''
                  AND TRIM(aula_asignada) != 'Por asignar'
            """)
            registros = cursor.fetchall()

        ocupacion: dict[str, dict] = {}
        for r in (registros or []):
            aula = r['aula_asignada']
            _, inicio, _ = _parse_horario_minutos(r.get('horario', ''))
            if inicio is None:
                continue
            if aula not in ocupacion:
                ocupacion[aula] = {'matutino': False, 'vespertino': False, 'carreras': []}
            if inicio < LIMITE_VESPERTINO:
                ocupacion[aula]['matutino'] = True
            else:
                ocupacion[aula]['vespertino'] = True

            # Agregar info de carrera, grado y horario
            lic = (r.get('licenciatura') or '').strip()
            horario = (r.get('horario') or '').strip()
            semestre = (r.get('semestre') or '').strip()
            cuatrimestre = (r.get('cuatrimestre') or '').strip()
            
            grado = semestre if semestre else cuatrimestre
            if not grado:
                grado = 'N/A'

            if lic:
                # Verificar si ya existe esta combinación
                ya_existe = any(
                    c['licenciatura'] == lic and c['horario'] == horario and c.get('grado', 'N/A') == grado
                    for c in ocupacion[aula]['carreras']
                )
                if not ya_existe:
                    ocupacion[aula]['carreras'].append({
                        'licenciatura': lic,
                        'horario': horario,
                        'grado': grado
                    })

        return ocupacion
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener ocupación: {str(e)}")
    finally:
        connection.close()


@app.get("/api/aulas")
def obtener_aulas():
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE aulas 
                SET en_mantenimiento = 0, inicio_mantenimiento = NULL, fin_mantenimiento = NULL, aula_temporal = NULL 
                WHERE en_mantenimiento = 1 AND fin_mantenimiento IS NOT NULL AND fin_mantenimiento <= NOW()
            """)
            connection.commit()
            cursor.execute("""
                SELECT id, nombre, edificio, capacidad, equipos, estado,
                       en_mantenimiento, inicio_mantenimiento, fin_mantenimiento, aula_temporal
                FROM aulas ORDER BY nombre
            """)
            aulas = cursor.fetchall()
        for aula in (aulas or []):
            # Parsear equipos de JSON string a lista
            if aula.get('equipos'):
                try:
                    aula['equipos'] = json.loads(aula['equipos'])
                except Exception:
                    aula['equipos'] = []
            # Serializar datetimes para JSON
            if aula.get('inicio_mantenimiento') and isinstance(aula['inicio_mantenimiento'], datetime):
                aula['inicio_mantenimiento'] = aula['inicio_mantenimiento'].isoformat()
            if aula.get('fin_mantenimiento') and isinstance(aula['fin_mantenimiento'], datetime):
                aula['fin_mantenimiento'] = aula['fin_mantenimiento'].isoformat()
        return aulas if aulas else []
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener aulas: {str(e)}")
    finally:
        connection.close()


@app.post("/api/aulas")
def crear_aula(aula: Aula, request: Request):
    """
    Crea una nueva aula
    """
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            usuario = obtener_usuario(request)
            # Convertir lista de equipos a JSON string
            equipos_json = json.dumps(aula.equipos)
            
            sql = "INSERT INTO aulas (nombre, edificio, capacidad, equipos, estado) VALUES (%s, %s, %s, %s, %s)"
            cursor.execute(sql, (
                aula.nombre, 
                aula.edificio, 
                aula.capacidad,
                equipos_json,
                aula.estado
            ))
            aula_id = cursor.lastrowid
            registrar_bitacora(
                cursor, usuario, 'crear_aula', 'aulas', registro_id=aula_id,
                aula_nueva=aula.nombre, estado_nuevo=aula.estado
            )
        connection.commit()
        return {"message": "Aula creada exitosamente"}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al crear aula: {str(e)}")
    finally:
        connection.close()


@app.put("/api/aulas/{aula_id}")
def actualizar_aula(aula_id: int, aula: Aula, request: Request):
    """Actualiza los datos de un aula existente."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            usuario = obtener_usuario(request)
            cursor.execute("SELECT nombre, estado FROM aulas WHERE id = %s", (aula_id,))
            ant = cursor.fetchone() or {}
            
            equipos_json = json.dumps(aula.equipos)
            cursor.execute(
                "UPDATE aulas SET nombre=%s, edificio=%s, capacidad=%s, equipos=%s, estado=%s WHERE id=%s",
                (aula.nombre, aula.edificio, aula.capacidad, equipos_json, aula.estado, aula_id)
            )
            registrar_bitacora(
                cursor, usuario, 'actualizar_aula', 'aulas', registro_id=aula_id,
                aula_anterior=ant.get('nombre'), aula_nueva=aula.nombre,
                estado_anterior=ant.get('estado'), estado_nuevo=aula.estado
            )
        connection.commit()
        return {"message": "Aula actualizada exitosamente"}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al actualizar aula: {str(e)}")
    finally:
        connection.close()


@app.post("/api/aulas/{aula_id}/mantenimiento")
def actualizar_mantenimiento(aula_id: int, datos: MantenimientoUpdate, request: Request):
    inicio_dt = None
    if datos.inicio_mantenimiento:
        try:
            inicio_dt = datetime.fromisoformat(datos.inicio_mantenimiento)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha de inicio inválido.")

    fin_dt = None
    if datos.fin_mantenimiento:
        try:
            fin_dt = datetime.fromisoformat(datos.fin_mantenimiento)
        except ValueError:
            raise HTTPException(status_code=400, detail="Formato de fecha de fin inválido.")

    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            usuario = obtener_usuario(request)
            cursor.execute("SELECT nombre FROM aulas WHERE id = %s", (aula_id,))
            row = cursor.fetchone() or {}
            
            cursor.execute(
                "UPDATE aulas SET en_mantenimiento=%s, inicio_mantenimiento=%s, fin_mantenimiento=%s, aula_temporal=%s WHERE id=%s",
                (
                    datos.en_mantenimiento,
                    inicio_dt if datos.en_mantenimiento else None,
                    fin_dt if datos.en_mantenimiento else None,
                    datos.aula_temporal if datos.en_mantenimiento else None,
                    aula_id
                )
            )
            registrar_bitacora(
                cursor, usuario, 'mantenimiento_aula', 'aulas', registro_id=aula_id,
                aula_anterior=row.get('nombre'),
                estado_nuevo="Mantenimiento" if datos.en_mantenimiento else "Disponible",
                motivo=f"Temporal: {datos.aula_temporal}" if datos.aula_temporal else ""
            )
        connection.commit()
        return {"message": "Estado de mantenimiento actualizado"}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al actualizar mantenimiento: {str(e)}")
    finally:
        connection.close()


@app.delete("/api/aulas/{aula_id}")
def eliminar_aula(aula_id: int, request: Request):
    """
    Elimina un aula por su ID
    """
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            usuario = obtener_usuario(request)
            cursor.execute("SELECT nombre FROM aulas WHERE id = %s", (aula_id,))
            row = cursor.fetchone() or {}
            
            cursor.execute("DELETE FROM aulas WHERE id = %s", (aula_id,))
            registrar_bitacora(
                cursor, usuario, 'eliminar_aula', 'aulas', registro_id=aula_id,
                aula_anterior=row.get('nombre')
            )
        connection.commit()
        return {"message": "Aula eliminada exitosamente"}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al eliminar aula: {str(e)}")
    finally:
        connection.close()


# ---------------------------------------------------------
# ENDPOINTS PARA GESTIÓN DE ARCHIVOS PDF
# ---------------------------------------------------------
@app.get("/api/archivos")
def obtener_archivos():
    """
    Obtiene lista de archivos únicos y estadísticas de los horarios
    """
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT DISTINCT archivo, 
                       COUNT(*) as total_horarios,
                       MIN(fecha_creacion) as fecha_carga,
                       MAX(fecha_creacion) as ultima_modificacion,
                       COUNT(CASE WHEN aula_asignada != 'Por asignar' THEN 1 END) as aulas_asignadas,
                       GROUP_CONCAT(DISTINCT CASE WHEN aula_asignada != 'Por asignar' AND TRIM(aula_asignada) != '' THEN aula_asignada END SEPARATOR ', ') as aulas_ocupadas,
                       GROUP_CONCAT(horario SEPARATOR '|') as todos_horarios,
                       MAX(NULLIF(TRIM(cuatrimestre), '')) as tiene_cuatri,
                       MAX(NULLIF(TRIM(semestre), '')) as tiene_semestre
                FROM horarios 
                GROUP BY archivo 
                ORDER BY fecha_carga DESC
            """)
            archivos = cursor.fetchall()
            
        LIMITE_VESPERTINO = 14 * 60 # 14:00
        for arch in (archivos or []):
            turnos = set()
            min_inicio = None
            max_fin = None
            if arch.get('todos_horarios'):
                horarios_lista = arch['todos_horarios'].split('|')
                for h in horarios_lista:
                    _, inicio, fin = _parse_horario_minutos(h)
                    if inicio is not None:
                        if min_inicio is None or inicio < min_inicio:
                            min_inicio = inicio
                        if fin is not None and (max_fin is None or fin > max_fin):
                            max_fin = fin
                        if inicio < LIMITE_VESPERTINO:
                            turnos.add("Matutino")
                        else:
                            turnos.add("Vespertino")
                    matches = re.findall(r'(\d{1,2}):(\d{2})\s*[-aA–—]\s*(\d{1,2}):(\d{2})', h)
                    for m_match in matches:
                        h_i, m_i, h_f, m_f = map(int, m_match)
                        st_m = h_i * 60 + m_i
                        en_m = h_f * 60 + m_f
                        if min_inicio is None or st_m < min_inicio:
                            min_inicio = st_m
                        if max_fin is None or en_m > max_fin:
                            max_fin = en_m
                        if st_m < LIMITE_VESPERTINO:
                            turnos.add("Matutino")
                        else:
                            turnos.add("Vespertino")

            if min_inicio is not None and max_fin is not None:
                h_i, m_i = min_inicio // 60, min_inicio % 60
                h_f, m_f = max_fin // 60, max_fin % 60
                arch['rango_horario'] = f"{h_i}:{m_i:02d} a {h_f}:{m_f:02d}"
            else:
                arch['rango_horario'] = None

            if "Matutino" in turnos and "Vespertino" in turnos:
                arch['turno'] = "Ambos Turnos"
            elif "Matutino" in turnos:
                arch['turno'] = "Matutino"
            elif "Vespertino" in turnos:
                arch['turno'] = "Vespertino"
            else:
                arch['turno'] = "Sin horario fijo"
            
            # Determinar el plan (Semestral o Cuatrimestral)
            if arch.get('tiene_cuatri') and arch.get('tiene_semestre'):
                arch['plan'] = 'Mixto'
            elif arch.get('tiene_cuatri'):
                arch['plan'] = 'Cuatrimestral'
            elif arch.get('tiene_semestre'):
                arch['plan'] = 'Semestral'
            else:
                arch['plan'] = 'No definido'
                
            # Limpiar campos pesados/innecesarios antes de devolver
            if 'todos_horarios' in arch: del arch['todos_horarios']
            if 'tiene_cuatri' in arch: del arch['tiene_cuatri']
            if 'tiene_semestre' in arch: del arch['tiene_semestre']

        return archivos if archivos else []
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener archivos: {str(e)}")
    finally:
        connection.close()


@app.get("/api/archivos/{nombre_archivo}/horarios")
def obtener_horarios_archivo(nombre_archivo: str):
    """
    Obtiene todos los horarios de un archivo específico
    """
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, docente, licenciatura, asignatura, horario, aula_asignada, fecha_creacion, semestre, cuatrimestre, grupo
                FROM horarios 
                WHERE archivo = %s
                ORDER BY horario
            """, (nombre_archivo,))
            horarios = cursor.fetchall()
        return horarios if horarios else []
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener horarios: {str(e)}")
    finally:
        connection.close()


@app.put("/api/horarios/{horario_id}")
def actualizar_horario(horario_id: int, datos: dict, request: Request):
    """Actualiza un horario específico (aula, docente, asignatura). Valida conflictos de aula."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            nueva_aula = datos.get("aula_asignada", "Por asignar")
            usuario = obtener_usuario(request)
            # Obtener el horario original para parsear día y hora
            cursor.execute("SELECT docente, asignatura, horario, aula_asignada FROM horarios WHERE id = %s", (horario_id,))
            actual = cursor.fetchone() or {}
            if actual:
                msg = _verificar_conflicto_aula(cursor, nueva_aula, actual.get('horario', ''), excluir_id=horario_id)
                if msg:
                    raise HTTPException(status_code=409, detail=msg)
            sql = "UPDATE horarios SET aula_asignada = %s, docente = %s, asignatura = %s WHERE id = %s"
            cursor.execute(sql, (
                nueva_aula,
                datos.get("docente", ""),
                datos.get("asignatura", ""),
                horario_id
            ))
            registrar_bitacora(
                cursor, usuario, 'actualizar_clase', 'horarios', registro_id=horario_id,
                docente=datos.get("docente", ""), asignatura=datos.get("asignatura", ""),
                aula_anterior=actual.get('aula_asignada'), aula_nueva=nueva_aula,
                datos_anteriores={"docente": actual.get("docente"), "asignatura": actual.get("asignatura")}
            )
        connection.commit()
        return {"message": "Horario actualizado exitosamente"}
    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al actualizar horario: {str(e)}")
    finally:
        connection.close()


@app.delete("/api/archivos/{nombre_archivo}")
def eliminar_archivo(nombre_archivo: str, request: Request):
    """
    Elimina un archivo y todos sus horarios asociados
    """
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            usuario = obtener_usuario(request)
            cursor.execute("DELETE FROM horarios WHERE archivo = %s", (nombre_archivo,))
            registrar_bitacora(
                cursor, usuario, 'eliminar_archivo_horarios', 'horarios',
                datos_anteriores={"archivo": nombre_archivo}
            )
        connection.commit()
        return {"message": f"Archivo '{nombre_archivo}' y sus horarios eliminados exitosamente"}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al eliminar archivo: {str(e)}")
    finally:
        connection.close()


# ---------------------------------------------------------
# ENDPOINTS PARA GESTIÓN DE DOCENTES
# ---------------------------------------------------------

@app.get("/api/docentes")
def obtener_docentes():
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM docentes ORDER BY nombre")
            docentes = cursor.fetchall()

            ahora = datetime.now()
            mins_ahora = ahora.hour * 60 + ahora.minute
            # isoweekday: Lun=1..Dom=7 → % 7 da Lun=1..Sab=6, Dom=0 (igual que JS getDay())
            dia_hoy = ahora.isoweekday() % 7

            cursor.execute("""
                SELECT s.*, ds.nombre as suplente_nombre
                FROM suplencias s
                LEFT JOIN docentes ds ON ds.id = s.suplente_id
                WHERE s.activa = TRUE AND s.fecha = CURDATE()
            """)
            suplencias_hoy = cursor.fetchall()

            # Sin filtro de fecha: usamos el campo horario (día semana + hora) igual que obtener_clases_hoy
            cursor.execute("""
                SELECT DISTINCT docente, horario FROM horarios
                WHERE docente IS NOT NULL AND TRIM(docente) != ''
            """)
            clases_semana = cursor.fetchall()

        for doc in (docentes or []):
            try:
                doc['materias'] = json.loads(doc.get('materias') or '[]')
            except Exception:
                doc['materias'] = []

            suplencia_activa = None
            for s in (suplencias_hoy or []):
                if s['docente_id'] != doc['id']:
                    continue
                mins_i = _time_to_mins(s['hora_inicio'])
                mins_f = _time_to_mins(s['hora_fin'])
                if mins_ahora >= mins_i and mins_ahora <= mins_f:
                    suplencia_activa = {
                        **{k: v for k, v in s.items() if k not in ('hora_inicio', 'hora_fin', 'fecha')},
                        'hora_inicio': _timedelta_to_str(s['hora_inicio']),
                        'hora_fin': _timedelta_to_str(s['hora_fin']),
                        'fecha': str(s['fecha']) if s.get('fecha') else ''
                    }
                    break

            en_clase = False
            if not suplencia_activa:
                for clase in (clases_semana or []):
                    if clase['docente'] != doc['nombre']:
                        continue
                    dia_clase, inicio, fin = _parse_horario_minutos(clase.get('horario', ''))
                    if dia_clase == dia_hoy and inicio is not None and inicio <= mins_ahora <= fin:
                        en_clase = True
                        break

            doc['suplencia_activa'] = suplencia_activa
            if suplencia_activa:
                doc['estado'] = 'suplente_asignado'
            elif en_clase:
                doc['estado'] = 'en_clase'
            else:
                doc['estado'] = 'disponible'

        return docentes if docentes else []
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener docentes: {str(e)}")
    finally:
        connection.close()


@app.post("/api/docentes")
def crear_docente(docente: Docente, request: Request):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "INSERT INTO docentes (nombre, especialidad, materias, correo) VALUES (%s, %s, %s, %s)",
            )
            doc_id = cursor.lastrowid
            registrar_bitacora(
                cursor, obtener_usuario(request), 'crear_docente', 'docentes', registro_id=doc_id,
                docente=docente.nombre.strip(),
                datos_nuevos={"especialidad": docente.especialidad.strip(), "correo": docente.correo.strip()}
            )
        connection.commit()
        return {"message": "Docente registrado exitosamente"}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al crear docente: {str(e)}")
    finally:
        connection.close()


@app.put("/api/docentes/{docente_id}")
def actualizar_docente(docente_id: int, docente: Docente, request: Request):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT nombre, especialidad, correo FROM docentes WHERE id = %s", (docente_id,))
            ant = cursor.fetchone() or {}
            
            cursor.execute(
                "UPDATE docentes SET nombre=%s, especialidad=%s, materias=%s, correo=%s WHERE id=%s",
                (docente.nombre.strip(), docente.especialidad.strip(), json.dumps(docente.materias), docente.correo.strip(), docente_id)
            )
            registrar_bitacora(
                cursor, obtener_usuario(request), 'actualizar_docente', 'docentes', registro_id=docente_id,
                docente=docente.nombre.strip(),
                datos_anteriores={"nombre": ant.get('nombre'), "especialidad": ant.get('especialidad'), "correo": ant.get('correo')},
                datos_nuevos={"nombre": docente.nombre.strip(), "especialidad": docente.especialidad.strip(), "correo": docente.correo.strip()}
            )
        connection.commit()
        return {"message": "Docente actualizado exitosamente"}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al actualizar docente: {str(e)}")
    finally:
        connection.close()


@app.delete("/api/docentes/{docente_id}")
def eliminar_docente(docente_id: int, request: Request):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT nombre FROM docentes WHERE id = %s", (docente_id,))
            ant = cursor.fetchone() or {}
            
            cursor.execute("DELETE FROM suplencias WHERE docente_id = %s OR suplente_id = %s", (docente_id, docente_id))
            cursor.execute("DELETE FROM docentes WHERE id = %s", (docente_id,))
            registrar_bitacora(
                cursor, obtener_usuario(request), 'eliminar_docente', 'docentes', registro_id=docente_id,
                docente=ant.get('nombre')
            )
        connection.commit()
        return {"message": "Docente eliminado exitosamente"}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al eliminar docente: {str(e)}")
    finally:
        connection.close()


# ---------------------------------------------------------
# ENDPOINTS PARA GESTIÓN DE SUPLENCIAS
# ---------------------------------------------------------

@app.get("/api/suplencias")
def obtener_suplencias():
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT s.*,
                       d.nombre  AS docente_nombre,
                       ds.nombre AS suplente_nombre
                FROM suplencias s
                LEFT JOIN docentes d  ON d.id  = s.docente_id
                LEFT JOIN docentes ds ON ds.id = s.suplente_id
                WHERE s.activa = TRUE AND s.fecha >= CURDATE()
                ORDER BY s.fecha, s.hora_inicio
            """)
            suplencias = cursor.fetchall()
        for s in (suplencias or []):
            s['hora_inicio'] = _timedelta_to_str(s['hora_inicio'])
            s['hora_fin']    = _timedelta_to_str(s['hora_fin'])
            if s.get('fecha'):
                s['fecha'] = str(s['fecha'])
        return suplencias if suplencias else []
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener suplencias: {str(e)}")
    finally:
        connection.close()


@app.post("/api/suplencias")
def crear_suplencia(s: Suplencia, request: Request):
    if s.docente_id == s.suplente_id:
        raise HTTPException(status_code=400, detail="El docente no puede ser su propio suplente.")
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id FROM suplencias
                WHERE suplente_id = %s AND activa = TRUE AND fecha = %s
                AND NOT (hora_fin <= %s OR hora_inicio >= %s)
            """, (s.suplente_id, s.fecha, s.hora_inicio, s.hora_fin))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail="El suplente ya tiene una suplencia asignada en ese horario.")
            cursor.execute("""
                INSERT INTO suplencias (docente_id, suplente_id, materia, dia, fecha, hora_inicio, hora_fin, activa)
                VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE)
            """, (s.docente_id, s.suplente_id, s.materia, s.dia, s.fecha, s.hora_inicio, s.hora_fin))
            suplencia_id = cursor.lastrowid
            
            cursor.execute("SELECT id, nombre FROM docentes WHERE id IN (%s, %s)", (s.docente_id, s.suplente_id))
            docentes = {row['id']: row['nombre'] for row in cursor.fetchall()}
            
            registrar_bitacora(
                cursor, obtener_usuario(request), 'asignar_suplente', 'suplencias', registro_id=suplencia_id,
                docente=docentes.get(s.docente_id, str(s.docente_id)),
                suplente=docentes.get(s.suplente_id, str(s.suplente_id)),
                asignatura=s.materia,
                hora_nueva=f"{s.hora_inicio}-{s.hora_fin}", dia_nuevo=s.dia
            )
        connection.commit()
        return {"message": "Suplencia asignada correctamente."}
    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al crear suplencia: {str(e)}")
    finally:
        connection.close()


@app.delete("/api/suplencias/{suplencia_id}")
def eliminar_suplencia(suplencia_id: int, request: Request):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT s.materia, d.nombre AS docente, ds.nombre AS suplente 
                FROM suplencias s 
                LEFT JOIN docentes d ON s.docente_id = d.id 
                LEFT JOIN docentes ds ON s.suplente_id = ds.id 
                WHERE s.id = %s
            """, (suplencia_id,))
            ant = cursor.fetchone() or {}
            
            cursor.execute("UPDATE suplencias SET activa = FALSE WHERE id = %s", (suplencia_id,))
            registrar_bitacora(
                cursor, obtener_usuario(request), 'retiro_suplente', 'suplencias', registro_id=suplencia_id,
                docente=ant.get('docente'), suplente=ant.get('suplente'), asignatura=ant.get('materia')
            )
        connection.commit()
        return {"message": "Suplencia cancelada."}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al cancelar suplencia: {str(e)}")
    finally:
        connection.close()


# ---------------------------------------------------------
# ENDPOINTS DE REPROGRAMACIÓN DE CLASES (POSPUESTAS Y REPOSICIONES)
# ---------------------------------------------------------

@app.get("/api/docentes/{docente_nombre}/clases-semana")
def obtener_clases_semana_docente(docente_nombre: str):
    """
    Retorna dinámicamente las clases programadas para un docente en la semana en curso.
    """
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT docente FROM horarios WHERE docente IS NOT NULL AND TRIM(docente) != ''")
            raw_docentes = [r['docente'] for r in (cursor.fetchall() or [])]
            mapa_canonico = _construir_mapa_docentes_canonicos(raw_docentes)

            nombres_coincidentes = [r for r in raw_docentes if mapa_canonico.get(r) == docente_nombre or r == docente_nombre]
            if not nombres_coincidentes:
                nombres_coincidentes = [docente_nombre]

            format_strings = ','.join(['%s'] * len(nombres_coincidentes))

            query = f"""
                SELECT
                    MIN(h.id) AS id,
                    h.docente,
                    h.asignatura,
                    h.horario,
                    h.aula_asignada,
                    MAX(h.licenciatura) AS licenciatura,
                    MAX(h.semestre) AS semestre,
                    MAX(h.cuatrimestre) AS cuatrimestre,
                    MAX(h.grupo) AS grupo,
                    MAX(h.archivo) AS archivo,
                    COALESCE(MAX(h.estado_slug), 'programada') AS estado_slug,
                    MAX(h.nota_reprogramacion) AS nota_reprogramacion,
                    MAX(h.fecha_reposicion) AS fecha_reposicion,
                    MAX(h.es_reposicion) AS es_reposicion
                FROM horarios h
                WHERE h.docente IN ({format_strings})
                  AND (h.estado_slug IS NULL OR h.estado_slug != 'reposicion_programada')
                  AND (h.es_reposicion IS NULL OR h.es_reposicion = FALSE)
                GROUP BY h.docente, h.asignatura, h.horario, h.aula_asignada
                ORDER BY h.horario
            """
            cursor.execute(query, tuple(nombres_coincidentes))
            clases = cursor.fetchall()
            return clases or []
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener clases del docente: {str(e)}")
    finally:
        connection.close()


@app.get("/api/aulas/disponibles-reprogramacion")
def obtener_aulas_disponibles_reprogramacion(
    fecha: str,
    hora_inicio: str,
    hora_fin: str,
    dia: Optional[str] = None
):
    """
    Retorna dinámicamente las aulas disponibles calculadas estrictamente por intersección de tiempo (Overlapping).
    Regla de choque: (hora_inicio_existente < hora_fin_solicitada) AND (hora_fin_existente > hora_inicio_solicitada)
    Permite asignar salones en "huecos" u horas muertas entre turnos.
    """
    def time_str_to_mins(t_str: str) -> int:
        if not t_str:
            return -1
        p = t_str.strip().split(':')
        return int(p[0]) * 60 + int(p[1])

    req_inicio = time_str_to_mins(hora_inicio)
    req_fin = time_str_to_mins(hora_fin)
    if req_inicio < 0 or req_fin < 0 or req_inicio >= req_fin:
        raise HTTPException(status_code=400, detail="Rango de horario inválido.")

    try:
        dt = datetime.strptime(fecha, "%Y-%m-%d")
        dia_sys_idx = (dt.isoweekday() % 7)  # 0=Domingo, 1=Lunes, ... 6=Sabado
    except Exception:
        raise HTTPException(status_code=400, detail="Fecha inválida. Formato esperado YYYY-MM-DD.")

    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            # 0. Limpiar mantenimientos vencidos automáticamente
            cursor.execute("""
                UPDATE aulas 
                SET en_mantenimiento = 0, inicio_mantenimiento = NULL, fin_mantenimiento = NULL, aula_temporal = NULL 
                WHERE en_mantenimiento = 1 AND fin_mantenimiento IS NOT NULL AND fin_mantenimiento <= NOW()
            """)
            connection.commit()

            # 1. Obtener todas las aulas que no estén en mantenimiento vigente
            cursor.execute("""
                SELECT nombre, capacidad, edificio AS ubicacion
                FROM aulas
                WHERE en_mantenimiento = FALSE
                   OR (en_mantenimiento = TRUE AND fin_mantenimiento IS NOT NULL AND fin_mantenimiento <= NOW())
                   OR (en_mantenimiento = TRUE AND inicio_mantenimiento IS NOT NULL AND inicio_mantenimiento > NOW())
                ORDER BY nombre
            """)
            todas_aulas = cursor.fetchall() or []
            aulas_disponibles_map = {a['nombre']: a for a in todas_aulas}

            # 2. Verificar cruces con clases regulares y reposiciones en la tabla horarios
            cursor.execute("""
                SELECT h.aula_asignada, h.horario, h.estado_slug, h.es_reposicion, h.fecha_reposicion
                FROM horarios h
                WHERE h.aula_asignada IS NOT NULL
                  AND TRIM(h.aula_asignada) != ''
                  AND TRIM(h.aula_asignada) != 'Por asignar'
            """)
            horarios_raw = cursor.fetchall() or []

            for h in horarios_raw:
                aula_nombre = h['aula_asignada']
                if aula_nombre not in aulas_disponibles_map:
                    continue

                # Si la clase original fue pospuesta, no bloquea el salón en su horario original
                if h.get('estado_slug') == 'pospuesta':
                    continue

                # Si es reposición, solo evalúa choque si la fecha coincide exactamente
                if h.get('es_reposicion'):
                    if not h.get('fecha_reposicion') or str(h['fecha_reposicion']) != fecha:
                        continue

                h_dia_idx, h_ini, h_fin = _parse_horario_minutos(h.get('horario', ''))
                if h_ini is None or h_fin is None:
                    continue

                # Para clases semanales regulares, debe coincidir el día de la semana
                if not h.get('es_reposicion'):
                    if h_dia_idx != dia_sys_idx:
                        continue

                # Fórmula estricta de intersección de tiempo (Overlapping):
                # (hora_inicio_existente < hora_fin_solicitada) AND (hora_fin_existente > hora_inicio_solicitada)
                if (h_ini < req_fin) and (h_fin > req_inicio):
                    aulas_disponibles_map.pop(aula_nombre, None)

            # 3. Verificar cruces con suplencias activas en esa fecha
            cursor.execute("""
                SELECT s.hora_inicio, s.hora_fin, h.aula_asignada
                FROM suplencias_horarios s
                LEFT JOIN horarios h ON h.docente = s.docente_nombre AND h.horario LIKE CONCAT('%%', s.dia, '%%')
                WHERE s.activa = TRUE AND s.fecha = %s
            """, (fecha,))
            suplencias = cursor.fetchall() or []
            for s in suplencias:
                aula_nombre = s.get('aula_asignada')
                if not aula_nombre or aula_nombre not in aulas_disponibles_map:
                    continue
                s_ini = _time_to_mins(_timedelta_to_str(s['hora_inicio']))
                s_fin = _time_to_mins(_timedelta_to_str(s['hora_fin']))
                if (s_ini < req_fin) and (s_fin > req_inicio):
                    aulas_disponibles_map.pop(aula_nombre, None)

            return list(aulas_disponibles_map.values())
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al verificar disponibilidad de aulas: {str(e)}")
    finally:
        connection.close()


@app.post("/api/reprogramaciones")
def crear_reprogramacion(req: ReprogramacionRequest, request: Request):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            ids_a_posponer = req.clases_originales_ids if req.clases_originales_ids else ([req.clase_original_id] if req.clase_original_id else [])
            if not ids_a_posponer:
                raise HTTPException(status_code=400, detail="Debe especificar al menos una clase original para posponer.")

            orig_primera = None
            horarios_orig = []
            for cid in ids_a_posponer:
                cursor.execute("""
                    SELECT * FROM horarios WHERE id = %s
                """, (cid,))
                orig = cursor.fetchone()
                if not orig:
                    continue
                if not orig_primera:
                    orig_primera = orig
                if orig['horario'] not in horarios_orig:
                    horarios_orig.append(orig['horario'])

                nota_rep = f"Pospuesta (Se repone el {req.nuevo_dia} a las {req.nueva_hora_inicio}-{req.nueva_hora_fin})"

                cursor.execute("""
                    UPDATE horarios
                    SET estado_slug = 'pospuesta',
                        nota_reprogramacion = %s
                    WHERE id = %s
                """, (nota_rep, cid))

            if not orig_primera:
                raise HTTPException(status_code=404, detail="No se encontró ninguna de las clases originales en la base de datos.")

            nuevo_horario_str = f"{req.nuevo_dia} {req.nueva_hora_inicio}-{req.nueva_hora_fin}"
            horarios_orig_str = " y ".join(horarios_orig)
            nota_reposicion_origen = f"Reposición de {horarios_orig_str} ({req.motivo or 'Clase pospuesta'})"

            cursor.execute("""
                INSERT INTO horarios (
                    docente, licenciatura, asignatura, horario, aula_asignada, archivo,
                    fecha_creacion, fecha_clase, semestre, cuatrimestre, grupo,
                    estado_slug, nota_reprogramacion, fecha_reposicion, es_reposicion
                ) VALUES (
                    %s, %s, %s, %s, %s, %s,
                    NOW(), CURDATE(), %s, %s, %s,
                    'reposicion_programada', %s, %s, TRUE
                )
            """, (
                orig_primera['docente'],
                orig_primera['licenciatura'] or '',
                req.clase_original_asignatura or orig_primera['asignatura'] or '',
                nuevo_horario_str,
                req.nueva_aula,
                orig_primera['archivo'] or '',
                orig_primera['semestre'] or '',
                orig_primera['cuatrimestre'] or '',
                orig_primera['grupo'] or '',
                nota_reposicion_origen,
                req.nueva_fecha
            ))
            
            reposicion_id = cursor.lastrowid
            registrar_bitacora(
                cursor, obtener_usuario(request), 'reprogramar_clase', 'horarios',
                registro_id=reposicion_id,
                docente=orig_primera['docente'], asignatura=req.clase_original_asignatura or orig_primera['asignatura'] or '',
                hora_anterior=horarios_orig_str, hora_nueva=nuevo_horario_str,
                aula_anterior=orig_primera['aula_asignada'], aula_nueva=req.nueva_aula,
                motivo=req.motivo or 'Clase pospuesta'
            )

        connection.commit()
        return {"message": "Clase reprogramada exitosamente en la base de datos."}
    except HTTPException:
        connection.rollback()
        raise
    except pymysql.Error as e:
        connection.rollback()
        raise HTTPException(status_code=500, detail=f"Error transaccional al reprogramar clase: {str(e)}")
    finally:
        connection.close()


@app.delete("/api/reprogramaciones/{clase_id}")
def cancelar_reprogramacion(clase_id: int, request: Request):
    """
    Revoca la reprogramación de una clase (ya sea por ID de clase pospuesta o por ID de clase de reposición):
    1. Reestablece la clase original a estado programada (sin nota de pospuesta).
    2. Elimina en BD el registro de la clase de reposición asociada del docente.
    """
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT * FROM horarios WHERE id = %s", (clase_id,))
            item = cursor.fetchone()
            if not item:
                raise HTTPException(status_code=404, detail="No se encontró la clase en la base de datos.")

            docente = item['docente']

            if item.get('es_reposicion') or item.get('estado_slug') == 'reposicion_programada':
                cursor.execute("DELETE FROM horarios WHERE id = %s", (clase_id,))
                cursor.execute("""
                    UPDATE horarios
                    SET estado_slug = NULL,
                        nota_reprogramacion = NULL
                    WHERE docente = %s AND (estado_slug = 'pospuesta' OR nota_reprogramacion LIKE 'Pospuesta%%')
                """, (docente,))
            else:
                cursor.execute("""
                    UPDATE horarios
                    SET estado_slug = NULL,
                        nota_reprogramacion = NULL
                    WHERE id = %s
                """, (clase_id,))
                cursor.execute("""
                    DELETE FROM horarios
                    WHERE docente = %s AND (es_reposicion = TRUE OR estado_slug = 'reposicion_programada')
                """, (docente,))

            registrar_bitacora(
                cursor, obtener_usuario(request), 'cancelar_reprogramacion', 'horarios',
                registro_id=clase_id, docente=docente, asignatura=item.get('asignatura')
            )
        connection.commit()
        return {"message": "Reprogramación cancelada y revertida en la base de datos."}
    except HTTPException:
        connection.rollback()
        raise
    except pymysql.Error as e:
        connection.rollback()
        raise HTTPException(status_code=500, detail=f"Error al cancelar reprogramación: {str(e)}")
    finally:
        connection.close()


def archivar_reposiciones_completadas(connection, effective_dt=None):
    """
    Archiva clases reprogramadas finalizadas en historial_reprogramaciones_semanal
    y normaliza el horario regular en horarios.
    """
    if effective_dt is None:
        effective_dt = datetime.now()
    fecha_hoy = effective_dt.date()
    hora_hoy_str = effective_dt.strftime("%H:%M")
    semana_str = effective_dt.strftime("%Y-W%W")

    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, docente, licenciatura, asignatura, horario, aula_asignada, fecha_reposicion, nota_reprogramacion
                FROM horarios
                WHERE (estado_slug = 'reposicion_programada' OR es_reposicion = TRUE)
                  AND fecha_reposicion IS NOT NULL
            """)
            repos = cursor.fetchall()
            for row in repos:
                frep = row['fecha_reposicion']
                if isinstance(frep, str):
                    try:
                        frep = datetime.strptime(frep, "%Y-%m-%d").date()
                    except:
                        continue

                hora_fin_str = "23:59"
                if row.get('horario'):
                    m = re.search(r'(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})', str(row['horario']))
                    if m:
                        hora_fin_str = m.group(2)

                terminada = False
                if frep < fecha_hoy:
                    terminada = True
                elif frep == fecha_hoy and hora_fin_str <= hora_hoy_str:
                    terminada = True

                if terminada:
                    cursor.execute("""
                        INSERT INTO historial_reprogramaciones_semanal (
                            docente, licenciatura, asignatura, horario_reposicion, fecha_reposicion,
                            aula_asignada, nota_reprogramacion, motivo, fecha_archivado, semana_anio
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (
                        row['docente'], row.get('licenciatura', ''), row.get('asignatura', ''),
                        row.get('horario', ''), frep, row.get('aula_asignada', ''),
                        row.get('nota_reprogramacion', ''), 'Reposición finalizada',
                        effective_dt, semana_str
                    ))

                    cursor.execute("DELETE FROM horarios WHERE id = %s", (row['id'],))

                    cursor.execute("""
                        UPDATE horarios
                        SET estado_slug = NULL,
                            nota_reprogramacion = NULL
                        WHERE docente = %s AND (estado_slug = 'pospuesta' OR nota_reprogramacion LIKE 'Pospuesta%%')
                    """, (row['docente'],))

        connection.commit()
    except Exception as e:
        print(f"Error archivar_reposiciones_completadas: {e}")

# ---------------------------------------------------------
# ENDPOINTS PARA HISTORIAL Y REPORTE SEMANAL DE REPROGRAMACIONES
# ---------------------------------------------------------

@app.get("/api/reprogramaciones/historial-semanal")
def obtener_historial_reprogramaciones_semanal(fecha_ref: Optional[str] = None):
    connection = get_db_connection()
    try:
        effective_dt = datetime.now()
        if fecha_ref:
            try:
                clean_dt = fecha_ref.split('T')[0]
                hora_part = '00:00'
                if 'T' in fecha_ref:
                    hora_part = fecha_ref.split('T')[1][:5]
                effective_dt = datetime.strptime(f"{clean_dt} {hora_part}", "%Y-%m-%d %H:%M")
            except Exception:
                pass
        archivar_reposiciones_completadas(connection, effective_dt)
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT * FROM historial_reprogramaciones_semanal
                ORDER BY fecha_reposicion DESC, id DESC
            """)
            filas = cursor.fetchall()
        return filas or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()

@app.post("/api/reprogramaciones/archivar-completadas")
def archivar_completadas_endpoint(fecha_ref: Optional[str] = None):
    connection = get_db_connection()
    try:
        effective_dt = datetime.now()
        if fecha_ref:
            try:
                clean_dt = fecha_ref.split('T')[0]
                hora_part = '00:00'
                if 'T' in fecha_ref:
                    hora_part = fecha_ref.split('T')[1][:5]
                effective_dt = datetime.strptime(f"{clean_dt} {hora_part}", "%Y-%m-%d %H:%M")
            except Exception:
                pass
        archivar_reposiciones_completadas(connection, effective_dt)
        return {"message": "Reposiciones archivadas y horarios normalizados correctamente."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()

@app.get("/api/reprogramaciones/exportar-historial-csv")
def exportar_historial_csv(
    docente: Optional[str] = None,
    licenciatura: Optional[str] = None,
    asignatura: Optional[str] = None,
    busqueda: Optional[str] = None,
    semana_anio: Optional[str] = None
):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            where_clauses = []
            params = []
            if docente:
                where_clauses.append("docente LIKE %s")
                params.append(f"%{docente}%")
            if licenciatura and licenciatura != 'todas':
                where_clauses.append("licenciatura LIKE %s")
                params.append(f"%{licenciatura}%")
            if asignatura and asignatura != 'todas':
                where_clauses.append("asignatura LIKE %s")
                params.append(f"%{asignatura}%")
            if busqueda:
                where_clauses.append("(docente LIKE %s OR asignatura LIKE %s OR aula_asignada LIKE %s)")
                params.extend([f"%{busqueda}%", f"%{busqueda}%", f"%{busqueda}%"])
            if semana_anio:
                where_clauses.append("semana_anio = %s")
                params.append(semana_anio)

            where_sql = " WHERE " + " AND ".join(where_clauses) if where_clauses else ""
            cursor.execute(f"""
                SELECT docente, licenciatura, asignatura, horario_reposicion, fecha_reposicion, aula_asignada, nota_reprogramacion, motivo, fecha_archivado, semana_anio
                FROM historial_reprogramaciones_semanal
                {where_sql}
                ORDER BY fecha_reposicion DESC, id DESC
            """, tuple(params))
            filas = cursor.fetchall()

        output = io.StringIO()
        output.write('\uFEFF')
        writer = csv.writer(output)
        writer.writerow(['Docente', 'Licenciatura', 'Asignatura', 'Horario Reposición', 'Fecha Reposición', 'Aula Asignada', 'Nota Reprogramación', 'Motivo', 'Fecha Archivado', 'Semana'])

        for f in (filas or []):
            writer.writerow([
                f.get('docente', ''),
                f.get('licenciatura', ''),
                f.get('asignatura', ''),
                f.get('horario_reposicion', ''),
                f.get('fecha_reposicion', ''),
                f.get('aula_asignada', ''),
                f.get('nota_reprogramacion', ''),
                f.get('motivo', ''),
                f.get('fecha_archivado', ''),
                f.get('semana_anio', '')
            ])

        content = output.getvalue()
        return Response(
            content=content,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=reporte_reprogramaciones_semanal.csv"}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


# ---------------------------------------------------------
# ENDPOINTS PARA DOCENTES DESDE HORARIOS (SIN TABLA PROPIA)
# ---------------------------------------------------------

@app.get("/api/docentes-horarios")
def obtener_docentes_horarios(fecha_ref: Optional[str] = None):
    """Docentes únicos de horarios. Incluye horarios_hoy para cálculo en tiempo real en el frontend."""
    connection = get_db_connection()
    try:
        effective_dt = datetime.now()
        if fecha_ref:
            try:
                clean_dt = fecha_ref.split('T')[0]
                hora_part = '00:00'
                if 'T' in fecha_ref:
                    hora_part = fecha_ref.split('T')[1][:5]
                effective_dt = datetime.strptime(f"{clean_dt} {hora_part}", "%Y-%m-%d %H:%M")
            except Exception:
                pass
        archivar_reposiciones_completadas(connection, effective_dt)
        with connection.cursor() as cursor:
            cursor.execute("SET SESSION group_concat_max_len = 65535")
            cursor.execute("""
                SELECT
                    docente AS nombre,
                    GROUP_CONCAT(DISTINCT asignatura  ORDER BY asignatura  SEPARATOR '|||') AS materias_raw,
                    GROUP_CONCAT(DISTINCT licenciatura ORDER BY licenciatura SEPARATOR '|||') AS licenciaturas_raw
                FROM horarios
                WHERE docente IS NOT NULL
                  AND TRIM(docente) != ''
                  AND TRIM(docente) != 'Sin especificar'
                GROUP BY docente
                ORDER BY docente
            """)
            filas = cursor.fetchall()

            # Todos los horarios semanales (sin filtro de fecha) para cálculo client-side
            cursor.execute("""
                SELECT docente, horario, asignatura,
                       MAX(NULLIF(TRIM(cuatrimestre), '')) as cuatrimestre
                FROM horarios
                WHERE docente IS NOT NULL AND TRIM(docente) != ''
                  AND TRIM(docente) != 'Sin especificar'
                GROUP BY docente, horario, asignatura
            """)
            clases_hoy_raw = cursor.fetchall()

            # Suplencias activas hoy
            cursor.execute("""
                SELECT * FROM suplencias_horarios
                WHERE activa = TRUE AND fecha = CURDATE()
            """)
            suplencias_hoy = cursor.fetchall()

            # Clases reprogramadas o reposiciones programadas (nueva clase)
            cursor.execute("""
                SELECT id, docente, asignatura, horario, aula_asignada, fecha_reposicion, nota_reprogramacion, estado_slug
                FROM horarios
                WHERE (estado_slug = 'reposicion_programada' OR es_reposicion = TRUE OR (nota_reprogramacion IS NOT NULL AND TRIM(nota_reprogramacion) != '' AND (estado_slug IS NULL OR estado_slug != 'pospuesta')))
                  AND docente IS NOT NULL AND TRIM(docente) != ''
                  AND TRIM(docente) != 'Sin especificar'
            """)
            clases_reprogramadas_raw = cursor.fetchall()

        # Construir mapa canónico de nombres para unificar duplicados ("Last, First" vs "First Last", segundo nombre omitido)
        nombres_raw_unicos = [f['nombre'] for f in (filas or []) if f.get('nombre')]
        mapa_canonico = _construir_mapa_docentes_canonicos(nombres_raw_unicos)

        docentes_dict = defaultdict(lambda: {
            'materias': set(),
            'licenciaturas': set(),
            'horarios_semana': [],
            'suplencias_hoy': [],
            'clases_reprogramadas': [],
            'nombres_raw': set()
        })

        for row in (filas or []):
            raw_n = row['nombre']
            canon_n = mapa_canonico.get(raw_n, _normalizar_nombre_formato(raw_n))
            docentes_dict[canon_n]['nombres_raw'].add(raw_n)

            if row.get('materias_raw'):
                for m in row['materias_raw'].split('|||'):
                    if m.strip():
                        docentes_dict[canon_n]['materias'].add(m.strip())
            if row.get('licenciaturas_raw'):
                for l in row['licenciaturas_raw'].split('|||'):
                    if l.strip():
                        docentes_dict[canon_n]['licenciaturas'].add(l.strip())

        for c in (clases_hoy_raw or []):
            raw_n = c['docente']
            canon_n = mapa_canonico.get(raw_n, _normalizar_nombre_formato(raw_n))
            dia_idx, inicio, fin = _parse_horario_minutos(c.get('horario', ''))
            if inicio is not None and dia_idx is not None:
                docentes_dict[canon_n]['horarios_semana'].append({
                    'asignatura':  c['asignatura'],
                    'dia_index':   dia_idx,   # 0=Dom,1=Lun...6=Sab (igual que JS getDay())
                    'inicio_mins': inicio,
                    'fin_mins':    fin,
                    'es_cuatri':   bool(c.get('cuatrimestre')),
                })

        for s in (suplencias_hoy or []):
            raw_n = s['docente_nombre']
            canon_n = mapa_canonico.get(raw_n, _normalizar_nombre_formato(raw_n))
            docentes_dict[canon_n]['suplencias_hoy'].append({
                'id':              s['id'],
                'suplente_nombre': s['suplente_nombre'],
                'materia':         s['materia'],
                'hora_inicio':     _timedelta_to_str(s['hora_inicio']),
                'hora_fin':        _timedelta_to_str(s['hora_fin']),
                'inicio_mins':     _time_to_mins(s['hora_inicio']),
                'fin_mins':        _time_to_mins(s['hora_fin']),
            })

        for cr in (clases_reprogramadas_raw or []):
            raw_n = cr['docente']
            canon_n = mapa_canonico.get(raw_n, _normalizar_nombre_formato(raw_n))
            docentes_dict[canon_n]['clases_reprogramadas'].append({
                'id':                  cr.get('id'),
                'asignatura':          cr.get('asignatura') or '',
                'horario':             cr.get('horario') or '',
                'aula_asignada':       cr.get('aula_asignada') or '',
                'fecha_reposicion':    str(cr.get('fecha_reposicion')) if cr.get('fecha_reposicion') else '',
                'nota_reprogramacion': cr.get('nota_reprogramacion') or '',
                'estado_slug':         cr.get('estado_slug') or '',
            })

        docentes = []
        for nombre_canonico, data in sorted(docentes_dict.items(), key=lambda x: x[0]):
            docentes.append({
                'nombre':               nombre_canonico,
                'materias':             sorted(list(data['materias'])),
                'licenciaturas':        sorted(list(data['licenciaturas'])),
                'horarios_semana':      data['horarios_semana'],
                'suplencias_hoy':       data['suplencias_hoy'],
                'clases_reprogramadas': data['clases_reprogramadas'],
                'nombres_raw':          list(data['nombres_raw']),
            })

        return docentes
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener docentes: {str(e)}")
    finally:
        connection.close()


@app.post("/api/suplencias-horarios")
def crear_suplencia_horarios(s: SuplenciaHorarios):
    if s.docente_nombre.strip() == s.suplente_nombre.strip():
        raise HTTPException(status_code=400, detail="El docente no puede ser su propio suplente.")
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id FROM suplencias_horarios
                WHERE suplente_nombre = %s AND activa = TRUE AND fecha = %s
                AND NOT (hora_fin <= %s OR hora_inicio >= %s)
            """, (s.suplente_nombre, s.fecha, s.hora_inicio, s.hora_fin))
            if cursor.fetchone():
                raise HTTPException(status_code=400, detail="El suplente ya tiene una suplencia asignada en ese horario.")
            cursor.execute("""
                INSERT INTO suplencias_horarios
                (docente_nombre, suplente_nombre, materia, dia, fecha, hora_inicio, hora_fin, activa)
                VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE)
            """, (s.docente_nombre, s.suplente_nombre, s.materia, s.dia, s.fecha, s.hora_inicio, s.hora_fin))
        connection.commit()
        return {"message": "Suplencia asignada correctamente."}
    except HTTPException:
        raise
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al crear suplencia: {str(e)}")
    finally:
        connection.close()


@app.delete("/api/suplencias-horarios/{suplencia_id}")
def cancelar_suplencia_horarios(suplencia_id: int):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("UPDATE suplencias_horarios SET activa = FALSE WHERE id = %s", (suplencia_id,))
        connection.commit()
        return {"message": "Suplencia cancelada."}
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al cancelar suplencia: {str(e)}")
    finally:
        connection.close()


@app.get("/api/suplencias-activas")
def obtener_suplencias_activas():
    """Todas las suplencias activas hoy con datos enriquecidos (licenciatura y aula del horario original)."""
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT sh.id, sh.docente_nombre, sh.suplente_nombre, sh.materia,
                       sh.dia, sh.fecha, sh.hora_inicio, sh.hora_fin,
                       MIN(h.licenciatura) AS licenciatura,
                       MIN(h.aula_asignada) AS aula_asignada
                FROM suplencias_horarios sh
                LEFT JOIN horarios h
                    ON h.docente = sh.docente_nombre AND h.asignatura = sh.materia
                WHERE sh.activa = TRUE AND sh.fecha = CURDATE()
                GROUP BY sh.id, sh.docente_nombre, sh.suplente_nombre, sh.materia,
                         sh.dia, sh.fecha, sh.hora_inicio, sh.hora_fin
            """)
            suplencias = cursor.fetchall()
        result = []
        for s in (suplencias or []):
            result.append({
                'id':              s['id'],
                'docente_nombre':  s['docente_nombre'],
                'suplente_nombre': s['suplente_nombre'],
                'materia':         s['materia'],
                'dia':             s['dia'],
                'fecha':           str(s['fecha']),
                'hora_inicio':     _timedelta_to_str(s['hora_inicio']),
                'hora_fin':        _timedelta_to_str(s['hora_fin']),
                'inicio_mins':     _time_to_mins(s['hora_inicio']),
                'fin_mins':        _time_to_mins(s['hora_fin']),
                'licenciatura':    s.get('licenciatura') or '',
                'aula_asignada':   s.get('aula_asignada') or '',
            })
        return result
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener suplencias activas: {str(e)}")
    finally:
        connection.close()


# ---------------------------------------------------------
# ENDPOINTS PARA EXPORTACIÓN, EMAIL, IMPRESIÓN Y AUDITORÍA
# ---------------------------------------------------------
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_decorations(num_pages)
            super().showPage()
        super().save()

    def draw_page_decorations(self, page_count):
        self.saveState()
        # Encabezado institucional azul
        self.setFillColor(colors.HexColor('#1c355e'))
        self.rect(0, 570, 792, 42, fill=True, stroke=False)
        self.setFillColor(colors.white)
        self.setFont("Helvetica-Bold", 11)
        self.drawString(36, 587, "UNIVERSIDAD LATINO  |  SIPREF - REPORTE DE HORARIOS Y EVENTOS ACADÉMICOS")
        
        # Pie de página
        self.setStrokeColor(colors.HexColor('#c5c6cf'))
        self.setLineWidth(0.5)
        self.line(36, 35, 792 - 36, 35)
        self.setFillColor(colors.HexColor('#75777f'))
        self.setFont("Helvetica", 8)
        self.drawString(36, 20, "Sistema de Gestión e Información Académica — Universidad Latino")
        page_text = f"Página {self._pageNumber} de {page_count}"
        self.drawRightString(792 - 36, 20, page_text)
        self.restoreState()


def _generar_pdf_reporte(registros: list[dict], filtros: dict, usuario_nombre: str = "Administrador", es_historial: bool = False) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=36,
        rightMargin=36,
        topMargin=54,
        bottomMargin=54
    )
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=15,
        leading=18,
        textColor=colors.HexColor('#1c355e')
    )
    sub_style = ParagraphStyle(
        'SubTitle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=11,
        textColor=colors.HexColor('#75777f')
    )
    cell_head_style = ParagraphStyle(
        'CellHead',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        leading=10,
        textColor=colors.white,
        alignment=1
    )
    cell_text_style = ParagraphStyle(
        'CellText',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        leading=10,
        textColor=colors.HexColor('#1b1c1e')
    )

    story = []
    ahora_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    tipo_repr = "HISTORIAL COMPLETO DE EVENTOS ACADÉMICOS" if es_historial else "REPORTE DE HORARIOS Y EVENTOS DE AULAS"
    
    story.append(Paragraph(f"<b>UNIVERSIDAD LATINO — {tipo_repr}</b>", title_style))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"Generado por: <b>{usuario_nombre}</b> | Fecha y Hora: <b>{ahora_str}</b> | Total Registros: <b>{len(registros)}</b>", sub_style))
    story.append(Spacer(1, 6))

    if filtros and not es_historial:
        filtros_str_list = [f"<b>{k}:</b> {v}" for k, v in filtros.items() if v]
        if filtros_str_list:
            filtros_text = " | ".join(filtros_str_list)
            filtros_style = ParagraphStyle('FiltrosText', parent=sub_style, fontSize=8, leading=10, textColor=colors.HexColor('#1c355e'))
            story.append(Paragraph(f"<b>Filtros Activos:</b> {filtros_text}", filtros_style))
            story.append(Spacer(1, 8))

    col_widths = [55, 120, 75, 60, 165, 80, 55, 110]
    table_data = [[
        Paragraph("Día", cell_head_style),
        Paragraph("Docente", cell_head_style),
        Paragraph("Licenciatura", cell_head_style),
        Paragraph("Nivel", cell_head_style),
        Paragraph("Asignatura", cell_head_style),
        Paragraph("Horario", cell_head_style),
        Paragraph("Aula", cell_head_style),
        Paragraph("Estado / Motivo", cell_head_style),
    ]]

    for r in registros:
        nivel = f"Sem: {r.get('semestre')}" if r.get('semestre') else (f"Cuat: {r.get('cuatrimestre')}" if r.get('cuatrimestre') else "")
        if r.get('grupo'): nivel += f" ({r.get('grupo')})"
        
        if r.get('estado_label'):
            estado_lbl = str(r.get('estado_label'))
            inc = str(r.get('incidencia_detalle') or '')
            estado_txt = f"{estado_lbl} ({inc})" if inc else estado_lbl
            aula_str = str(r.get('aula_actual') or r.get('aula_original') or r.get('aula_asignada') or '—')
        else:
            estado_txt = str(r.get('estadoTiempo', 'programada')).replace('_', ' ').capitalize()
            if r.get('esAulaLiberada'):
                motivo = r.get('infoLiberacion', {}).get('motivo', 'Salida anticipada') if isinstance(r.get('infoLiberacion'), dict) else 'Aula Liberada'
                estado_txt = f"Aula Liberada ({motivo})"
            elif r.get('es_suplencia'):
                estado_txt = f"Suplencia (Cubre a {r.get('docente_ausente', '')})"
            elif r.get('estadoRazon'):
                estado_txt += f" ({r.get('estadoRazon')})"

            aula_str = r.get('aula_asignada', '—')
            if r.get('esAulaLiberada'):
                aula_str = f"Fuera de Aula ({aula_str})"
            elif r.get('aula_reasignada'):
                aula_str = f"{r.get('aula_asignada')} (Temp)"

        dia_val = str(r.get('diaOriginal') or r.get('dia_semana') or r.get('dia') or '—').capitalize()
        doc_val = str(r.get('docente_nombre') or r.get('docente') or '—')
        lic_val = str(r.get('licenciatura_codigo') or r.get('licenciatura') or '—')
        asig_val = str(r.get('asignatura_nombre') or r.get('asignatura') or '—')
        hora_val = f"{r.get('horario_inicio')}-{r.get('horario_fin')}" if r.get('horario_inicio') and r.get('horario_fin') else str(r.get('textoHora', r.get('horario', '—')))

        row = [
            Paragraph(dia_val, cell_text_style),
            Paragraph(doc_val, cell_text_style),
            Paragraph(lic_val, cell_text_style),
            Paragraph(nivel or '—', cell_text_style),
            Paragraph(asig_val, cell_text_style),
            Paragraph(hora_val, cell_text_style),
            Paragraph(str(aula_str), cell_text_style),
            Paragraph(str(estado_txt), cell_text_style),
        ]
        table_data.append(row)

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1c355e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e8')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    story.append(t)
    doc.build(story, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer.getvalue()


def _generar_excel_reporte(registros: list[dict], filtros: dict, usuario_nombre: str = "Administrador", es_historial: bool = False) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Horarios"

    header_fill = PatternFill(start_color="1C355E", end_color="1C355E", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1C355E")
    meta_font = Font(name="Calibri", size=9, italic=True, color="555555")
    cell_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E0E0E8'),
        right=Side(style='thin', color='E0E0E8'),
        top=Side(style='thin', color='E0E0E8'),
        bottom=Side(style='thin', color='E0E0E8')
    )
    zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    tipo_repr = "HISTORIAL COMPLETO DE EVENTOS ACADÉMICOS" if es_historial else "REPORTE DE HORARIOS Y EVENTOS DE AULAS"
    ws.cell(row=1, column=1, value=f"UNIVERSIDAD LATINO — {tipo_repr}").font = title_font
    
    ahora_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row=2, column=1, value=f"Generado por: {usuario_nombre} | Fecha y hora: {ahora_str} | Total registros: {len(registros)}").font = meta_font

    row_offset = 4
    if filtros and not es_historial:
        filtros_str = ", ".join([f"{k}: {v}" for k, v in filtros.items() if v])
        if filtros_str:
            ws.cell(row=3, column=1, value=f"Filtros aplicados: {filtros_str}").font = meta_font
            row_offset = 5

    headers = ["Día", "Docente", "Licenciatura", "Semestre", "Cuatrimestre", "Grupo", "Asignatura", "Horario", "Aula", "Estado", "Detalle / Motivo"]
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=row_offset, column=col_num, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r_idx, r in enumerate(registros, start=row_offset + 1):
        if r.get('estado_label'):
            estado_txt = str(r.get('estado_label'))
            motivo = str(r.get('incidencia_detalle') or '')
            aula_str = str(r.get('aula_actual') or r.get('aula_original') or r.get('aula_asignada') or '—')
        else:
            estado_txt = str(r.get('estadoTiempo', 'programada')).replace('_', ' ').capitalize()
            motivo = ""
            if r.get('esAulaLiberada'):
                estado_txt = "Aula Liberada"
                motivo = r.get('infoLiberacion', {}).get('motivo', 'Salida anticipada') if isinstance(r.get('infoLiberacion'), dict) else 'Salida anticipada'
            elif r.get('es_suplencia'):
                estado_txt = "Suplencia"
                motivo = f"Cubre a {r.get('docente_ausente', '')}"
            elif r.get('estadoRazon'):
                motivo = str(r.get('estadoRazon'))

            aula_str = str(r.get('aula_asignada', '—'))
            if r.get('esAulaLiberada'):
                aula_str = f"Fuera de Aula ({aula_str})"

        values = [
            str(r.get('diaOriginal') or r.get('dia_semana') or r.get('dia') or '—').capitalize(),
            str(r.get('docente_nombre') or r.get('docente') or '—'),
            str(r.get('licenciatura_codigo') or r.get('licenciatura') or '—'),
            str(r.get('semestre', '—')),
            str(r.get('cuatrimestre', '—')),
            str(r.get('grupo', '—')),
            str(r.get('asignatura_nombre') or r.get('asignatura') or '—'),
            f"{r.get('horario_inicio')}-{r.get('horario_fin')}" if r.get('horario_inicio') and r.get('horario_fin') else str(r.get('textoHora', r.get('horario', '—'))),
            aula_str,
            estado_txt,
            motivo
        ]

        use_zebra = (r_idx % 2 == 0)
        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = cell_font
            cell.border = thin_border
            if use_zebra:
                cell.fill = zebra_fill
            if c_idx in [1, 4, 5, 6, 8, 9, 10]:
                cell.alignment = Alignment(horizontal="center")

    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


class ExportarRequest(BaseModel):
    usuario_nombre: str = "Administrador"
    usuario_correo: str = "admin@universidadlatino.edu.mx"
    es_historial_completo: bool = False
    filtros_aplicados: dict = {}
    registros: list = []
    licenciatura: Optional[str] = None
    semestre: Optional[str] = None
    asignatura: Optional[str] = None
    filtros_bitacora: Optional[dict] = None
    filtros_aplicados: Optional[dict] = {}
    registros: list[dict] = []
    formato: Optional[str] = "excel"

class EnviarReporteEmailRequest(BaseModel):
    usuario_nombre: Optional[str] = "Administrador"
    usuario_correo: Optional[str] = ""
    destinatarios: str
    cc: Optional[str] = ""
    cco: Optional[str] = ""
    asunto: Optional[str] = "Reporte de Horarios y Eventos Académicos — ULA"
    mensaje: Optional[str] = ""
    adjuntar_pdf: Optional[bool] = True
    adjuntar_excel: Optional[bool] = False
    es_historial_completo: Optional[bool] = False
    licenciatura: Optional[str] = None
    semestre: Optional[str] = None
    asignatura: Optional[str] = None
    filtros_aplicados: Optional[dict] = {}
    registros: list[dict] = []


def _registrar_auditoria_exportacion(usuario_nombre, usuario_correo, tipo, es_historial, filtros, cantidad, enviado_email=False, destinatarios=""):
    connection = get_db_connection()
    try:
        filtros_str = json.dumps(filtros, ensure_ascii=False) if isinstance(filtros, dict) else str(filtros)
        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO auditoria_exportaciones
                (usuario_nombre, usuario_correo, tipo_exportacion, es_historial_completo, filtros_aplicados, cantidad_registros, enviado_email, destinatarios_email)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """, (usuario_nombre or 'Administrador', usuario_correo or '', tipo, es_historial, filtros_str, cantidad, enviado_email, destinatarios))
            connection.commit()
    except Exception as e:
        print(f"Error al registrar auditoria de exportacion: {e}")
    finally:
        connection.close()


@app.get("/api/historial/filtros-disponibles")
def obtener_filtros_disponibles_historial():
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT licenciatura_codigo FROM clases_historico WHERE licenciatura_codigo IS NOT NULL AND TRIM(licenciatura_codigo) != '' ORDER BY licenciatura_codigo")
            lics = [row['licenciatura_codigo'] if isinstance(row, dict) else row[0] for row in cursor.fetchall()]
            
            cursor.execute("SELECT DISTINCT semestre FROM clases_historico WHERE semestre IS NOT NULL AND TRIM(semestre) != '' ORDER BY semestre")
            sems = [str(row['semestre']) if isinstance(row, dict) else str(row[0]) for row in cursor.fetchall()]
            
            cursor.execute("SELECT DISTINCT asignatura_nombre FROM clases_historico WHERE asignatura_nombre IS NOT NULL AND TRIM(asignatura_nombre) != '' ORDER BY asignatura_nombre")
            asigs = [row['asignatura_nombre'] if isinstance(row, dict) else row[0] for row in cursor.fetchall()]
            
            return {
                "licenciaturas": lics,
                "semestres": sems,
                "asignaturas": asigs
            }
    except Exception as e:
        print(f"Error en filtros-disponibles historial: {e}")
        return {"licenciaturas": [], "semestres": [], "asignaturas": []}
    finally:
        connection.close()



def _consultar_bitacora_para_exportacion(filtros: dict) -> list[dict]:
    fecha_inicio = filtros.get('fecha_inicio') if filtros else None
    fecha_fin = filtros.get('fecha_fin') if filtros else None
    modulo = filtros.get('modulo') if filtros else None
    usuario = filtros.get('usuario') if filtros else None

    connection = get_db_connection()
    try:
        query = "SELECT * FROM bitacora_cambios WHERE 1=1"
        params = []
        if fecha_inicio:
            query += " AND DATE(fecha_cambio) >= %s"
            params.append(fecha_inicio)
        if fecha_fin:
            query += " AND DATE(fecha_cambio) <= %s"
            params.append(fecha_fin)
        if modulo and modulo.lower() != "todos":
            query += " AND modulo_afectado = %s"
            params.append(modulo.lower())
        if usuario and usuario.lower() != "todos":
            query += " AND usuario = %s"
            params.append(usuario)
            
        query += " ORDER BY fecha_cambio DESC LIMIT 1000"
        with connection.cursor() as cursor:
            cursor.execute(query, params)
            registros = cursor.fetchall()
        for r in (registros or []):
            if r.get('fecha_cambio'):
                r['fecha_cambio'] = r['fecha_cambio'].strftime("%Y-%m-%d %H:%M:%S")
        return registros if registros else []
    except Exception as e:
        print(f"Error consultando bitacora para exportacion: {e}")
        return []
    finally:
        connection.close()

def _generar_pdf_bitacora(registros: list[dict], filtros: dict, usuario_nombre: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), leftMargin=36, rightMargin=36, topMargin=54, bottomMargin=54)
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle('DocTitle', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=15, leading=18, textColor=colors.HexColor('#1c355e'))
    sub_style = ParagraphStyle('SubTitle', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=11, textColor=colors.HexColor('#75777f'))
    cell_head_style = ParagraphStyle('CellHead', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=8, leading=10, textColor=colors.white, alignment=1)
    cell_text_style = ParagraphStyle('CellText', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10, textColor=colors.HexColor('#1b1c1e'))

    story = []
    ahora_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    story.append(Paragraph("<b>UNIVERSIDAD LATINO — BITÁCORA DE CAMBIOS DEL SISTEMA</b>", title_style))
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"Generado por: <b>{usuario_nombre}</b> | Fecha y Hora: <b>{ahora_str}</b> | Total Registros: <b>{len(registros)}</b>", sub_style))
    story.append(Spacer(1, 6))

    if filtros:
        filtros_str_list = [f"<b>{k}:</b> {v}" for k, v in filtros.items() if v]
        if filtros_str_list:
            filtros_text = " | ".join(filtros_str_list)
            filtros_style = ParagraphStyle('FiltrosText', parent=sub_style, fontSize=8, leading=10, textColor=colors.HexColor('#1c355e'))
            story.append(Paragraph(f"<b>Filtros Activos:</b> {filtros_text}", filtros_style))
            story.append(Spacer(1, 8))

    col_widths = [90, 80, 80, 70, 80, 220, 100]
    table_data = [[
        Paragraph("Fecha", cell_head_style),
        Paragraph("Usuario", cell_head_style),
        Paragraph("Módulo", cell_head_style),
        Paragraph("Acción", cell_head_style),
        Paragraph("Entidad", cell_head_style),
        Paragraph("Detalles", cell_head_style),
        Paragraph("IP", cell_head_style),
    ]]

    for r in registros:
        det_txt = str(r.get('detalles_json') or '')
        if len(det_txt) > 200: det_txt = det_txt[:197] + "..."
        row = [
            Paragraph(str(r.get('fecha_cambio') or '—'), cell_text_style),
            Paragraph(str(r.get('usuario') or '—'), cell_text_style),
            Paragraph(str(r.get('modulo_afectado') or '—').capitalize(), cell_text_style),
            Paragraph(str(r.get('accion') or '—'), cell_text_style),
            Paragraph(f"{r.get('entidad_tipo') or ''} {r.get('entidad_id') or ''}", cell_text_style),
            Paragraph(det_txt, cell_text_style),
            Paragraph(str(r.get('ip_address') or '—'), cell_text_style),
        ]
        table_data.append(row)

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1c355e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e8')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    story.append(t)
    doc.build(story, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer.getvalue()

def _generar_excel_bitacora(registros: list[dict], filtros: dict, usuario_nombre: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bitácora"

    header_fill = PatternFill(start_color="1C355E", end_color="1C355E", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1C355E")
    meta_font = Font(name="Calibri", size=9, italic=True, color="555555")
    cell_font = Font(name="Calibri", size=10)
    thin_border = Border(left=Side(style='thin', color='E0E0E8'), right=Side(style='thin', color='E0E0E8'), top=Side(style='thin', color='E0E0E8'), bottom=Side(style='thin', color='E0E0E8'))

    ws.cell(row=1, column=1, value="UNIVERSIDAD LATINO — BITÁCORA DE CAMBIOS DEL SISTEMA").font = title_font
    ws.cell(row=2, column=1, value=f"Generado por: {usuario_nombre} | Fecha y hora: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total registros: {len(registros)}").font = meta_font

    row_offset = 4
    if filtros:
        filtros_str = ", ".join([f"{k}: {v}" for k, v in filtros.items() if v])
        if filtros_str:
            ws.cell(row=3, column=1, value=f"Filtros aplicados: {filtros_str}").font = meta_font
            row_offset = 5

    headers = ["Fecha", "Usuario", "Módulo", "Acción", "Entidad Tipo", "Entidad ID", "Detalles", "IP"]
    for col_num, h in enumerate(headers, 1):
        c = ws.cell(row=row_offset, column=col_num, value=h)
        c.fill = header_fill
        c.font = header_font
        c.border = thin_border
        
    for i, r in enumerate(registros):
        row_num = row_offset + 1 + i
        vals = [
            str(r.get('fecha_cambio') or '—'),
            str(r.get('usuario') or '—'),
            str(r.get('modulo_afectado') or '—').capitalize(),
            str(r.get('accion') or '—'),
            str(r.get('entidad_tipo') or '—'),
            str(r.get('entidad_id') or '—'),
            str(r.get('detalles_json') or '—'),
            str(r.get('ip_address') or '—')
        ]
        for col_num, val in enumerate(vals, 1):
            c = ws.cell(row=row_num, column=col_num, value=val)
            c.font = cell_font
            c.border = thin_border
            
    col_w = [20, 20, 15, 15, 15, 10, 40, 15]
    for i, cw in enumerate(col_w, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = cw

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def _consultar_historial_para_exportacion(licenciatura: Optional[str] = None, semestre: Optional[str] = None, asignatura: Optional[str] = None):
    _evaluar_y_cerrar_clases_vencidas()
    query = "SELECT * FROM clases_historico WHERE 1=1"
    params = []
    if licenciatura and str(licenciatura).strip() not in ("", "Todos", "Todas las Licenciaturas"):
        query += " AND licenciatura_codigo = %s"
        params.append(str(licenciatura).strip())
    if semestre and str(semestre).strip() not in ("", "Todos", "Todos los Grados"):
        query += " AND semestre = %s"
        params.append(str(semestre).strip())
    if asignatura and str(asignatura).strip() not in ("", "Todos", "Todas las Asignaturas"):
        query += " AND asignatura_nombre = %s"
        params.append(str(asignatura).strip())
    query += " ORDER BY fecha DESC, horario_inicio ASC"
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(query, params)
            filas = cursor.fetchall() or []
        resultado = []
        for r in filas:
            resultado.append({
                "id": r["id"],
                "clase_id": r.get("clase_id"),
                "fecha": str(r["fecha"]),
                "dia_semana": r["dia_semana"],
                "semana_numero": r["semana_numero"],
                "docente_nombre": r["docente_nombre"],
                "licenciatura_codigo": r["licenciatura_codigo"],
                "semestre": r["semestre"],
                "grupo": r["grupo"],
                "asignatura_nombre": r["asignatura_nombre"],
                "horario_inicio": _timedelta_to_str(r["horario_inicio"]),
                "horario_fin": _timedelta_to_str(r["horario_fin"]),
                "aula_original": r["aula_original"],
                "aula_actual": r["aula_actual"],
                "estado_slug": r["estado_slug"],
                "estado_label": r["estado_label"],
                "incidencia_detalle": r.get("incidencia_detalle") or ""
            })
        return resultado
    except Exception as e:
        print(f"Error consultando historial para exportacion: {e}")
        return []
    finally:
        connection.close()


@app.post("/api/exportar/pdf")
def exportar_pdf(req: ExportarRequest):
    try:
        if req.es_historial_completo:
            registros_data = _consultar_bitacora_para_exportacion(req.filtros_bitacora)
            pdf_bytes = _generar_pdf_bitacora(registros_data, req.filtros_aplicados, req.usuario_nombre)
        else:
            registros_data = req.registros
            pdf_bytes = _generar_pdf_reporte(registros_data, req.filtros_aplicados, req.usuario_nombre, False)
        
        _registrar_auditoria_exportacion(req.usuario_nombre, req.usuario_correo, "PDF", req.es_historial_completo, req.filtros_aplicados, len(registros_data))
        
        fecha_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"Bitacora_Eventos_{fecha_str}.pdf" if req.es_historial_completo else f"Dashboard_Clases_{fecha_str}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(e)}")


@app.post("/api/exportar/excel")
def exportar_excel(req: ExportarRequest):
    try:
        if req.es_historial_completo:
            registros_data = _consultar_bitacora_para_exportacion(req.filtros_bitacora)
            excel_bytes = _generar_excel_bitacora(registros_data, req.filtros_aplicados, req.usuario_nombre)
        else:
            registros_data = req.registros
            excel_bytes = _generar_excel_reporte(registros_data, req.filtros_aplicados, req.usuario_nombre, False)

        _registrar_auditoria_exportacion(req.usuario_nombre, req.usuario_correo, "Excel", req.es_historial_completo, req.filtros_aplicados, len(registros_data))

        fecha_str = datetime.now().strftime("%Y-%m-%d")
        filename = f"Bitacora_Eventos_{fecha_str}.xlsx" if req.es_historial_completo else f"Dashboard_Clases_{fecha_str}.xlsx"
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar Excel: {str(e)}")


def _generar_excel_docentes(registros: list[dict], filtros_str: str, usuario_nombre: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Directorio de Docentes"

    header_fill = PatternFill(start_color="1C355E", end_color="1C355E", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="1C355E")
    meta_font = Font(name="Calibri", size=9, italic=True, color="555555")
    cell_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='E0E0E8'),
        right=Side(style='thin', color='E0E0E8'),
        top=Side(style='thin', color='E0E0E8'),
        bottom=Side(style='thin', color='E0E0E8')
    )
    zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    ws.cell(row=1, column=1, value="UNIVERSIDAD LATINO — DIRECTORIO DE DOCENTES Y ESTADO ACTUAL").font = title_font
    
    ahora_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.cell(row=2, column=1, value=f"Generado por: {usuario_nombre} | Fecha y hora: {ahora_str} | Total registros: {len(registros)}").font = meta_font

    row_offset = 4
    if filtros_str:
        ws.cell(row=3, column=1, value=f"Filtros aplicados: {filtros_str}").font = meta_font
        row_offset = 5

    headers = ["Docente", "Estado Actual", "Licenciaturas", "Asignaturas"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_offset, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    col_widths = {"A": 40, "B": 25, "C": 50, "D": 60}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    for i, r in enumerate(registros, start=1):
        row_num = row_offset + i
        ws.cell(row=row_num, column=1, value=r.get('nombre', '')).font = cell_font
        ws.cell(row=row_num, column=2, value=r.get('estado', '')).font = cell_font
        ws.cell(row=row_num, column=3, value=r.get('licenciaturas', '')).font = cell_font
        ws.cell(row=row_num, column=4, value=r.get('asignaturas', '')).font = cell_font

        for col_num in range(1, len(headers) + 1):
            c = ws.cell(row=row_num, column=col_num)
            c.border = thin_border
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if i % 2 == 0:
                c.fill = zebra_fill

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


@app.post("/api/exportar/docentes-excel")
def exportar_docentes_excel(req: ExportarRequest):
    try:
        filtros_str = ""
        if req.filtros_aplicados and isinstance(req.filtros_aplicados, dict):
            filtros_str = ", ".join([f"{k}: {v}" for k, v in req.filtros_aplicados.items() if v])
        elif req.filtros_aplicados:
            filtros_str = str(req.filtros_aplicados)

        fecha_str = datetime.now().strftime("%Y-%m-%d")
        formato = req.formato or "excel"

        if formato == 'pdf':
            pdf_bytes = _generar_pdf_docentes(req.registros, filtros_str, req.usuario_nombre or "Administrador")
            filename = f"Directorio_Docentes_{fecha_str}.pdf"
            return Response(
                content=pdf_bytes,
                media_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename="{filename}"'}
            )

        excel_bytes = _generar_excel_docentes(req.registros, filtros_str, req.usuario_nombre or "Administrador")
        filename = f"Directorio_Docentes_{fecha_str}.xlsx"
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al generar reporte: {str(e)}")


def _generar_pdf_docentes(registros: list[dict], filtros_str: str, usuario_nombre: str) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=60, bottomMargin=40)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle('TitleDoc', parent=styles['Title'], fontSize=14, textColor=colors.HexColor('#1c355e'), spaceAfter=6)
    sub_style = ParagraphStyle('SubDoc', parent=styles['Normal'], fontSize=8, textColor=colors.HexColor('#555555'), spaceAfter=4)
    cell_head_style = ParagraphStyle('CellHeadDoc', parent=styles['Normal'], fontSize=8, textColor=colors.white, fontName='Helvetica-Bold')
    cell_text_style = ParagraphStyle('CellTextDoc', parent=styles['Normal'], fontSize=8, leading=10)

    story.append(Paragraph("UNIVERSIDAD LATINO — DIRECTORIO DE DOCENTES", title_style))
    ahora_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    story.append(Paragraph(f"Generado por: {usuario_nombre} | {ahora_str} | Total: {len(registros)} docentes", sub_style))
    if filtros_str:
        story.append(Paragraph(f"<b>Filtros:</b> {filtros_str}", sub_style))
    story.append(Spacer(1, 12))

    col_widths = [200, 100, 200, 250]
    table_data = [[
        Paragraph("Docente", cell_head_style),
        Paragraph("Estado", cell_head_style),
        Paragraph("Licenciaturas", cell_head_style),
        Paragraph("Asignaturas", cell_head_style),
    ]]

    for r in registros:
        row = [
            Paragraph(str(r.get('nombre', '—')), cell_text_style),
            Paragraph(str(r.get('estado', '—')), cell_text_style),
            Paragraph(str(r.get('licenciaturas', '—')), cell_text_style),
            Paragraph(str(r.get('asignaturas', '—')), cell_text_style),
        ]
        table_data.append(row)

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1c355e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e0e0e8')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
    ]))
    story.append(t)
    doc.build(story, canvasmaker=NumberedCanvas)
    buffer.seek(0)
    return buffer.getvalue()


class ExportarDocentesEmailRequest(BaseModel):
    usuario_nombre: str = "Administrador"
    usuario_correo: str = "admin@universidadlatino.edu.mx"
    destinatarios: str
    cc: Optional[str] = ""
    cco: Optional[str] = ""
    asunto: Optional[str] = "Reporte de Docentes — ULA"
    mensaje: Optional[str] = ""
    adjuntar_pdf: Optional[bool] = True
    adjuntar_excel: Optional[bool] = False
    filtros_aplicados: Optional[dict] = {}
    registros: list[dict] = []


@app.post("/api/exportar/docentes-email")
def exportar_docentes_email(req: ExportarDocentesEmailRequest):
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    try:
        filtros_str = ""
        if req.filtros_aplicados and isinstance(req.filtros_aplicados, dict):
            filtros_str = ", ".join([f"{k}: {v}" for k, v in req.filtros_aplicados.items() if v])

        fecha_str = datetime.now().strftime("%Y-%m-%d")

        msg = MIMEMultipart()
        msg['From'] = os.getenv("SMTP_FROM", req.usuario_correo)
        msg['To'] = req.destinatarios
        if req.cc:
            msg['Cc'] = req.cc
        msg['Subject'] = req.asunto or f"Directorio de Docentes — {fecha_str}"
        body = req.mensaje or f"Se adjunta el reporte del directorio de docentes generado el {fecha_str}."
        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        if req.adjuntar_pdf:
            pdf_bytes = _generar_pdf_docentes(req.registros, filtros_str, req.usuario_nombre)
            part = MIMEBase('application', 'pdf')
            part.set_payload(pdf_bytes)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="Directorio_Docentes_{fecha_str}.pdf"')
            msg.attach(part)

        if req.adjuntar_excel:
            excel_bytes = _generar_excel_docentes(req.registros, filtros_str, req.usuario_nombre)
            part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            part.set_payload(excel_bytes)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="Directorio_Docentes_{fecha_str}.xlsx"')
            msg.attach(part)

        smtp_server = os.getenv("SMTP_HOST", "smtp.gmail.com")
        smtp_port = int(os.getenv("SMTP_PORT", "587"))
        smtp_user = os.getenv("SMTP_USER", "")
        smtp_pass = os.getenv("SMTP_PASS", "")

        all_recipients = [e.strip() for e in req.destinatarios.split(',') if e.strip()]
        if req.cc:
            all_recipients += [e.strip() for e in req.cc.split(',') if e.strip()]
        if req.cco:
            all_recipients += [e.strip() for e in req.cco.split(',') if e.strip()]

        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            if smtp_user and smtp_pass:
                server.login(smtp_user, smtp_pass)
            server.sendmail(msg['From'], all_recipients, msg.as_string())

        return {"message": "Correo enviado exitosamente", "destinatarios": len(all_recipients)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar correo: {str(e)}")


REPORTES_COMPARTIDOS_CACHE = {}

def _guardar_reporte_compartido(nombre_archivo: str, tipo: str, contenido: bytes) -> str:
    report_id = str(uuid.uuid4())
    REPORTES_COMPARTIDOS_CACHE[report_id] = {
        "nombre_archivo": nombre_archivo,
        "tipo": tipo,
        "contenido": contenido
    }
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute("""
                    CREATE TABLE IF NOT EXISTS reportes_compartidos (
                        id VARCHAR(64) PRIMARY KEY,
                        fecha_hora DATETIME DEFAULT NOW(),
                        nombre_archivo VARCHAR(255) NOT NULL,
                        tipo VARCHAR(50) NOT NULL,
                        contenido LONGBLOB NOT NULL
                    )
                """)
                cursor.execute("""
                    INSERT INTO reportes_compartidos (id, nombre_archivo, tipo, contenido)
                    VALUES (%s, %s, %s, %s)
                """, (report_id, nombre_archivo, tipo, contenido))
            conn.commit()
    except Exception as e:
        print(f"Advertencia al guardar reporte compartido en DB: {e}")
    return report_id


@app.get("/api/reportes/descargar/{report_id}")
def descargar_reporte_compartido(report_id: str):
    data = REPORTES_COMPARTIDOS_CACHE.get(report_id)
    if not data:
        try:
            with get_db_connection() as conn:
                with conn.cursor() as cursor:
                    cursor.execute("""
                        SELECT nombre_archivo, tipo, contenido FROM reportes_compartidos WHERE id = %s
                    """, (report_id,))
                    row = cursor.fetchone()
                    if row:
                        data = {
                            "nombre_archivo": row[0],
                            "tipo": row[1],
                            "contenido": row[2]
                        }
                        REPORTES_COMPARTIDOS_CACHE[report_id] = data
        except Exception as e:
            print(f"Error al consultar reporte compartido de DB: {e}")

    if not data:
        raise HTTPException(status_code=404, detail="El reporte solicitado no existe o ha expirado.")

    media_type = "application/pdf" if data["tipo"] == "pdf" else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    headers = {
        "Content-Disposition": f'attachment; filename="{data["nombre_archivo"]}"'
    }
    return Response(content=data["contenido"], media_type=media_type, headers=headers)


@app.post("/api/exportar/email")
def enviar_reporte_email(req: EnviarReporteEmailRequest, request: Request):
    try:
        destinos = [d.strip() for d in req.destinatarios.replace(';', ',').split(',') if d.strip()]
        if not destinos:
            raise HTTPException(status_code=400, detail="Debe especificar al menos un correo destinatario.")
        
        cc_list = [d.strip() for d in (req.cc or '').replace(';', ',').split(',') if d.strip()]
        cco_list = [d.strip() for d in (req.cco or '').replace(';', ',').split(',') if d.strip()]

        fecha_str = datetime.now().strftime("%Y-%m-%d")
        asunto = req.asunto or f"Reporte de Eventos Académicos ULA - {fecha_str}"

        dest_str = f"To: {','.join(destinos)}"
        if cc_list: dest_str += f" | CC: {','.join(cc_list)}"
        if cco_list: dest_str += f" | CCO: {','.join(cco_list)}"
        tipo_str = "Email PDF+Excel" if (req.adjuntar_pdf and req.adjuntar_excel) else ("Email PDF" if req.adjuntar_pdf else "Email Excel")

        registros_data = _consultar_historial_para_exportacion(req.licenciatura, req.semestre, req.asignatura) if req.es_historial_completo else req.registros

        pdf_bytes = None
        xlsx_bytes = None
        id_pdf = None
        id_excel = None

        if req.adjuntar_pdf:
            pdf_bytes = _generar_pdf_reporte(registros_data, req.filtros_aplicados, req.usuario_nombre, req.es_historial_completo)
            id_pdf = _guardar_reporte_compartido(f"Reporte_Academico_{fecha_str}.pdf", "pdf", pdf_bytes)

        if req.adjuntar_excel:
            xlsx_bytes = _generar_excel_reporte(registros_data, req.filtros_aplicados, req.usuario_nombre, req.es_historial_completo)
            id_excel = _guardar_reporte_compartido(f"Reporte_Academico_{fecha_str}.xlsx", "xlsx", xlsx_bytes)

        backend_url = os.getenv("RENDER_EXTERNAL_URL") or os.getenv("BACKEND_URL") or str(request.base_url).rstrip('/')
        if backend_url.endswith('/'):
            backend_url = backend_url[:-1]

        botones_html = ""
        if id_pdf:
            url_pdf = f"{backend_url}/api/reportes/descargar/{id_pdf}"
            botones_html += f"""
            <a href="{url_pdf}" style="background-color: #1c355e; color: #ffffff; padding: 12px 24px; text-decoration: none; border-radius: 6px; font-weight: bold; display: inline-block; margin: 6px; font-size: 14px;">
                📄 Descargar Reporte PDF
            </a>
            """
        if id_excel:
            url_excel = f"{backend_url}/api/reportes/descargar/{id_excel}"
            botones_html += f"""
            <a href="{url_excel}" style="background-color: #16a34a; color: #ffffff; padding: 12px 24px; text-decoration: none; border-radius: 6px; font-weight: bold; display: inline-block; margin: 6px; font-size: 14px;">
                📊 Descargar Reporte Excel
            </a>
            """

        body_html = f"""
        <html>
        <body style="font-family: Arial, sans-serif; color: #333; background-color: #f4f6fb; padding: 20px;">
            <div style="max-width: 600px; margin: 0 auto; background: #ffffff; padding: 24px; border-radius: 10px; border: 1px solid #e0e6ed;">
                <h2 style="color: #1c355e; margin-top: 0; border-bottom: 2px solid #1c355e; padding-bottom: 12px;">
                    Reporte de Horarios y Eventos Académicos — ULA
                </h2>
                <p style="font-size: 15px;">Hola,</p>
                <p style="font-size: 15px; color: #475569;">
                    {req.mensaje or 'El reporte generado por el Sistema de Horarios ULA está listo para su descarga.'}
                </p>
                <div style="margin: 28px 0; text-align: center;">
                    {botones_html}
                </div>
                <hr style="border: 0; border-top: 1px solid #e2e8f0; margin: 24px 0;" />
                <p style="font-size: 12px; color: #64748b; margin-bottom: 0;">
                    Generado por <strong>{req.usuario_nombre}</strong> ({req.usuario_correo}) el {datetime.now().strftime('%d/%m/%Y %H:%M')}.
                </p>
            </div>
        </body>
        </html>
        """

        # Intentar envío SMTP si está configurado
        sender_email = os.getenv("SMTP_USER", "")
        smtp_host = os.getenv("SMTP_HOST", "smtp.gmail.com")
        smtp_port = int(os.getenv("SMTP_PORT", 587))
        smtp_pass = os.getenv("SMTP_PASSWORD", "") or os.getenv("SMTP_PASS", "")

        if not sender_email:
            sender_email = "soporte.sipref.software@gmail.com"

        from email.mime.base import MIMEBase
        from email import encoders
        
        msg = MIMEMultipart()
        msg['From'] = f"Sistema Horarios ULA <{sender_email}>"
        msg['To'] = ", ".join(destinos)
        if cc_list: msg['Cc'] = ", ".join(cc_list)
        msg['Subject'] = asunto
        msg.attach(MIMEText(body_html, 'html'))

        if pdf_bytes:
            part = MIMEBase('application', 'pdf')
            part.set_payload(pdf_bytes)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="Reporte_Academico_{fecha_str}.pdf"')
            msg.attach(part)

        if xlsx_bytes:
            part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            part.set_payload(xlsx_bytes)
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename="Reporte_Academico_{fecha_str}.xlsx"')
            msg.attach(part)

        enviado_smtp = False
        if sender_email and smtp_pass:
            try:
                with smtplib.SMTP(smtp_host, smtp_port, timeout=4.0) as server:
                    server.starttls()
                    server.login(sender_email, smtp_pass)
                    server.sendmail(sender_email, destinos + cc_list + cco_list, msg.as_string())
                enviado_smtp = True
            except Exception as e_smtp:
                print(f"SMTP no disponible en Render o falló ({e_smtp}), usando fallback de EmailJS...")

        if not enviado_smtp:
            for dest in (destinos + cc_list + cco_list):
                _enviar_smtp(dest, asunto, body_html)

        _registrar_auditoria_exportacion(req.usuario_nombre, req.usuario_correo, tipo_str, req.es_historial_completo, req.filtros_aplicados, len(req.registros), enviado_email=True, destinatarios=dest_str)
        return {"message": f"Reporte enviado exitosamente a {', '.join(destinos)}"}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al enviar correo: {str(e)}")


@app.get("/api/exportar/auditoria")
def obtener_auditoria_exportaciones():
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT id, usuario_nombre, usuario_correo, fecha_hora, tipo_exportacion, es_historial_completo, filtros_aplicados, cantidad_registros, enviado_email, destinatarios_email
                FROM auditoria_exportaciones
                ORDER BY id DESC
                LIMIT 200
            """)
            return cursor.fetchall() or []
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener auditoría: {str(e)}")
    finally:
        connection.close()


# ---------------------------------------------------------
# PERSISTENCIA HISTÓRICA DE CLASES (clases_historico)
# ---------------------------------------------------------

def _norm_time(t) -> str:
    if not t:
        return ""
    try:
        parts = str(t).strip().split(":")
        if len(parts) >= 2:
            return f"{int(parts[0]):02d}:{int(parts[1]):02d}"
    except Exception:
        pass
    return str(t)[:5]

def _guardar_registro_historico(
    cursor,
    clase_id: Optional[int],
    fecha_obj: date,
    docente_nombre: str,
    licenciatura_codigo: str,
    semestre: str,
    grupo: str,
    asignatura_nombre: str,
    horario_inicio_str: str,
    horario_fin_str: str,
    aula_original: str,
    aula_actual: str,
    estado_slug: str,
    estado_label: str,
    incidencia_detalle: str = "",
    historico_map: Optional[dict] = None
):
    try:
        dias_es = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        dia_semana = dias_es[fecha_obj.weekday()]
        semana_numero = fecha_obj.isocalendar()[1]

        cache_key = (docente_nombre, asignatura_nombre, _norm_time(horario_inicio_str))
        existing = None
        if historico_map is not None and cache_key in historico_map:
            ex_row = historico_map[cache_key]
            ex_id = ex_row['id']
            if ex_row.get('estado_slug') == estado_slug and (ex_row.get('incidencia_detalle') or '') == incidencia_detalle and (ex_row.get('aula_actual') or '') == (aula_actual or ''):
                return
            existing = {'id': ex_id}
        elif historico_map is None:
            cursor.execute("""
                SELECT id FROM clases_historico
                WHERE fecha = %s AND docente_nombre = %s AND asignatura_nombre = %s AND horario_inicio = %s
            """, (fecha_obj, docente_nombre, asignatura_nombre, horario_inicio_str))
            existing = cursor.fetchone()

        if existing:
            ex_id = existing['id'] if isinstance(existing, dict) else existing[0]
            cursor.execute("""
                UPDATE clases_historico
                SET estado_slug = %s, estado_label = %s, incidencia_detalle = %s, aula_actual = %s
                WHERE id = %s
            """, (estado_slug, estado_label, incidencia_detalle, aula_actual or '', ex_id))
            if historico_map is not None and cache_key in historico_map:
                historico_map[cache_key]['estado_slug'] = estado_slug
                historico_map[cache_key]['incidencia_detalle'] = incidencia_detalle
                historico_map[cache_key]['aula_actual'] = aula_actual or ''
            return

        cursor.execute("""
            INSERT INTO clases_historico (
                clase_id, fecha, dia_semana, semana_numero, docente_nombre,
                licenciatura_codigo, semestre, grupo, asignatura_nombre,
                horario_inicio, horario_fin, aula_original, aula_actual,
                estado_slug, estado_label, incidencia_detalle
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            clase_id, fecha_obj, dia_semana, semana_numero, docente_nombre,
            licenciatura_codigo or '', semestre or '', grupo or '', asignatura_nombre,
            horario_inicio_str, horario_fin_str, aula_original or '', aula_actual or '',
            estado_slug, estado_label, incidencia_detalle
        ))
    except Exception as e:
        print(f"Advertencia guardando snapshot historico: {e}")


def _evaluar_y_cerrar_clases_vencidas():
    """Evalúa clases de hoy cuyo horario_fin ya haya transcurrido y las persiste en clases_historico."""
    ahora = datetime.now()
    fecha_hoy = ahora.date()
    dia_hoy_idx = ahora.isoweekday() % 7
    mins_ahora = ahora.hour * 60 + ahora.minute

    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            # Precargar aulas liberadas y suplencias del día en diccionarios O(1)
            cursor.execute("SELECT aula_nombre, motivo FROM aulas_liberadas WHERE fecha = %s AND activa = TRUE", (fecha_hoy,))
            aulas_lib_raw = cursor.fetchall() or []
            aulas_lib_map = {r['aula_nombre'] if isinstance(r, dict) else r[0]: (r['motivo'] if isinstance(r, dict) else r[1]) for r in aulas_lib_raw}

            cursor.execute("SELECT docente_nombre, materia, suplente_nombre FROM suplencias_horarios WHERE fecha = %s AND activa = TRUE", (fecha_hoy,))
            suplencias_raw = cursor.fetchall() or []
            suplencias_map = {(r['docente_nombre'], r['materia']) if isinstance(r, dict) else (r[0], r[1]): (r['suplente_nombre'] if isinstance(r, dict) else r[2]) for r in suplencias_raw}

            cursor.execute("SELECT id, docente_nombre, asignatura_nombre, horario_inicio, estado_slug, incidencia_detalle, aula_actual FROM clases_historico WHERE fecha = %s", (fecha_hoy,))
            hist_raw = cursor.fetchall() or []
            historico_map = {
                (r['docente_nombre'] if isinstance(r, dict) else r[1],
                 r['asignatura_nombre'] if isinstance(r, dict) else r[2],
                 _norm_time(r['horario_inicio'] if isinstance(r, dict) else r[3])): (
                     r if isinstance(r, dict) else {'id': r[0], 'estado_slug': r[4], 'incidencia_detalle': r[5], 'aula_actual': r[6]}
                 )
                for r in hist_raw
            }

            cursor.execute("""
                SELECT h.id, h.docente, h.licenciatura, h.asignatura, h.horario, h.aula_asignada,
                       h.semestre, h.cuatrimestre, h.grupo, h.estado_slug, h.nota_reprogramacion,
                       h.fecha_reposicion, h.es_reposicion
                FROM horarios h
                WHERE h.docente IS NOT NULL AND TRIM(h.docente) != ''
            """)
            horarios_raw = cursor.fetchall()
            
            for h in (horarios_raw or []):
                dia_idx, inicio_m, fin_m = _parse_horario_minutos(h.get('horario', ''))
                if dia_idx == dia_hoy_idx and fin_m is not None:
                    h_ini_str = f"{inicio_m // 60:02d}:{inicio_m % 60:02d}:00"
                    h_fin_str = f"{fin_m // 60:02d}:{fin_m % 60:02d}:00"
                    sem_val = h.get('cuatrimestre') or h.get('semestre') or ''
                    aula_orig = h.get('aula_asignada', '')
                    aula_act = aula_orig

                    est_slug = 'finalizada'
                    est_label = 'Finalizada'
                    inc_det = 'Clase impartida sin incidencias'
                    es_incidencia = False

                    # 1. ¿Tuvo aula liberada hoy por cuestiones seleccionadas?
                    if aula_orig in aulas_lib_map:
                        motivo_lib = aulas_lib_map[aula_orig]
                        est_slug = 'aula_liberada'
                        est_label = 'Aula Liberada'
                        inc_det = f"Aula liberada: {motivo_lib or 'Cuestiones seleccionadas'}"
                        aula_act = f"Fuera de Aula ({aula_orig})"
                        es_incidencia = True

                    # 2. ¿Qué maestro mandó suplente hoy?
                    sup_key = (h['docente'], h['asignatura'])
                    if sup_key in suplencias_map:
                        suplente_nom = suplencias_map[sup_key]
                        est_slug = 'suplencia'
                        est_label = 'Con Suplente'
                        inc_det = f"Suplente: Prof. {suplente_nom}"
                        es_incidencia = True

                    # 3. ¿Reprogramó su clase para otro día?
                    nota_rep = h.get('nota_reprogramacion') or ''
                    if h.get('estado_slug') == 'pospuesta' or 'pospuesta' in str(nota_rep).lower() or 'reprogramada' in str(nota_rep).lower():
                        est_slug = 'reprogramada'
                        est_label = 'Reprogramada'
                        inc_det = nota_rep or 'Clase reprogramada para otro día'
                        es_incidencia = True
                    elif h.get('es_reposicion'):
                        est_slug = 'impartida'
                        est_label = 'Impartida (Reposición)'
                        inc_det = nota_rep or 'Clase de reposición'
                        es_incidencia = True

                    # Guardar en histórico si la clase ya venció o si tiene alguna incidencia (suplencia, liberación, reprogramación)
                    if mins_ahora > fin_m or es_incidencia:
                        _guardar_registro_historico(
                            cursor,
                            clase_id=h.get('id'),
                            fecha_obj=fecha_hoy,
                            docente_nombre=h['docente'],
                            licenciatura_codigo=h.get('licenciatura', ''),
                            semestre=sem_val,
                            grupo=h.get('grupo', ''),
                            asignatura_nombre=h['asignatura'],
                            horario_inicio_str=h_ini_str,
                            horario_fin_str=h_fin_str,
                            aula_original=aula_orig,
                            aula_actual=aula_act,
                            estado_slug=est_slug,
                            estado_label=est_label,
                            incidencia_detalle=inc_det,
                            historico_map=historico_map
                        )
            connection.commit()
    except Exception as e:
        print(f"Error evaluando cierre de clases vencidas: {e}")
    finally:
        connection.close()



@app.get("/api/clases-historico")
def obtener_clases_historico(
    rango_fecha: Optional[str] = "esta_semana",
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    licenciatura: Optional[str] = None,
    asignatura: Optional[str] = None,
    docente: Optional[str] = None,
    dia: Optional[str] = None,
    estado: Optional[str] = None,
    semestre: Optional[str] = None
):
    """Obtiene el historial de clases (getHistoricalClasses) filtrado por periodo, licenciatura, docente, estado, etc."""
    _evaluar_y_cerrar_clases_vencidas()

    ahora = datetime.now()
    fecha_hoy = ahora.date()

    if rango_fecha == "hoy":
        f_ini = f_fin = fecha_hoy
    elif rango_fecha == "esta_semana":
        f_ini = fecha_hoy - timedelta(days=fecha_hoy.weekday())
        f_fin = f_ini + timedelta(days=6)
    elif rango_fecha == "semana_pasada":
        lunes_pasado = fecha_hoy - timedelta(days=fecha_hoy.weekday() + 7)
        f_ini = lunes_pasado
        f_fin = lunes_pasado + timedelta(days=6)
    elif rango_fecha == "rango_personalizado" and fecha_inicio and fecha_fin:
        f_ini = datetime.strptime(fecha_inicio, "%Y-%m-%d").date()
        f_fin = datetime.strptime(fecha_fin, "%Y-%m-%d").date()
    else:
        f_ini = fecha_hoy - timedelta(days=fecha_hoy.weekday())
        f_fin = f_ini + timedelta(days=6)

    query = "SELECT * FROM clases_historico WHERE fecha BETWEEN %s AND %s"
    params = [f_ini, f_fin]

    if licenciatura:
        query += " AND licenciatura_codigo = %s"
        params.append(licenciatura)
    if asignatura:
        query += " AND asignatura_nombre LIKE %s"
        params.append(f"%{asignatura}%")
    if docente:
        query += " AND docente_nombre LIKE %s"
        params.append(f"%{docente}%")
    if dia:
        query += " AND dia_semana = %s"
        params.append(dia)
    if estado:
        query += " AND estado_slug = %s"
        params.append(estado)
    if semestre:
        query += " AND semestre = %s"
        params.append(semestre)

    query += " ORDER BY fecha DESC, horario_inicio ASC"

    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(query, params)
            filas = cursor.fetchall() or []

        resultado = []
        for r in filas:
            resultado.append({
                "id": r["id"],
                "clase_id": r.get("clase_id"),
                "fecha": str(r["fecha"]),
                "dia_semana": r["dia_semana"],
                "semana_numero": r["semana_numero"],
                "docente_nombre": r["docente_nombre"],
                "licenciatura_codigo": r["licenciatura_codigo"],
                "semestre": r["semestre"],
                "grupo": r["grupo"],
                "asignatura_nombre": r["asignatura_nombre"],
                "horario_inicio": _timedelta_to_str(r["horario_inicio"]),
                "horario_fin": _timedelta_to_str(r["horario_fin"]),
                "aula_original": r["aula_original"],
                "aula_actual": r["aula_actual"],
                "estado_slug": r["estado_slug"],
                "estado_label": r["estado_label"],
                "incidencia_detalle": r.get("incidencia_detalle") or "",
                "created_at": r["created_at"].strftime("%Y-%m-%d %H:%M:%S") if r.get("created_at") else ""
            })
        return resultado
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al consultar historial: {str(e)}")
    finally:
        connection.close()

@app.get("/api/bitacora")
def obtener_bitacora(
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    modulo: Optional[str] = None,
    usuario: Optional[str] = None
):
    connection = get_db_connection()
    try:
        query = "SELECT * FROM bitacora_cambios WHERE 1=1"
        params = []
        if fecha_inicio:
            query += " AND DATE(fecha_cambio) >= %s"
            params.append(fecha_inicio)
        if fecha_fin:
            query += " AND DATE(fecha_cambio) <= %s"
            params.append(fecha_fin)
        if modulo and modulo.lower() != "todos":
            query += " AND modulo_afectado = %s"
            params.append(modulo.lower())
        if usuario and usuario.lower() != "todos":
            query += " AND usuario = %s"
            params.append(usuario)
            
        query += " ORDER BY fecha_cambio DESC LIMIT 500"
        with connection.cursor() as cursor:
            cursor.execute(query, params)
            registros = cursor.fetchall()
        for r in (registros or []):
            if r.get('fecha_cambio'):
                r['fecha_cambio'] = r['fecha_cambio'].isoformat()
        return registros if registros else []
    except pymysql.Error as e:
        raise HTTPException(status_code=500, detail=f"Error al obtener bitacora: {str(e)}")
    finally:
        connection.close()

@app.get("/api/bitacora/filtros-disponibles")
def obtener_filtros_bitacora():
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT DISTINCT usuario FROM bitacora_cambios WHERE usuario IS NOT NULL AND usuario != ''")
            usuarios = [row['usuario'] for row in cursor.fetchall()]
        return {
            "usuarios": usuarios,
            "modulos": ["Aulas", "Docentes", "Horarios", "Suplencias"]
        }
    except pymysql.Error:
        return {"usuarios": [], "modulos": []}
    finally:
        connection.close()

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)


# ---------------------------------------------------------
# ENDPOINTS DE HISTORIAL (CIERRE DE PERIODO)
# ---------------------------------------------------------
class CierrePeriodo(BaseModel):
    nombre_ciclo: str
    modalidad: str

@app.post("/api/horarios/cierre-periodo")
def cierre_periodo(payload: CierrePeriodo):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            # Construir filtro de modalidad
            where_clause = "1=1"
            if payload.modalidad == "semestral":
                where_clause = "semestre IS NOT NULL AND semestre != ''"
            elif payload.modalidad == "cuatrimestral":
                where_clause = "cuatrimestre IS NOT NULL AND cuatrimestre != ''"

            # Verificar si hay horarios para archivar con esa modalidad
            cursor.execute(f"SELECT COUNT(*) as total FROM horarios WHERE {where_clause}")
            count = cursor.fetchone()
            if not count or count['total'] == 0:
                return {"message": "No hay horarios actuales para archivar en esta modalidad."}
                
            # Insertar en bloque
            cursor.execute(f"""
                INSERT INTO historial_periodos_cerrados
                (nombre_ciclo, tipo_periodo, dia, horario, asignatura, docente, aula_asignada, carrera, semestre, cuatrimestre, grupo)
                SELECT %s, %s, '', horario, asignatura, docente, aula_asignada, licenciatura, semestre, cuatrimestre, grupo
                FROM horarios
                WHERE {where_clause}
            """, (payload.nombre_ciclo, payload.modalidad))
            
            # Limpiar los horarios archivados de la vista principal
            cursor.execute(f"DELETE FROM horarios WHERE {where_clause}")
            
            connection.commit()
            return {"message": f"Periodo '{payload.nombre_ciclo}' cerrado y archivado correctamente en lote."}
    except Exception as e:
        connection.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


@app.get("/api/historial-periodos")
def obtener_historial_periodos():
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                SELECT 
                    *,
                    DATE_FORMAT(fecha_archivado, '%Y-%m-%d %H:%i:%s') as fecha_archivado_str
                FROM historial_periodos_cerrados
                ORDER BY fecha_archivado DESC, carrera, semestre, cuatrimestre
            """)
            historial = cursor.fetchall()
            return historial
    finally:
        connection.close()

class RenombrarHistorialRequest(BaseModel):
    nombre_ciclo_antiguo: str
    tipo_periodo: str
    nuevo_nombre_ciclo: str

@app.delete("/api/historial-periodos")
def eliminar_historial_periodo(nombre_ciclo: str, tipo_periodo: str):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                DELETE FROM historial_periodos_cerrados
                WHERE nombre_ciclo = %s AND tipo_periodo = %s
            """, (nombre_ciclo, tipo_periodo))
            connection.commit()
            if cursor.rowcount == 0:
                raise HTTPException(status_code=404, detail="No se encontraron registros para eliminar.")
            return {"message": "Historial eliminado correctamente"}
    except Exception as e:
        connection.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()

@app.put("/api/historial-periodos/renombrar")
def renombrar_historial_periodo(req: RenombrarHistorialRequest):
    connection = get_db_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                UPDATE historial_periodos_cerrados
                SET nombre_ciclo = %s
                WHERE nombre_ciclo = %s AND tipo_periodo = %s
            """, (req.nuevo_nombre_ciclo, req.nombre_ciclo_antiguo, req.tipo_periodo))
            connection.commit()
            if cursor.rowcount == 0:
                raise HTTPException(status_code=404, detail="No se encontraron registros para renombrar.")
            return {"message": "Historial renombrado correctamente"}
    except Exception as e:
        connection.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        connection.close()


from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook

@app.get("/api/historial-periodos/exportar")
def exportar_historial_periodos(ciclo: str = None, modalidad: str = None, fecha: str = None, semestre: str = None, cuatrimestre: str = None, carrera: str = None):
    connection = get_db_connection()
    try:
        with connection.cursor(pymysql.cursors.DictCursor) as cursor:
            query = "SELECT dia, horario, asignatura, docente, aula_asignada, carrera, semestre, cuatrimestre, grupo, nombre_ciclo as ciclo, tipo_periodo as modalidad, DATE_FORMAT(fecha_archivado, '%%Y-%%m-%%d') as fecha_archivado FROM historial_periodos_cerrados WHERE 1=1"
            params = []
            
            if ciclo and ciclo != 'undefined' and ciclo != 'null':
                query += " AND nombre_ciclo = %s"
                params.append(ciclo)
            if modalidad and modalidad != 'undefined' and modalidad != 'null':
                query += " AND tipo_periodo = %s"
                params.append(modalidad)
            if fecha and fecha != 'undefined' and fecha != 'null':
                query += " AND DATE(fecha_archivado) = %s"
                params.append(fecha)
            if semestre and semestre != 'undefined' and semestre != 'null':
                query += " AND semestre = %s"
                params.append(semestre)
            if cuatrimestre and cuatrimestre != 'undefined' and cuatrimestre != 'null':
                query += " AND cuatrimestre = %s"
                params.append(cuatrimestre)
            if carrera and carrera != 'undefined' and carrera != 'null':
                query += " AND carrera = %s"
                params.append(carrera)
            
            query += " ORDER BY fecha_archivado DESC, carrera, semestre, cuatrimestre"
            
            cursor.execute(query, params)
            registros = cursor.fetchall()
            
            if not registros:
                raise HTTPException(status_code=404, detail="No se encontraron registros.")
                
            wb = Workbook()
            ws = wb.active
            ws.title = "Historial"
            
            # Escribir encabezados
            headers = ['Día', 'Horario', 'Asignatura', 'Docente', 'Aula Asignada', 'Carrera', 'Semestre', 'Cuatrimestre', 'Grupo', 'Ciclo', 'Modalidad', 'Fecha Archivado']
            ws.append(headers)
            
            # Escribir datos
            for row in registros:
                ws.append([
                    row.get('dia', ''),
                    row.get('horario', ''),
                    row.get('asignatura', ''),
                    row.get('docente', ''),
                    row.get('aula_asignada', ''),
                    row.get('carrera', ''),
                    row.get('semestre', ''),
                    row.get('cuatrimestre', ''),
                    row.get('grupo', ''),
                    row.get('ciclo', ''),
                    row.get('modalidad', ''),
                    row.get('fecha_archivado', '')
                ])
                
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            safe_ciclo = ciclo if (ciclo and ciclo != 'undefined' and ciclo != 'null') else "Completo"
            headers_dict = {
                'Content-Disposition': f'attachment; filename="Historial_{safe_ciclo}.xlsx"'
            }
            
            return StreamingResponse(
                output, 
                headers=headers_dict, 
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    finally:
        connection.close()
